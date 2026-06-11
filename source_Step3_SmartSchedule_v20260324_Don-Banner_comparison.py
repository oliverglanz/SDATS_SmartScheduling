#!/usr/bin/env python
# coding: utf-8

# # Course Comparison: Banner vs SEMWorkingSpreadsheet
# 
# This notebook compares course data between:
# - **banner.csv**: CSV file with all course data
# - **New_SEM_course_updates_YYYYsemester.xlsx**: Excel file with courses split across multiple department tabs
# 
# 
# 
# 
# 
# **IMPORTANT!!!:** 
# The procedure below works only on computers (MacOS, Windows) where Excel is installed.
# 
# **Why?:** With the latest update we have now added Formulas in the DonSheet for the columns "Instructor Email {Instr Email}" and "Instructor ID {Instr ID}". These formulas will fill automatically once the instructor names in column "Instructor Name {Instr Name}" is being added. The automated filling of those fields is done from a hidden Sheet called "InstructorMap". Since the fields "Instructor Email {Instr Email}" and "Instructor ID {Instr ID}" contain Excel Formulas rather than values. It is important that the DonSheet has been run/opened on a computer so that the formulas have been executed before the comparision notebook is doing its work. In order to automate this "run/opening" of the file I have added this code:
# 
# ```python
# # Force Excel to calculate formulas and store cached values
# with xw.App(visible=False) as app:
#     wb = app.books.open(SEM_course_updates_INPUT)
#     wb.app.calculate()
#     wb.save()
#     wb.close()
# print("✔ Excel formulas recalculated and cached")
# ```
# 
# This will automatically - in the background - open the excel file and then close it again to make sure the Formulas have been used.
# 
# In addition, we need to make sure that the resulting values of the formulas are being compared and not the formula itself as it would result in detecting false "mismatches". I have therefore altered the following line:
# 
# ```python
# wb_working_xlsx = load_workbook(file_path, data_only=False)
# ```
# 
# into:
# 
# ```python
# wb_working_xlsx = load_workbook(file_path, data_only=True)
# ```
# 
# If one does not want to use this updated procedure and rather compare values. One can either remove:
# 
# ```python
# # Force Excel to calculate formulas and store cached values
# with xw.App(visible=False) as app:
#     wb = app.books.open(SEM_course_updates_INPUT)
#     wb.app.calculate()
#     wb.save()
#     wb.close()
# print("✔ Excel formulas recalculated and cached")
# ```
# 
# and change:
# ```python
# wb_working_xlsx = load_workbook(file_path, data_only=True)
# ```
# into
# ```python
# wb_working_xlsx = load_workbook(file_path, data_only=False)
# ```
# 
# Or use the previous version of this code with the name "source_Step3_SmartSchedule_v20251208_Don-Banner_comparison.ipynb".
# 

# In[ ]:


import os
import re
import glob
import numpy as np
from copy import copy
from collections import defaultdict
from IPython.display import display
import pandas as pd
import openpyxl
import xlwings as xw
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation


# In[2]:


# Making sure we are working within the same directory as the jupyter notebook (this will allow for poratbility of the project)
import os
## Verify we're in the right directory
print("Current directory:", os.getcwd())


# In[3]:


import pandas as pd

# Show all columns
pd.set_option('display.max_columns', None)

# Show all rows (optional — be careful if your BannerRecent is huge)
pd.set_option('display.max_rows', None)

# Don’t truncate long strings
pd.set_option('display.max_colwidth', None)

# Make the display wider in the notebook
pd.set_option('display.width', 0)


# ## Configuration

# In[4]:


# File Paths
import glob
## First Input File: Downloaded AU Course Schedule (supports CSV, XLS, XLSX)
BannerRecentFile = (
    glob.glob('Step3_input_downloaded_AU_BannerMostRecent/*.csv') +
    glob.glob('Step3_input_downloaded_AU_BannerMostRecent/*.xls') + 
    glob.glob('Step3_input_downloaded_AU_BannerMostRecent/*.xlsx')
    )
if not BannerRecentFile:
    raise FileNotFoundError("No CSV, XLS, or XLSX files found in Step3_input_downloaded_AU_BannerMostRecent directory")
BannerRecentFile_INPUT = BannerRecentFile[0]
print(f"Found Banner file: {BannerRecentFile_INPUT}")


### Second Input File: WorkingSpreadsheet
#SEM_course_updates_file = glob.glob('Step3_input_SEM_CourseUpdates_Spreadsheet/*.csv') + glob.glob('Step3_input_SEM_CourseUpdates_Spreadsheet/*.xls') + glob.glob('Step3_input_SEM_CourseUpdates_Spreadsheet/*.xlsx')
#if not SEM_course_updates_file:
#    raise FileNotFoundError("No CSV, XLS, or XLSX files found in Step3_input_SEM_CourseUpdates_Spreadsheet directory")
#SEM_course_updates_INPUT = SEM_course_updates_file[0]
#print(f"Found Banner file: {SEM_course_updates_INPUT}")

## Second Input File: WorkingSpreadsheet
SEM_course_updates_file = (
    glob.glob('x_SEM_CourseUpdate_LiveFolder_office365/*.csv') +
    glob.glob('x_SEM_CourseUpdate_LiveFolder_office365/*.xls') +
    glob.glob('x_SEM_CourseUpdate_LiveFolder_office365/*.xlsx')
    )

if not SEM_course_updates_file:
    raise FileNotFoundError(
        "No CSV, XLS, or XLSX files found in x_SEM_CourseUpdate_LiveFolder_office365 directory"
    )

SEM_course_updates_INPUT = SEM_course_updates_file[0]
print(f"Found SEM course update file: {SEM_course_updates_INPUT}")

# Force Excel to calculate formulas and store cached values
with xw.App(visible=False) as app:
    wb = app.books.open(SEM_course_updates_INPUT)
    wb.app.calculate()
    wb.save()
    wb.close()
print("✔ Excel formulas recalculated and cached")


DropDown_RETRIEVAL = '0000_BuildingFiles/source_DropDownMenus.xlsx'
print(f"Found Banner file: {DropDown_RETRIEVAL}")

## First Output File: Basic Calculated Spreadsheet File
Step3_OUTPUT_DiscrepencyReport_Color_full = 'Step3_output_DiscrepencyReport/ColorReport_SEM_course_updates.xlsx'
Step3_OUTPUT_DiscrepencyReport_Simple_full = 'Step3_output_DiscrepencyReport/SimpleReport_SEM_course_updates.xlsx'
Step3_OUTPUT_DiscrepencyReport_Simple_OTST_missingRows = 'Step3_output_DiscrepencyReport/SimpleReport_OTST_missingRows.xlsx'
Step3_OUTPUT_DiscrepencyReport_Simple_OTST_discrepencies = 'Step3_output_DiscrepencyReport/SimpleReport_OTST_discrepencies.xlsx'


# # Checking out the "DonSheet"

# In[5]:


import pandas as pd

xl = pd.ExcelFile(SEM_course_updates_INPUT)
print(xl.sheet_names)


# In[6]:


import pandas as pd
from openpyxl import load_workbook
import os

def make_unique_headers(headers):
    seen = {}
    new_headers = []

    for i, h in enumerate(headers):
        name = "" if h is None else str(h).strip()

        if name == "":
            name = f"Unnamed_{i+1}"

        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0

        new_headers.append(name)

    return new_headers


file_path = SEM_course_updates_INPUT
ext = os.path.splitext(file_path)[1].lower()

print("Loading working_df from:", file_path)

working_dfs = []
working_sheet_names = []
wb_working_xlsx = None

if ext in [".xlsx", ".xlsm"]:
    wb_working_xlsx = load_workbook(file_path, data_only=False)
    working_sheet_names = [
        s for s in wb_working_xlsx.sheetnames
        if not s.startswith("DropDown") and s != "InstructorMap"
    ]

    for sheet in working_sheet_names:
        ws = wb_working_xlsx[sheet]
        df = pd.DataFrame(ws.values)

        if df.shape[0] == 0:
            continue

        # Remove fully empty rows/columns first
        df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")

        if df.empty:
            continue

        # Make headers unique
        headers = make_unique_headers(df.iloc[0].tolist())

        # Apply headers and remove header row from data
        df = df.iloc[1:].copy()
        df.columns = headers
        df["SheetName"] = sheet

        working_dfs.append(df)

elif ext == ".xls":
    xl = pd.ExcelFile(file_path)
    working_sheet_names = [
        s for s in xl.sheet_names
        if not str(s).startswith("DropDown") and str(s) != "InstructorMap"
    ]

    for sheet in working_sheet_names:
        df = pd.read_excel(xl, sheet_name=sheet, dtype=str)

        if df.empty:
            continue

        df.columns = make_unique_headers(df.columns.tolist())
        df["SheetName"] = sheet
        working_dfs.append(df)

elif ext == ".csv":
    df = pd.read_csv(file_path, dtype=str)
    df.columns = make_unique_headers(df.columns.tolist())
    df["SheetName"] = "CSV_Input"
    working_sheet_names = ["CSV_Input"]
    working_dfs.append(df)

else:
    raise ValueError(f"Unsupported file type: {ext}")

if not working_dfs:
    raise ValueError("No usable sheets were loaded.")

# Concatenate sheets (if multiple)
working_df = pd.concat(working_dfs, ignore_index=True)
working_df.head()


# # Loading the BannerRecent Data

# ## Identical Cells in `Step3 code` (exception: `BannerRecent` [Step1] = `BannerRecentFile` [Step3])

# ### Loading Banner Data

# In[7]:


import pandas as pd
import re

def find_header_row(file_path, max_rows_to_check=20):
    """
    Automatically identify the header row in an Excel file.
    Reads everything as text to preserve exact cell values,
    but detects headers by checking for mostly non-numeric strings.
    """
    # Read file as text
    BannerRecent_raw = pd.read_excel(
        file_path,
        header=None,
        nrows=max_rows_to_check,
        dtype=str,
        na_filter=False
    )

    best_row = 0
    max_score = 0

    for idx in range(len(BannerRecent_raw)):
        row = BannerRecent_raw.iloc[idx].fillna('')

        # Define helper functions
        def looks_numeric(val):
            return bool(re.match(r"^\s*-?\d+(\.\d+)?\s*$", val.strip()))

        # Count non-empty cells
        non_empty_count = sum(val.strip() != '' for val in row)

        # Count values that look like text (not numeric)
        text_like_count = sum((val.strip() != '' and not looks_numeric(val)) for val in row)

        # Scoring logic:
        #   reward text-like cells, reward non-empty rows,
        #   penalize rows with mostly numeric values
        score = text_like_count * 2 + non_empty_count

        # Bonus for having a reasonable fill ratio
        if non_empty_count > len(row) * 0.3:
            score += 10

        if score > max_score:
            max_score = score
            best_row = idx

    return best_row


def load_dataframe(file_path, max_rows_to_check=20):
    """
    Load Excel file with auto header detection and exact text preservation.
    """
    header_row = find_header_row(file_path, max_rows_to_check)
    print(f"Header row identified at row index: {header_row}")

    BannerRecent = pd.read_excel(
        file_path,
        header=header_row,
        dtype=str,
        na_filter=False
    )
    return BannerRecent


# === Example usage ===
if __name__ == "__main__":
    file_path = BannerRecentFile_INPUT
    BannerRecent = load_dataframe(file_path)

    print(f"\nDataFrame shape: {BannerRecent.shape}")
    print(f"\nColumn names:")
    print(BannerRecent.columns.tolist())
    print(f"\nFirst 2 rows:")
    print(BannerRecent.head(2))
    print(f"\nData types:")
    print(BannerRecent.dtypes)


# ### Adapting Banner Input

# In[8]:


BannerRecent.head(2)


# In[9]:


import pandas as pd

def add_crn_sorted_column(BannerRecent):
    print("=== Starting add_crn_sorted_column ===")

    df = BannerRecent.copy()

    if 'CRN' not in df.columns:
        raise KeyError("CRN column not found in DataFrame")
    print("Found CRN column")

    crn_series = df['CRN'].astype(str).str.strip()
    print("Converted CRN to string")

    crn_numeric = pd.to_numeric(crn_series, errors='coerce')
    is_digitlike = crn_series.str.match(r'^\d+$')
    df['__crn_group'] = crn_series.where(~is_digitlike, crn_numeric.astype('Int64').astype(str))
    print("Created __crn_group helper column")

    df['__occurrence'] = df.groupby('__crn_group', sort=False).cumcount() + 1
    print("Created __occurrence helper column")

    crn_sorted = crn_series + '-' + df['__occurrence'].astype(str).str.zfill(2)
    empty_mask = crn_series == ''
    crn_sorted.loc[empty_mask] = ''
    print("Built CRN sorted {not in Banner} values")

    df.drop(columns=['__crn_group', '__occurrence'], inplace=True)
    print("Dropped helper columns")

    crn_pos = df.columns.get_loc('CRN')
    if 'CRN sorted {not in Banner}' in df.columns:
        df.drop(columns='CRN sorted {not in Banner}', inplace=True)
    df.insert(crn_pos + 1, 'CRN sorted {not in Banner}', crn_sorted)
    print("Inserted 'CRN sorted {not in Banner}' column successfully")

    print("=== Function complete ===")
    return df


# In[10]:


BannerRecent = add_crn_sorted_column(BannerRecent)
print(BannerRecent.columns.tolist())
print(BannerRecent.head())


# In[11]:


def normalize_empty_values(series):
    """
    Normalize a pandas Series by:
    - Converting 'nan', 'none', 'na', 'n/a' (case-insensitive) to np.nan
    - Converting empty strings or whitespace-only to np.nan
    - Stripping leading/trailing spaces
    - Preserving original casing otherwise
    """
    # First ensure string type and strip whitespace
    s = series.astype(str).str.strip()

    # Define values to replace
    null_values = ['', 'nan', 'NaN', 'NAN', 'none', 'None', 'NA', 'N/A', 'na', 'n/a']
    
    # Use mask with isin for cleaner replacement
    s = s.mask(s.isin(null_values), np.nan)

    return s


columns_to_process = [
    'CRN sorted {not in Banner}', 'CRN', 'Activity Date', 'Term Code', 'Subject', 'Crse Num',
    'Seq Crse Num', 'X Lst', 'Catalog Title', 'Section Title', 'Status',
    'Cat Crs', 'Sect Crs', 'Bill Hrs', 'Sect Sch Type', 'Preq Areas',
    'Inst Method', 'Link Ind', 'Integration Code', 'Link Conn', 'Crs Attr',
    'Comments', 'Ssasect Campus', 'Camp Restr CC', 'Major Restr',
    'Scacrse College', 'Scacrse Dept', 'Special Apvl', 'Max Enrl', 'Enrolled',
    'Waitlist Capacity', 'Instr ID', 'Instr Name', 'Instr Email',
    '% Responsibility', 'Primary Ind', 'Override', 'Part-Of-Term',
    'Soaterm Start Date', 'Soaterm End Date', 'Weeks', '1st Date Reg Opens',
    'Last Date Reg Opens', 'OL Range From Date', 'OL Range To Date',
    'OL Numb Units', 'OL Duration Code', 'Grade Mode', 'Gradable Ind',
    'Tuit Waiver', 'Meeting Type', 'SUN', 'MON', 'TUE', 'WED', 'THU', 'FRI',
    'SAT', 'Meet Start Date', 'Meet End Date', 'Meet Beg Time', 'Meet End Time',
    'Meet Bldg', 'Meet Room', 'Meet Override', 'Fee Term', 'Fee Level',
    'Fees Amt', 'Fee Type', 'Fee Ind', 'Detl Code', 'Rate Code', 'Cohort Code',
    'St Attr Code', 'Level SC', 'Camp SC', 'Coll SC', 'DegC SC', 'Prog SC',
    'FOS Type', 'FOS Code', 'Dept Code', 'Admit Term', 'Curr Rate', 'Curr Styp',
    'Curricula'
]

for col in columns_to_process:
    if col in BannerRecent.columns:
        BannerRecent[col] = normalize_empty_values(BannerRecent[col])
        print(f"{col}: {BannerRecent[col].dtype}")
    else:
        print(f"Warning: Column '{col}' not found in BannerRecent")


# In[12]:


BannerRecent.head()


# In[13]:


# ==============================================================================
# BANNER COLUMN RENAMING
# ==============================================================================

rename_dict = {
    'Crse Num': 'Course Number {Crse Num}',
    'Seq Crse Num': 'Course Section {Seq Crse Num}',
    'Cat Crs': 'Credit Category {Cat Crs}',
    'Sect Crs': 'Credits {Sect Crs}',
    'Preq Areas': 'Prerequisites {Preq Areas}',
    'Inst Method': 'Instruction Method {Inst Method}',
    'Camp Restr CC': 'Campus Restriction {Camp Restr CC}',
    'Scacrse Dept': 'SEM Department {Scacrse Dept}',
    'Special Apvl': 'Special Approval {Special Apvl}',
    'Max Enrl': 'Enrollment Cap {Max Enrl}',
    'Instr ID': 'Instructor ID {Instr ID}',
    'Instr Name': 'Instructor Name {Instr Name}',
    'Instr Email': 'Instructor Email {Instr Email}',
    'Meet Start Date': 'Course Start Date {Meet Start Date}',
    'Meet End Date': 'Course End Date {Meet End Date}',
    'Meet Beg Time': 'Course Beginning Time {Meet Beg Time}',
    'Meet End Time': 'Course Ending Time {Meet End Time}',
    'Meet Bldg': 'Building {Meet Bldg}',
    'Meet Room': 'Room {Meet Room}',
    'Meet Override': 'Course Override {Meet Override}',
    'Fees Amt': 'Fee Amount {Fees Amt}',
    'Detl Code': 'Fee Code {Detl Code}',
    'Level SC': 'Level {Level SC}',
    'Ssasect Campus': 'Campus {Ssasect Campus}',
    'Soaterm Start Date':'Semester Start Date {Soaterm Start Date}',
    'Soaterm End Date':'Semester End Date {Soaterm End Date}',
    'Cat Crs':'Credit Catalog {Cat Crs}'
}

BannerRecent.rename(columns=rename_dict, inplace=True)

BannerRecent.head(2)


# In[14]:


site_name_dict = {
'ACW':'Wellness Center [ACW]',	
'BH':'Bell Hall [BH]',	
'BUL':'Buller Hall [BUL]',	
'CSH':'Chan Shun Hall [CSH]',	
'GHA':'Griggs Hall A [GHA]',
'GHB':'Griggs Hall B [GHB]',	
'HORN':'Horn Museum [HORN]',	
'JGYM':'Johnson Gym [JGYM]',	
'JWL':'James White Library [JWL]',	
'NH':'Nethery Hall [NH]',	
'OCARG':'Univ Adventista del Plata [OCARG]',	
'OCBRA1':'Centro Univ Adv de Sao Paulo [OCBRA1]',	
'OCCABU':'Burman University [OCCABU]',	
'OCCALL':'Loma Linda University [OCCALL]',	
'OCCASC':'Southeastern California Conf [OCCASC]',	
'OCCHK':'Hong Kong Adventist College [OCCHK]',	
'OCFLAH':'Advent Health University [OCFLAH]',	
'OCFLFC':'Florida Conference of SDA [OCFLFC]',	
'OCFLFL':'Forest Lake SDA Church [OCFLFL]',	
'OCGBR':'Newbold College [OCGBR]',	
'OCMDND':'North American Division of SDA [OCMDND]',	
'OCMENE':'Northern New England Conf [OCMENE]',	
'OCNEUC':'Union College [OCNEUC]',	
'OCOKOC':'Oklahoma City Central [OCOKOC]',	
'OCPOL':'Polish Senior College Theo&Hum [OCPOL]',	
'OCROU':'Universitatea Adventus din Cer [OCROU]',	
'OCRUS':'Zaokski Theo Seminary [OCRUS]',	
'OCTHA':'Asia-Pacific Int Univ AIU [OCTHA]',	
'OCTWN':'Taiwan Adventist College [OCTWN]',	
'OCUKR':'Ukrainan Adv Center Higher ED [OCUKR]',	
'OCWANP':'North Pacific Union Conference [OCWANP]',	
'PMC':'Pioneer Memorial Church [PMC]',	
'SEM':'Seminary Building [SEM]'
}

def replace_site_prefix(value):
    """Replace site prefix before ':' using site_name_dict."""
    if not isinstance(value, str) or ':' not in value:
        return value  # leave non-string or missing values unchanged

    prefix, rest = value.split(':', 1)
    prefix = prefix.strip()
    rest = rest.strip()

    # Replace prefix if found in dict
    if prefix in site_name_dict:
        return f"{site_name_dict[prefix]}: {rest}"
    else:
        return value

# Apply to Meet Bldg column
# Replace short codes with full names
BannerRecent["Building {Meet Bldg}"] = BannerRecent["Building {Meet Bldg}"].replace(site_name_dict)

print("✅ 'Building {Meet Bldg}' values replaced successfully.")

print("✅ 'Building {Meet Bldg}' prefixes replaced using site_name_dict")
BannerRecent["Building {Meet Bldg}"].head(2)


# In[15]:


BannerRecent["Building {Meet Bldg}"].unique()


# In[16]:


BannerRecent.head(2)


# In[17]:


BannerRecent = pd.concat(
    [pd.DataFrame(columns=["Don {not in Banner}", "Mismatching {not in Banner}", "Dept/Prog Admin {not in Banner}", "Mona {not in Banner}", "Karen {not in Banner}", "DONE {not in Banner}", "Notes {not in Banner}", "Program {not in Banner}", "Crosslist Details {not in Banner}", "Schedule Type {not in Banner}", "Pre-work Start Date {not in Banner}", "Pre-work End Date {not in Banner}", "Post-work Start Date {not in Banner}", "Post-work End Date {not in Banner}", "load/contract {not in Banner}", "costs per credit {not in Banner}", "total costs {not in Banner}", "account to be charged {not in Banner}"]), BannerRecent],
    axis=1
)
BannerRecent.head(2)


# In[18]:


# Define the exact column order
desired_order = [
    "Don {not in Banner}", 
    "Mismatching {not in Banner}", 
    "Dept/Prog Admin {not in Banner}",
    "Mona {not in Banner}", 
    "Karen {not in Banner}",
    "DONE {not in Banner}", 
    "Notes {not in Banner}",
    "CRN sorted {not in Banner}",
    "CRN",
    "Subject",
    "Course Number {Crse Num}",
    "Course Section {Seq Crse Num}",
    "Program {not in Banner}",
    "Catalog Title",
    "Section Title",
    "Crosslist Details {not in Banner}", #added and not in banner
    "X Lst",
    "Campus {Ssasect Campus}",
    "Campus Restriction {Camp Restr CC}",   #Did 'Campus' get back in?? WHen, where, and how can we get it out
    "Schedule Type {not in Banner}",
    "Instruction Method {Inst Method}", #not working
    "Level {Level SC}",
    "Credit Catalog {Cat Crs}",
    "Credits {Sect Crs}",
    "Enrollment Cap {Max Enrl}",
    "Meeting Type",
    "Semester Start Date {Soaterm Start Date}",
    "Pre-work Start Date {not in Banner}",
    "Pre-work End Date {not in Banner}",
    "Course Start Date {Meet Start Date}",
    "Course End Date {Meet End Date}",
    "Post-work Start Date {not in Banner}",
    "Post-work End Date {not in Banner}",
    "Semester End Date {Soaterm End Date}",
    "Course Beginning Time {Meet Beg Time}",
    "Course Ending Time {Meet End Time}",
    "SUN",
    "MON",
    "TUE",
    "WED",
    "THU",
    "FRI",
    "SAT",
    "Room {Meet Room}",
    "Building {Meet Bldg}",
    "Instructor Name {Instr Name}",
    "Instructor Email {Instr Email}",
    "Instructor ID {Instr ID}",
    "% Responsibility",
    "load/contract {not in Banner}",
    "costs per credit {not in Banner}",
    "total costs {not in Banner}", 
    "account to be charged {not in Banner}",
    "Fee Amount {Fees Amt}",
    "Fee Code {Detl Code}",
    "Fee Type",
    "Fee Level",
    "Fee Ind",
    "Fee Term",
    "SEM Department {Scacrse Dept}",
    "Activity Date",
    "Status",
    "Bill Hrs",
    "Sect Sch Type",
    "Prerequisites {Preq Areas}",
    "Link Ind",
    "Integration Code",
    "Link Conn",
    "Crs Attr",
    "Comments",
    "Major Restr",
    "Scacrse College",
    "Enrolled",
    "Waitlist Capacity",
    "Primary Ind",
    "Override",
    "1st Date Reg Opens",
    "Last Date Reg Opens",
    "OL Range From Date",
    "OL Range To Date",
    "OL Numb Units",
    "OL Duration Code",
    "Grade Mode",
    "Gradable Ind",
    "Tuit Waiver",
    "Course Override {Meet Override}",
    "Lvl Res Ind",
    "Cmp Res Ind", 
    "Rate Code",
    "Cohort Code",
    "St Attr Code",
    "Coll SC",
    "DegC SC",
    "Prog SC",
    "FOS Type",
    "Special Approval {Special Apvl}",
    "Weeks",
    "Term Code",
    "Part-Of-Term",
    "Camp SC" #Is this the one tha appears as 'Campus' get back in?? WHen, where, and how can we get it out

]

# Reorder columns
BannerRecent = BannerRecent[[col for col in desired_order if col in BannerRecent.columns]]

# (Optional) Report any columns that were missing
missing_cols = [col for col in desired_order if col not in BannerRecent.columns]
if missing_cols:
    print("⚠️ The following columns were not found in BannerRecent:")
    for col in missing_cols:
        print("  -", col)
else:
    print("✅ All columns reordered successfully.")

# Preview the reordered DataFrame
BannerRecent.head(2)


# In[19]:


# Save an unfiltered copy for diagnostics
BannerRecent_original = BannerRecent.copy()

# ==============================================================================
# FILTER BANNERRENEWAL FOR SEMINARY COURSES (Scacrse College == "70")
# ==============================================================================

before_count = len(BannerRecent)

# Ensure Scacrse College exists
if "Scacrse College" not in BannerRecent.columns:
    raise RuntimeError(
        "Column 'Scacrse College' is missing in BannerRecent. "
        "Cannot filter for Seminary courses."
    )

# Apply the filter
BannerRecent = BannerRecent[BannerRecent["Scacrse College"] == "70"].copy()
after_count = len(BannerRecent)

print(f"Filtered BannerRecent rows: {after_count} out of {before_count}")

# If NO rows remain, abort the whole process
if after_count == 0:
    print("Values found in 'Scacrse College' in the original Banner file:")
    print(BannerRecent_original["Scacrse College"].value_counts())

    raise SystemExit(
        "🚫 A comparison report cannot be created because the Banner report does "
        "not contain seminary-relevant course information.\n\n"
        "Specifically: Banner has *no* rows where 'Scacrse College' == '70'.\n"
        "Please verify that the Banner export includes SEM course data."
    )


# ### Minor Testing

# In[20]:


BannerRecentDummy = BannerRecent
DUMMY = BannerRecentDummy[(
    (
        #(BannerRecentDummy["Scacrse College"] == 70) 
        #& 
        (BannerRecentDummy['CRN'] == "4730")
        & 
        (BannerRecentDummy['Course Section {Seq Crse Num}'] == 'XXX')
    )
)].copy()
DUMMY.head()


# In[21]:


BannerRecentDummy = BannerRecent
DUMMY = BannerRecentDummy[(
    (
        #(BannerRecentDummy["Scacrse College"] == 70) 
        #& 
        (BannerRecentDummy['Subject'] == "OTST")
        #& 
        #(BannerRecentDummy['Course Section {Seq Crse Num}'] == 'XXX')
    )
)].copy()
DUMMY.head()


# In[22]:


BannerRecent.head()


# # Creating DropDown Menu (Identical to Step1!)

# In[23]:


import pandas as pd
import pprint  # pretty-print

# --- Step 1: Load the Excel file ---
# Replace 'your_file.xlsx' with your filename
DropDownMenu_dict = pd.read_excel(DropDown_RETRIEVAL, dtype=str)  # forces all columns to be read as strings
DropDownMenu_dict.head()


# In[24]:


# --- Build Instructor mapping (Name -> Email, ID) BEFORE unique lists are created ---
instr_name_col  = "Instructor Name {Instr Name}"
instr_email_col = "Instructor Email {Instr Email}"
instr_id_col    = "Instructor ID {Instr ID}"

InstructorMap_df = None
if all(c in DropDownMenu_dict.columns for c in [instr_name_col, instr_email_col, instr_id_col]):
    InstructorMap_df = (
        DropDownMenu_dict[[instr_name_col, instr_email_col, instr_id_col]]
        .copy()
    )

    # normalize
    for c in [instr_name_col, instr_email_col, instr_id_col]:
        InstructorMap_df[c] = InstructorMap_df[c].fillna("").astype(str).str.strip()

    # keep only rows that actually have a name
    InstructorMap_df = InstructorMap_df[InstructorMap_df[instr_name_col] != ""]

    # dedupe by name (keep first)
    InstructorMap_df = InstructorMap_df.drop_duplicates(subset=[instr_name_col], keep="first")



print("=== DEBUG: InstructorMap_df ===")
print("InstructorMap_df is None:", InstructorMap_df is None)
if InstructorMap_df is not None:
    print("InstructorMap_df rows:", len(InstructorMap_df))
    print(InstructorMap_df.head(10))
    # quick sanity check for empties
    empties = (InstructorMap_df[instr_email_col] == "").sum()
    print("Empty emails:", empties)
    empties_id = (InstructorMap_df[instr_id_col] == "").sum()
    print("Empty IDs:", empties_id)


# In[25]:


# --- Step 2: Create dictionaries ---
# Each column becomes a key; values are the unique entries (as strings)
DropDownMenu_dict = {
    col: DropDownMenu_dict[col].dropna().astype(str).unique().tolist()
    for col in DropDownMenu_dict.columns
}

# --- Step 3: (Optional) Display the result nicely ---
for key, values in DropDownMenu_dict.items():
    print(f"{key}: {values[:5]}{'...' if len(values) > 5 else ''}")


# # Preprocessing of Banner Data

# In[26]:


BannerRecent.columns.tolist()


# In[ ]:


# =============================================================================
# PREPROCESS BannerRecent for comparison
# - keep original BannerRecent unchanged
# - collapse duplicate Banner rows based on a POSITIVE list of relevant columns
# - rebuild "CRN sorted {not in Banner}" after deduplication
# =============================================================================

import pandas as pd
import numpy as np

# ---------------------------------------------------------------------
# 1) Columns that DO define meaningful Banner row differences
# ---------------------------------------------------------------------
BANNER_RELEVANT_FOR_ROW_MULTIPLICATION = [
    'CRN',
    'Subject',
    'Course Number {Crse Num}',
    'Course Section {Seq Crse Num}',
    'Section Title',
    'Campus {Ssasect Campus}',
    'Campus Restriction {Camp Restr CC}',
    'Instruction Method {Inst Method}',
    'Credits {Sect Crs}',
    'Enrollment Cap {Max Enrl}',
    'Meeting Type',
    'Course Start Date {Meet Start Date}',
    'Course End Date {Meet End Date}',
    'Course Beginning Time {Meet Beg Time}',
    'Course Ending Time {Meet End Time}',
    'SUN',
    'MON',
    'TUE',
    'WED',
    'THU',
    'FRI',
    'Room {Meet Room}',
    'Building {Meet Bldg}',
    'Instructor Name {Instr Name}',
    'Instructor Email {Instr Email}',
    'Instructor ID {Instr ID}',
    '% Responsibility',
]

# ---------------------------------------------------------------------
# 2) Helper functions
# ---------------------------------------------------------------------
def normalize_empty_value(val):
    if pd.isna(val):
        return None
    if isinstance(val, str) and val.strip() == "":
        return None
    return val

def normalize_text_for_dedupe(val):
    v = normalize_empty_value(val)
    if v is None:
        return ""
    return str(v).strip()

def normalize_key_series(series):
    s = series.astype(str).str.strip()
    null_like = {"", "nan", "NaN", "NONE", "None", "Null", "null"}
    s = s.mask(s.isin(null_like), np.nan)
    return s

def add_crn_sorted_column(df, crn_col="CRN", out_col="CRN sorted {not in Banner}", sort_cols=None):
    """
    Rebuild values like:
        5832-01, 5832-02, ...
    within each CRN after deduplication.
    """
    df = df.copy()

    if sort_cols is None:
        sort_cols = []

    usable_sort_cols = [c for c in sort_cols if c in df.columns and c != crn_col]

    if usable_sort_cols:
        df = df.sort_values(
            by=[crn_col] + usable_sort_cols,
            kind="stable",
            na_position="last"
        ).copy()
    else:
        df = df.sort_values(
            by=[crn_col],
            kind="stable",
            na_position="last"
        ).copy()

    seq = df.groupby(crn_col, dropna=False).cumcount().add(1)
    df[out_col] = (
        df[crn_col].astype(str).str.strip()
        + "-"
        + seq.astype(str).str.zfill(2)
    )

    return df

# ---------------------------------------------------------------------
# 3) Build Banner comparison dataframe
# ---------------------------------------------------------------------
banner_original_df = BannerRecent.copy()
banner_compare_df = banner_original_df.copy()

if "CRN" not in banner_compare_df.columns:
    raise KeyError("'CRN' must exist in BannerRecent before preprocessing.")

banner_compare_df["CRN"] = normalize_key_series(banner_compare_df["CRN"])
banner_compare_df = banner_compare_df[banner_compare_df["CRN"].notna()].copy()

# ---------------------------------------------------------------------
# 4) Define deduplication subset from POSITIVE list
# ---------------------------------------------------------------------
dedupe_subset = [
    c for c in BANNER_RELEVANT_FOR_ROW_MULTIPLICATION
    if c in banner_compare_df.columns
]

missing_relevant_cols = [
    c for c in BANNER_RELEVANT_FOR_ROW_MULTIPLICATION
    if c not in banner_compare_df.columns
]

if missing_relevant_cols:
    print(f"⚠️ These relevant columns were not found in BannerRecent: {missing_relevant_cols}")

if not dedupe_subset:
    raise ValueError("No valid columns found for Banner deduplication.")

print(f"🔎 Banner dedupe subset ({len(dedupe_subset)} columns):")
print(dedupe_subset)

# ---------------------------------------------------------------------
# 5) Normalize dedupe subset so blank / NaN differences do not block collapse
# ---------------------------------------------------------------------
banner_dedupe_view = banner_compare_df[dedupe_subset].copy()

for col in banner_dedupe_view.columns:
    banner_dedupe_view[col] = banner_dedupe_view[col].map(normalize_text_for_dedupe)

# ---------------------------------------------------------------------
# 6) Drop duplicate Banner rows for comparison
# ---------------------------------------------------------------------
keep_mask = ~banner_dedupe_view.duplicated(keep="first")
banner_compare_df = banner_compare_df.loc[keep_mask].copy()

print(f"📉 Banner rows before preprocessing: {len(banner_original_df)}")
print(f"📉 Banner rows after comparison-deduplication: {len(banner_compare_df)}")
print(f"📉 Collapsed duplicate rows: {len(banner_original_df) - len(banner_compare_df)}")

# ---------------------------------------------------------------------
# 7) Rebuild CRN sorted {not in Banner}
# ---------------------------------------------------------------------
CRN_SORT_ORDER = [
    "Meeting Type",
    "Course Start Date {Meet Start Date}",
    "Course End Date {Meet End Date}",
    "Course Beginning Time {Meet Beg Time}",
    "Course Ending Time {Meet End Time}",
    "Instructor ID {Instr ID}",
    "% Responsibility",
    "Room {Meet Room}",
    "Building {Meet Bldg}",
    "SUN",
    "MON",
    "TUE",
    "WED",
    "THU",
    "FRI",
]

banner_compare_df = add_crn_sorted_column(
    banner_compare_df,
    crn_col="CRN",
    out_col="CRN sorted {not in Banner}",
    sort_cols=CRN_SORT_ORDER
)

# ---------------------------------------------------------------------
# 8) Optional diagnostic table: which CRNs shrank?
# ---------------------------------------------------------------------
before_counts = (
    banner_original_df.assign(CRN=normalize_key_series(banner_original_df["CRN"]))
    .dropna(subset=["CRN"])
    .groupby("CRN")
    .size()
    .rename("RowsBefore")
)

after_counts = (
    banner_compare_df.groupby("CRN")
    .size()
    .rename("RowsAfter")
)

banner_preprocess_summary = (
    pd.concat([before_counts, after_counts], axis=1)
    .fillna(0)
    .astype(int)
    .reset_index()
)

banner_preprocess_summary["RowsCollapsed"] = (
    banner_preprocess_summary["RowsBefore"] - banner_preprocess_summary["RowsAfter"]
)

banner_preprocess_summary = banner_preprocess_summary.sort_values(
    ["RowsCollapsed", "CRN"],
    ascending=[False, True]
).reset_index(drop=True)

print("✅ Banner preprocessing complete.")
#display(banner_preprocess_summary.head(20))
print(banner_preprocess_summary.head(20))

# ---------------------------------------------------------------------
# 9) From here on, use banner_compare_df in the comparison code
#    instead of BannerRecent / banner_df
# ---------------------------------------------------------------------


# In[28]:


crn_test = "3627"

tmp = banner_compare_df[banner_compare_df["CRN"].astype(str) == crn_test].copy()
print("Rows remaining for CRN", crn_test, ":", len(tmp))

varying_cols = []
for col in tmp.columns:
    vals = tmp[col].map(lambda x: "" if pd.isna(x) else str(x).strip()).unique()
    vals = [v for v in vals if v != ""]
    if len(vals) > 1:
        varying_cols.append((col, vals))

for col, vals in varying_cols:
    print(f"\nCOLUMN: {col}")
    print(vals)


# In[29]:


banner_compare_df.head()


# In[30]:


banner_compare_df[banner_compare_df["CRN"] == "3621"]


# # Comparison Process

# In[31]:


# =============================================================================
# STEP 3 — CELL #1
# Unified working loader (CSV/XLS/XLSX) + CRN-based comparison → discrepancy master
# =============================================================================
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import warnings

# ---- Config -----------------------------------------------------------------
BASE_KEY = "CRN"

# Keep this minimal so we do NOT consume important comparison columns.
ROW_MATCH_COLS_CANDIDATES = [
    "Instructor Email {Instr Email}",
]

ARR_EQUIVALENTS = {"Arr.", "Arr", "arr.", "arr"}

Step3_OUTPUT_SummaryReport = os.path.join(
    "Step3_output_DiscrepencyReport",
    "Step3_OUTPUT_SummaryReport.xlsx"
)
Step3_OUTPUT_DetailedValueDiscrepancyReport = os.path.join(
    "Step3_output_DiscrepencyReport",
    "Step3_OUTPUT_DetailedValueDiscrepancyReport.xlsx"
)

os.makedirs("Step3_output_DiscrepencyReport", exist_ok=True)

# ---- Helpers ----------------------------------------------------------------
def normalize_empty_value(val, col_name=None):
    """
    Normalize scalar values for comparison.
    Allows column-specific equivalences.
    """
    if pd.isna(val):
        return None

    if isinstance(val, str):
        val = val.strip()
        if val == "":
            return None

    # Treat Arr / Arr. variants as empty for these columns
    if col_name in {
        "Building {Meet Bldg}",
        "Room {Meet Room}",
        "Course Beginning Time {Meet Beg Time}",
        "Course Ending Time {Meet End Time}",
    }:
        if isinstance(val, str) and val in ARR_EQUIVALENTS:
            return None

    # Schedule Type mapping
    if col_name == "Schedule Type {not in Banner}":
        if isinstance(val, str):
            v = val.strip().lower()
            if v in {"blended learning", "bl"}:
                return "BL"

    return val

def values_are_equal(val1, val2, col_name=None):
    """
    Compare two values with lenient type handling
    plus column-specific equivalence rules.
    """
    v1 = normalize_empty_value(val1, col_name=col_name)
    v2 = normalize_empty_value(val2, col_name=col_name)

    if v1 is None and v2 is None:
        return True
    if v1 is None or v2 is None:
        return False

    try:
        return float(v1) == float(v2)
    except Exception:
        return str(v1).strip() == str(v2).strip()

def normalize_key_series(series):
    s = series.astype(str).str.strip()
    null_like = {"", "nan", "NaN", "NONE", "None", "Null", "null"}
    s = s.mask(s.isin(null_like), np.nan)
    return s

def make_unique_headers(headers):
    seen = {}
    out = []

    for i, h in enumerate(headers):
        name = "" if h is None else str(h).strip()

        if name == "":
            name = f"Unnamed_{i+1}"

        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0

        out.append(name)

    return out

def clean_loaded_sheet(df):
    if df.shape[0] == 0:
        return None

    df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
    if df.empty:
        return None

    headers = make_unique_headers(df.iloc[0].tolist())
    df = df.iloc[1:].copy()
    df.columns = headers
    df = df.dropna(axis=0, how="all")

    if df.empty:
        return None

    return df

def normalize_for_signature(val):
    v = normalize_empty_value(val)
    if v is None:
        return ""
    return str(v).strip()

def build_minimal_row_signature(df, candidate_cols):
    """
    Build a MINIMAL helper signature.
    This is only for pairing repeated rows inside a CRN.
    It should not consume major comparison columns.
    """
    usable_cols = [c for c in candidate_cols if c in df.columns]

    if not usable_cols:
        return pd.Series(["ROW"] * len(df), index=df.index)

    sig = df[usable_cols].apply(
        lambda row: " | ".join(normalize_for_signature(v) for v in row),
        axis=1
    )

    sig = sig.replace("", "ROW")
    return sig

def add_pairing_columns(df, base_key, candidate_cols):
    """
    Pairing logic for cross-system comparison:
    1. use CRN as base key
    2. build a minimal helper signature
    3. sort rows stably within each CRN + helper signature
    4. assign sequence numbers within each (CRN, helper signature)
    5. create _RowMatchKey = CRN || helper signature || sequence
    """
    df = df.copy()

    df["_RowSignature"] = build_minimal_row_signature(df, candidate_cols)
    df["_OriginalOrder"] = np.arange(len(df))

    sort_cols = [base_key, "_RowSignature"]

    if "CRN sorted {not in Banner}" in df.columns:
        sort_cols.append("CRN sorted {not in Banner}")

    sort_cols.append("_OriginalOrder")

    df = df.sort_values(sort_cols, kind="stable", na_position="last").copy()

    df["_RowSeqWithinSignature"] = (
        df.groupby([base_key, "_RowSignature"], dropna=False)
          .cumcount()
          .add(1)
    )

    df["_RowMatchKey"] = (
        df[base_key].astype(str).str.strip()
        + " || "
        + df["_RowSignature"].astype(str)
        + " || #"
        + df["_RowSeqWithinSignature"].astype(str)
    )

    return df

# ---- 1) Load SEM_course_updates_INPUT into one unified DataFrame ------------
working_ext = os.path.splitext(SEM_course_updates_INPUT)[1].lower()
all_working_rows = []
working_sheet_names = []
wb_working_xlsx = None

print(f"📥 Loading WorkingSpreadsheet: {SEM_course_updates_INPUT}")

if working_ext in [".xlsx", ".xlsm"]:
    wb_working_xlsx = load_workbook(SEM_course_updates_INPUT, data_only=True)

    working_sheet_names = [
        s for s in wb_working_xlsx.sheetnames
        if not s.startswith("DropDown") and s != "InstructorMap"
    ]

    for sheet in working_sheet_names:
        ws = wb_working_xlsx[sheet]
        raw_df = pd.DataFrame(ws.values)
        df = clean_loaded_sheet(raw_df)

        if df is None:
            continue

        df["SheetName"] = sheet
        all_working_rows.append(df)

elif working_ext == ".xls":
    xls = pd.ExcelFile(SEM_course_updates_INPUT)

    working_sheet_names = [
        s for s in xls.sheet_names
        if not str(s).startswith("DropDown") and str(s) != "InstructorMap"
    ]

    for sheet in working_sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet, dtype=str)

        if df.empty:
            continue

        df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
        if df.empty:
            continue

        df.columns = make_unique_headers(df.columns.tolist())
        df["SheetName"] = sheet
        all_working_rows.append(df)

elif working_ext == ".csv":
    df = pd.read_csv(SEM_course_updates_INPUT, dtype=str)

    if not df.empty:
        df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
        df.columns = make_unique_headers(df.columns.tolist())
        df["SheetName"] = "CSV_Input"
        working_sheet_names = ["CSV_Input"]
        all_working_rows.append(df)

else:
    raise ValueError(f"Unsupported WorkingSpreadsheet file type: {working_ext}")

if not all_working_rows:
    raise RuntimeError("No rows found in SEM_course_updates_INPUT.")

working_df = pd.concat(all_working_rows, ignore_index=True)
print(f"✅ Working loaded: {len(working_df)} rows across {len(working_sheet_names)} sheet(s)")

# ---- 2) Use your already-prepped BannerRecent directly ----------------------
banner_df = banner_compare_df.copy()
banner_df.columns = make_unique_headers(banner_df.columns.tolist())

# ---- 3) Key normalization ---------------------------------------------------
if BASE_KEY not in working_df.columns or BASE_KEY not in banner_df.columns:
    raise KeyError(f"'{BASE_KEY}' must exist in both Working and Banner data.")

working_df[BASE_KEY] = normalize_key_series(working_df[BASE_KEY])
banner_df[BASE_KEY]  = normalize_key_series(banner_df[BASE_KEY])

working_df = working_df[working_df[BASE_KEY].notna()].copy()
banner_df  = banner_df[banner_df[BASE_KEY].notna()].copy()

# ---- 4) Decide comparable columns ------------------------------------------
COMPARE_COLS_MAP = {
    "Subject": "Subject",
    "Course Number {Crse Num}": "Course Number {Crse Num}",
    "Course Section {Seq Crse Num}": "Course Section {Seq Crse Num}",
    "Program {not in Banner}": "Program {not in Banner}",
    "Catalog Title": "Catalog Title",
    "Section Title": "Section Title",
    "Schedule Type {not in Banner}": "Sect Sch Type",
    "Instruction Method {Inst Method}": "Instruction Method {Inst Method}",
    "Level {Level SC}": "Level {Level SC}",
    "Credits {Sect Crs}": "Credits {Sect Crs}",
    "Enrollment Cap {Max Enrl}": "Enrollment Cap {Max Enrl}",
    "Pre-work Start Date {not in Banner}": "Pre-work Start Date {not in Banner}",
    "Pre-work End Date {not in Banner}": "Pre-work End Date {not in Banner}",
    "Course Start Date {Meet Start Date}": "Course Start Date {Meet Start Date}",
    "Course End Date {Meet End Date}": "Course End Date {Meet End Date}",
    "Post-work Start Date {not in Banner}": "Post-work Start Date {not in Banner}",
    "Post-work End Date {not in Banner}": "Post-work End Date {not in Banner}",
    "Meeting Type": "Meeting Type",
    "Course Beginning Time {Meet Beg Time}": "Course Beginning Time {Meet Beg Time}",
    "Course Ending Time {Meet End Time}": "Course Ending Time {Meet End Time}",
    "SUN": "SUN",
    "MON": "MON",
    "TUE": "TUE",
    "WED": "WED",
    "THU": "THU",
    "FRI": "FRI",
    "Room {Meet Room}": "Room {Meet Room}",
    "Building {Meet Bldg}": "Building {Meet Bldg}",
    "Instructor Name {Instr Name}": "Instructor Name {Instr Name}",
    "Instructor Email {Instr Email}": "Instructor Email {Instr Email}",
    "Instructor ID {Instr ID}": "Instructor ID {Instr ID}",
    "% Responsibility": "% Responsibility",
    "load/contract {not in Banner}": "load/contract {not in Banner}",
    "costs per credit {not in Banner}": "costs per credit {not in Banner}",
    "total costs {not in Banner}": "total costs {not in Banner}",
    "account to be charged {not in Banner}": "account to be charged {not in Banner}",
    "Weeks": "Weeks",
    "Term Code": "Term Code",
    "Part-Of-Term": "Part-Of-Term",
    "Camp SC": "Camp SC",
}

compare_pairs = [
    (w_col, b_col)
    for w_col, b_col in COMPARE_COLS_MAP.items()
    if w_col in working_df.columns and b_col in banner_df.columns
]

missing_from_working = [
    w_col for w_col, b_col in COMPARE_COLS_MAP.items()
    if w_col not in working_df.columns
]
missing_from_banner = [
    b_col for w_col, b_col in COMPARE_COLS_MAP.items()
    if b_col not in banner_df.columns
]

if missing_from_working:
    print(f"⚠️ Missing from Working: {missing_from_working}")

if missing_from_banner:
    print(f"⚠️ Missing from Banner: {missing_from_banner}")

print(f"🔎 Comparable mapped columns: {len(compare_pairs)}")

# ---- 5) Build pairing key ---------------------------------------------------
working_df = add_pairing_columns(working_df, BASE_KEY, ROW_MATCH_COLS_CANDIDATES)
banner_df  = add_pairing_columns(banner_df, BASE_KEY, ROW_MATCH_COLS_CANDIDATES)

print("🔑 Pairing key built from CRN + minimal helper signature + stable sequence.")

# ---- 6) True missing-course detection by CRN --------------------------------
working_crns = set(working_df[BASE_KEY].dropna())
banner_crns  = set(banner_df[BASE_KEY].dropna())

missing_in_banner  = working_crns - banner_crns
missing_in_working = banner_crns - working_crns

print(f"⚠️ CRNs missing in Banner (present only in Working): {len(missing_in_banner)}")
print(f"⚠️ CRNs missing in DonSheet (present only in Banner): {len(missing_in_working)}")

# ---- 7) Shared CRNs only ----------------------------------------------------
shared_crns = sorted(working_crns & banner_crns)
working_shared = working_df[working_df[BASE_KEY].isin(shared_crns)].copy()
banner_shared  = banner_df[banner_df[BASE_KEY].isin(shared_crns)].copy()

# ---- 8) Row-level presence mismatch within shared CRNs ----------------------
working_row_keys = set(working_shared["_RowMatchKey"])
banner_row_keys  = set(banner_shared["_RowMatchKey"])

row_missing_in_banner  = working_row_keys - banner_row_keys
row_missing_in_working = banner_row_keys - working_row_keys

print(f"⚠️ Row variants missing in Banner within shared CRNs: {len(row_missing_in_banner)}")
print(f"⚠️ Row variants missing in DonSheet within shared CRNs: {len(row_missing_in_working)}")

# ---- 9) Merge only matched row variants ------------------------------------
working_merge_cols = [BASE_KEY, "_RowMatchKey", "SheetName"] + [w for w, b in compare_pairs]
banner_merge_cols  = [BASE_KEY, "_RowMatchKey"] + [b for w, b in compare_pairs]

merged = working_shared[working_merge_cols].merge(
    banner_shared[banner_merge_cols],
    on=[BASE_KEY, "_RowMatchKey"],
    how='inner',
    suffixes=('_Working', '_Banner')
)

records = []

# ---- 10) Entire CRN missing in Banner --------------------------------------
for crn in sorted(missing_in_banner):
    w_rows = working_df.loc[working_df[BASE_KEY] == crn]
    sheet_guess = w_rows["SheetName"].iloc[0] if len(w_rows) else "(OnlyInWorking)"
    crn_sorted_val = (
        w_rows["CRN sorted {not in Banner}"].iloc[0]
        if "CRN sorted {not in Banner}" in w_rows.columns and len(w_rows)
        else None
    )
    subject_val = w_rows["Subject"].iloc[0] if "Subject" in w_rows.columns and len(w_rows) else None
    course_num_val = w_rows["Course Number {Crse Num}"].iloc[0] if "Course Number {Crse Num}" in w_rows.columns and len(w_rows) else None
    section_val = w_rows["Course Section {Seq Crse Num}"].iloc[0] if "Course Section {Seq Crse Num}" in w_rows.columns and len(w_rows) else None

    records.append({
        'SheetName': sheet_guess,
        'CRN sorted {not in Banner}': crn_sorted_val,
        'CRN': crn,
        'Subject': subject_val,
        'Course Number {Crse Num}': course_num_val,
        'Course Section {Seq Crse Num}': section_val,
        'RowMatchKey': None,
        'Column': '(entire CRN)',
        'DonSheet Value': '(exists in Working only)',
        'Banner Value': '(missing CRN)',
        'Type': 'CRNMissingInBanner'
    })

# ---- 11) Entire CRN missing in DonSheet ------------------------------------
for crn in sorted(missing_in_working):
    b_rows = banner_df.loc[banner_df[BASE_KEY] == crn]
    sheet_guess = str(b_rows['Subject'].iloc[0]) if 'Subject' in b_rows.columns and len(b_rows) else '(OnlyInBanner)'
    crn_sorted_val = (
        b_rows["CRN sorted {not in Banner}"].iloc[0]
        if "CRN sorted {not in Banner}" in b_rows.columns and len(b_rows)
        else None
    )
    subject_val = b_rows["Subject"].iloc[0] if "Subject" in b_rows.columns and len(b_rows) else None
    course_num_val = b_rows["Course Number {Crse Num}"].iloc[0] if "Course Number {Crse Num}" in b_rows.columns and len(b_rows) else None
    section_val = b_rows["Course Section {Seq Crse Num}"].iloc[0] if "Course Section {Seq Crse Num}" in b_rows.columns and len(b_rows) else None

    records.append({
        'SheetName': sheet_guess,
        'CRN sorted {not in Banner}': crn_sorted_val,
        'CRN': crn,
        'Subject': subject_val,
        'Course Number {Crse Num}': course_num_val,
        'Course Section {Seq Crse Num}': section_val,
        'RowMatchKey': None,
        'Column': '(entire CRN)',
        'DonSheet Value': '(missing CRN)',
        'Banner Value': '(exists in Banner only)',
        'Type': 'CRNMissingInDonSheet'
    })

# ---- 12) Row variant missing in Banner -------------------------------------
for row_key in sorted(row_missing_in_banner):
    w_rows = working_shared.loc[working_shared["_RowMatchKey"] == row_key]
    if w_rows.empty:
        continue
    r0 = w_rows.iloc[0]

    records.append({
        'SheetName': r0.get('SheetName'),
        'CRN sorted {not in Banner}': r0.get('CRN sorted {not in Banner}'),
        'CRN': r0.get('CRN'),
        'Subject': r0.get('Subject'),
        'Course Number {Crse Num}': r0.get('Course Number {Crse Num}'),
        'Course Section {Seq Crse Num}': r0.get('Course Section {Seq Crse Num}'),
        'RowMatchKey': row_key,
        'Column': '(row variant)',
        'DonSheet Value': '(exists in Working only)',
        'Banner Value': '(missing row variant)',
        'Type': 'RowVariantMissingInBanner'
    })

# ---- 13) Row variant missing in DonSheet -----------------------------------
for row_key in sorted(row_missing_in_working):
    b_rows = banner_shared.loc[banner_shared["_RowMatchKey"] == row_key]
    if b_rows.empty:
        continue
    r0 = b_rows.iloc[0]
    sheet_guess = r0.get('Subject', '(OnlyInBanner)')

    records.append({
        'SheetName': sheet_guess,
        'CRN sorted {not in Banner}': r0.get('CRN sorted {not in Banner}'),
        'CRN': r0.get('CRN'),
        'Subject': r0.get('Subject'),
        'Course Number {Crse Num}': r0.get('Course Number {Crse Num}'),
        'Course Section {Seq Crse Num}': r0.get('Course Section {Seq Crse Num}'),
        'RowMatchKey': row_key,
        'Column': '(row variant)',
        'DonSheet Value': '(missing row variant)',
        'Banner Value': '(exists in Banner only)',
        'Type': 'RowVariantMissingInDonSheet'
    })

# ---- 14) Cell-level differences for matched row variants --------------------
for _, row in merged.iterrows():
    crn = row[BASE_KEY]
    row_key = row["_RowMatchKey"]
    sheet_for_row = row["SheetName"]
    crn_sorted_val = row.get("CRN sorted {not in Banner}_Working", None)
    subject_val = row.get("Subject_Working", row.get("Subject_Banner", None))
    course_num_val = row.get("Course Number {Crse Num}_Working", row.get("Course Number {Crse Num}_Banner", None))
    section_val = row.get("Course Section {Seq Crse Num}_Working", row.get("Course Section {Seq Crse Num}_Banner", None))

    for w_col, b_col in compare_pairs:
        w_val = row.get(f"{w_col}_Working")
        b_val = row.get(f"{b_col}_Banner")

        if values_are_equal(w_val, b_val, col_name=w_col):
            continue

        w_norm = normalize_empty_value(w_val, col_name=w_col)
        b_norm = normalize_empty_value(b_val, col_name=w_col)

        if w_norm is not None and b_norm is None:
            dtype = 'Red'
        elif w_norm is None and b_norm is not None:
            dtype = 'Green'
        else:
            dtype = 'Blue'

        records.append({
            'SheetName': sheet_for_row,
            'CRN sorted {not in Banner}': crn_sorted_val,
            'CRN': crn,
            'Subject': subject_val,
            'Course Number {Crse Num}': course_num_val,
            'Course Section {Seq Crse Num}': section_val,
            'RowMatchKey': row_key,
            'Column': w_col,
            'DonSheet Value': w_val,
            'Banner Value': b_val,
            'Type': dtype
        })

discrepancy_master_df = pd.DataFrame.from_records(records)

if discrepancy_master_df.empty:
    print("✅ No discrepancies detected — skipping coloring/extra sheets.")
    discrepancy_master_df = pd.DataFrame(columns=[
        'SheetName',
        'CRN sorted {not in Banner}',
        'CRN',
        'Subject',
        'Course Number {Crse Num}',
        'Course Section {Seq Crse Num}',
        'RowMatchKey',
        'Column',
        'DonSheet Value',
        'Banner Value',
        'Type'
    ])
else:
    print(f"🧮 Discrepancy master built: {len(discrepancy_master_df)} rows.")


# In[32]:


# =============================================================================
# STEP 3 — CELL #2
# Build color-coded workbook from the master (preserve gradients & dropdowns)
# =============================================================================
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, GradientFill
from openpyxl.utils import get_column_letter
from copy import copy

RED_GRAD    = GradientFill(stop=("152EFF", "ADD8E6"), degree=90)
GREEN_GRAD  = GradientFill(stop=("152EFF", "ADD8E6"), degree=90)
BLUE_GRAD   = GradientFill(stop=("152EFF", "ADD8E6"), degree=90)
YELLOW_GRAD = GradientFill(stop=("152EFF", "ADD8E6"), degree=90)
ORANGE_GRAD = GradientFill(stop=("152EFF", "ADD8E6"), degree=90)

def apply_discrepancies_to_sheet(ws, disc_df_for_sheet, id_col="CRN sorted {not in Banner}", base_key_col="CRN"):
    """
    Apply cell-level and row-level fills to an individual worksheet.
    Prefer CRN sorted for row mapping if available; otherwise fall back to CRN.
    """
    header_to_col = {}
    for c in range(1, ws.max_column + 1):
        hval = ws.cell(1, c).value
        if hval:
            header_to_col[hval] = c

    id_col_available = id_col in header_to_col
    base_col_available = base_key_col in header_to_col

    if not id_col_available and not base_col_available:
        return

    # Primary mapping
    row_lookup = {}

    if id_col_available:
        id_col_idx = header_to_col[id_col]
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, id_col_idx).value
            if val is not None and str(val).strip() != "":
                row_lookup[("id", str(val).strip())] = r

    if base_col_available:
        base_col_idx = header_to_col[base_key_col]
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, base_col_idx).value
            if val is not None and str(val).strip() != "":
                row_lookup.setdefault(("crn", str(val).strip()), r)

    def find_row(rec):
        crn_sorted = rec.get("CRN sorted {not in Banner}")
        crn = rec.get("CRN")

        if pd.notna(crn_sorted):
            r = row_lookup.get(("id", str(crn_sorted).strip()))
            if r:
                return r

        if pd.notna(crn):
            r = row_lookup.get(("crn", str(crn).strip()))
            if r:
                return r

        return None

    mismatch_idx = header_to_col.get("Mismatching {not in Banner}", None)

    # Entire-CRN missing in Banner
    for _, rec in disc_df_for_sheet.loc[
        disc_df_for_sheet['Type'] == 'CRNMissingInBanner'
    ].iterrows():
        r = find_row(rec)
        if r:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = YELLOW_GRAD

    # Row variant missing in Banner
    for _, rec in disc_df_for_sheet.loc[
        disc_df_for_sheet['Type'] == 'RowVariantMissingInBanner'
    ].iterrows():
        r = find_row(rec)
        if r:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = YELLOW_GRAD

    # Cell-level colors
    cell_df = disc_df_for_sheet[disc_df_for_sheet['Type'].isin(['Red', 'Green', 'Blue'])]

    for _, rec in cell_df.iterrows():
        col_name = rec['Column']
        r = find_row(rec)
        c = header_to_col.get(col_name)

        if r is None or c is None:
            continue

        if rec['Type'] == 'Red':
            ws.cell(r, c).fill = RED_GRAD
            if mismatch_idx is not None:
                ws.cell(r, mismatch_idx).fill = RED_GRAD

        elif rec['Type'] == 'Green':
            ws.cell(r, c).fill = GREEN_GRAD
            if mismatch_idx is not None:
                ws.cell(r, mismatch_idx).fill = GREEN_GRAD

        elif rec['Type'] == 'Blue':
            ws.cell(r, c).fill = BLUE_GRAD
            if mismatch_idx is not None:
                ws.cell(r, mismatch_idx).fill = BLUE_GRAD

with warnings.catch_warnings():
    warnings.filterwarnings('ignore', category=UserWarning)

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    if wb_working_xlsx is not None:
        for sheet in working_sheet_names:
            ws_src = wb_working_xlsx[sheet]
            ws_dst = wb_out.create_sheet(sheet)

            for row in ws_src.iter_rows():
                for cell in row:
                    nc = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        nc.font = copy(cell.font)
                        nc.fill = copy(cell.fill)
                        nc.border = copy(cell.border)
                        nc.alignment = copy(cell.alignment)
                        nc.number_format = cell.number_format
                        nc.protection = copy(cell.protection)

            if not discrepancy_master_df.empty:
                df_sheet = discrepancy_master_df[
                    discrepancy_master_df['SheetName'] == sheet
                ]
                if not df_sheet.empty:
                    apply_discrepancies_to_sheet(ws_dst, df_sheet)

    else:
        for sheet in working_sheet_names:
            ws_dst = wb_out.create_sheet(sheet)
            df_sheet = working_df[working_df['SheetName'] == sheet].copy()
            headers = [c for c in df_sheet.columns if c != 'SheetName']

            for j, h in enumerate(headers, start=1):
                cell = ws_dst.cell(row=1, column=j, value=h)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for i, row in enumerate(df_sheet[headers].itertuples(index=False), start=2):
                for j, val in enumerate(row, start=1):
                    ws_dst.cell(row=i, column=j, value=val)

            if not discrepancy_master_df.empty:
                df_sheet_disc = discrepancy_master_df[
                    discrepancy_master_df['SheetName'] == sheet
                ]
                if not df_sheet_disc.empty:
                    apply_discrepancies_to_sheet(ws_dst, df_sheet_disc)

    # Extra sheet for CRNs / row variants only in Banner
    only_in_banner_crns = discrepancy_master_df.loc[
        discrepancy_master_df['Type'].isin(['CRNMissingInDonSheet', 'RowVariantMissingInDonSheet']),
        'CRN'
    ].dropna().astype(str).unique()

    if len(only_in_banner_crns) > 0:
        ws_extra = wb_out.create_sheet("OnlyInBanner")
        ws_extra.sheet_properties.tabColor = "FF0000"
        subset = banner_df[banner_df[BASE_KEY].astype(str).isin(only_in_banner_crns)].copy()
        cols = list(subset.columns)

        for j, name in enumerate(cols, start=1):
            cell = ws_extra.cell(row=1, column=j, value=name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i, row in enumerate(subset.itertuples(index=False), start=2):
            for j, val in enumerate(row, start=1):
                ws_extra.cell(row=i, column=j, value=val)
            for j in range(1, len(cols) + 1):
                ws_extra.cell(row=i, column=j).fill = ORANGE_GRAD

    try:
        dropdown_ranges = create_dropdown_menu_sheet(wb_out, DropDownMenu_dict)
        for sheet_name in working_sheet_names:
            format_sheet(wb_out[sheet_name], dropdown_ranges)
    except NameError:
        pass

    wb_out.save(Step3_OUTPUT_DiscrepencyReport_Color_full)

print("✅ Color-coded workbook created (CRN-based comparison, row-variant aware).")


# In[33]:


# =============================================================================
# STEP 3 — CELL #3 (CLEAN DETAILED SummaryReport ONLY) + APPLY CELL-2 FORMATTING
# =============================================================================
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string, quote_sheetname
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd

# -----------------------------------------------------------------------------
# 0) Enrich discrepancy_master_df with identifier columns
# -----------------------------------------------------------------------------
extra_cols = ["Subject", "CRN", "Course Number {Crse Num}", "Course Section {Seq Crse Num}"]

# Prefer CRN sorted for lookup when available; otherwise fall back to CRN
WORKING_LOOKUP_KEY = (
    "CRN sorted {not in Banner}"
    if "CRN sorted {not in Banner}" in working_df.columns
    else "CRN"
)

BANNER_LOOKUP_KEY = (
    "CRN sorted {not in Banner}"
    if "CRN sorted {not in Banner}" in banner_df.columns
    else "CRN"
)

DISCREPANCY_LOOKUP_KEY = (
    "CRN sorted {not in Banner}"
    if "CRN sorted {not in Banner}" in discrepancy_master_df.columns
    else "CRN"
)

working_lookup = (
    working_df.drop_duplicates(subset=[WORKING_LOOKUP_KEY], keep="first")
    .set_index(WORKING_LOOKUP_KEY)[[c for c in extra_cols if c in working_df.columns]]
    if WORKING_LOOKUP_KEY in working_df.columns
    else pd.DataFrame()
)

banner_lookup = (
    banner_df.drop_duplicates(subset=[BANNER_LOOKUP_KEY], keep="first")
    .set_index(BANNER_LOOKUP_KEY)[[c for c in extra_cols if c in banner_df.columns]]
    if BANNER_LOOKUP_KEY in banner_df.columns
    else pd.DataFrame()
)

for col in extra_cols:
    if col not in discrepancy_master_df.columns:
        discrepancy_master_df[col] = None

    if col in working_lookup.columns and DISCREPANCY_LOOKUP_KEY in discrepancy_master_df.columns:
        mask = discrepancy_master_df[col].isna()
        discrepancy_master_df.loc[mask, col] = discrepancy_master_df.loc[mask, DISCREPANCY_LOOKUP_KEY].map(
            working_lookup[col]
        )

    if col in banner_lookup.columns and DISCREPANCY_LOOKUP_KEY in discrepancy_master_df.columns:
        mask = discrepancy_master_df[col].isna()
        discrepancy_master_df.loc[mask, col] = discrepancy_master_df.loc[mask, DISCREPANCY_LOOKUP_KEY].map(
            banner_lookup[col]
        )

# Extra fallback by plain CRN if still missing
if "CRN" in discrepancy_master_df.columns:
    working_lookup_crn = (
        working_df.drop_duplicates(subset=["CRN"], keep="first")
        .set_index("CRN")[[c for c in extra_cols if c in working_df.columns and c != "CRN"]]
        if "CRN" in working_df.columns
        else pd.DataFrame()
    )

    banner_lookup_crn = (
        banner_df.drop_duplicates(subset=["CRN"], keep="first")
        .set_index("CRN")[[c for c in extra_cols if c in banner_df.columns and c != "CRN"]]
        if "CRN" in banner_df.columns
        else pd.DataFrame()
    )

    for col in [c for c in extra_cols if c != "CRN"]:
        if col in working_lookup_crn.columns:
            mask = discrepancy_master_df[col].isna()
            discrepancy_master_df.loc[mask, col] = discrepancy_master_df.loc[mask, "CRN"].map(
                working_lookup_crn[col]
            )

        if col in banner_lookup_crn.columns:
            mask = discrepancy_master_df[col].isna()
            discrepancy_master_df.loc[mask, col] = discrepancy_master_df.loc[mask, "CRN"].map(
                banner_lookup_crn[col]
            )

# =============================================================================
# 1) Build / rebuild SummaryReport sheet
# =============================================================================
wb_final = load_workbook(Step3_OUTPUT_DiscrepencyReport_Color_full)

if "SummaryReport" in wb_final.sheetnames:
    del wb_final["SummaryReport"]
ws_sum = wb_final.create_sheet("SummaryReport")
ws_sum.sheet_properties.tabColor = "FF0000"

ws_sum.append([
    "Sheet", "Subject", "CRN sorted {not in Banner}", "CRN",
    "Course Number {Crse Num}", "Course Section {Seq Crse Num}",
    "Column", "Type", "Banner Value", "DonSheet Value", "Status"
])

if not discrepancy_master_df.empty:
    df_disp = discrepancy_master_df.copy()
    df_disp["Sheet"] = df_disp["SheetName"]

    df_disp["Status"] = df_disp["Type"].map({
        "Red": "Value missing in Banner",
        "Green": "Value missing in DonSheet",
        "Blue": "Values differ",
        "CRNMissingInBanner": "CRN missing in Banner",
        "CRNMissingInDonSheet": "CRN missing in DonSheet",
        "RowVariantMissingInBanner": "Row variant missing in Banner",
        "RowVariantMissingInDonSheet": "Row variant missing in DonSheet",
    }).fillna("")

    mask_crn_missing_working = df_disp["Type"] == "CRNMissingInDonSheet"
    df_disp.loc[mask_crn_missing_working, "DonSheet Value"] = df_disp.loc[mask_crn_missing_working, "DonSheet Value"].replace(
        {pd.NA: "(missing CRN)", None: "(missing CRN)", "": "(missing CRN)"}
    )
    df_disp.loc[mask_crn_missing_working, "Banner Value"] = df_disp.loc[mask_crn_missing_working, "Banner Value"].replace(
        {pd.NA: "(exists in Banner only)", None: "(exists in Banner only)", "": "(exists in Banner only)"}
    )

    df_disp = df_disp[[
        "Sheet", "Subject", "CRN sorted {not in Banner}", "CRN",
        "Course Number {Crse Num}", "Course Section {Seq Crse Num}",
        "Column", "Type", "Banner Value", "DonSheet Value", "Status"
    ]].sort_values(
        ["Sheet", "Subject", "CRN", "CRN sorted {not in Banner}", "Column"],
        na_position="last"
    )

    for r in dataframe_to_rows(df_disp, index=False, header=False):
        ws_sum.append(r)

wb_final.save(Step3_OUTPUT_DiscrepencyReport_Color_full)
print("✅ SummaryReport sheet rebuilt.")

# =============================================================================
# 2) External Reports: Detailed + Summary
# =============================================================================
detailed_cols = [
    "SheetName", "Subject", "CRN sorted {not in Banner}", "CRN",
    "Course Number {Crse Num}", "Course Section {Seq Crse Num}",
    "Column", "DonSheet Value", "Banner Value", "Type"
]

detailed_df = (
    discrepancy_master_df[detailed_cols]
    .copy()
    .sort_values(["SheetName", "Subject", "CRN", "CRN sorted {not in Banner}", "Column"])
)

detailed_df.to_excel(Step3_OUTPUT_DetailedValueDiscrepancyReport, index=False)
print(f"✅ DetailedValueDiscrepancyReport saved → {Step3_OUTPUT_DetailedValueDiscrepancyReport}")

summary_counts = (
    discrepancy_master_df.groupby(["SheetName", "Type"])
    .size()
    .reset_index(name="Count")
    .sort_values(["SheetName", "Type"])
)
summary_counts.to_excel(Step3_OUTPUT_SummaryReport, index=False)
print(f"✅ SummaryReport (counts) saved separately → {Step3_OUTPUT_SummaryReport}")

# =============================================================================
# 3) APPLY THE SAME FORMATTING AS YOUR "CELL 2"
# =============================================================================
thin_grey_border = Border(
    left=Side(style='thin', color="D3D3D3"),
    right=Side(style='thin', color="D3D3D3"),
    top=Side(style='thin', color="D3D3D3"),
    bottom=Side(style='thin', color="D3D3D3"),
)

COLOR_DATA_CELLS = False

BLUE_HEADERS = [
    "Don {not in Banner}",
    "Mismatching {not in Banner}",
    "Dept/Prog Admin {not in Banner}",
    "Mona {not in Banner}",
    "Karen {not in Banner}",
    "DONE {not in Banner}",
    "Notes {not in Banner}",
    "Subject",
    "Course Number {Crse Num}",
    "Course Section {Seq Crse Num}",
    "Program {not in Banner}",
    "Catalog Title",
    "Section Title",
    "Crosslist Details {not in Banner}",
    "Campus {Ssasect Campus}",
    "Campus Restriction {Camp Restr CC}",
    "Schedule Type {not in Banner}",
    "Instruction Method {Inst Method}",
    "Level {Level SC}",
    "Part-Of-Term",
    "Credits {Sect Crs}",
    "Enrollment Cap {Max Enrl}",
    "Meeting Type",
    "Pre-work Start Date {not in Banner}",
    "Pre-work End Date {not in Banner}",
    "Course Start Date {Meet Start Date}",
    "Course End Date {Meet End Date}",
    "Post-work Start Date {not in Banner}",
    "Post-work End Date {not in Banner}",
    "Course Beginning Time {Meet Beg Time}",
    "Course Ending Time {Meet End Time}",
    "SUN", "MON", "TUE", "WED", "THU", "FRI", "SAT",
    "Room {Meet Room}",
    "Building {Meet Bldg}",
    "Instructor Name {Instr Name}",
    "Instructor Email {Instr Email}",
    "Instructor ID {Instr ID}",
    "% Responsibility",
    "load/contract {not in Banner}",
    "costs per credit {not in Banner}",
    "total costs {not in Banner}",
    "account to be charged {not in Banner}",
]

GREY_HEADERS = [
    "CRN sorted {not in Banner}",
    "CRN",
    "X Lst",
    "Special Approval {Special Apvl}",
    "Credit Catalog {Cat Crs}",
    "Semester Start Date {Soaterm Start Date}",
    "Semester End Date {Soaterm End Date}",
    "Weeks",
    "Fee Amount {Fees Amt}",
    "Fee Code {Detl Code}",
    "Fee Type",
    "Fee Level",
    "Fee Ind",
    "Fee Term",
    "SEM Department {Scacrse Dept}",
    "Activity Date",
    "Status",
    "Bill Hrs",
    "Sect Sch Type",
    "Prerequisites {Preq Areas}",
    "Link Ind",
    "Integration Code",
    "Link Conn",
    "Crs Attr",
    "Comments",
    "Major Restr",
    "Scacrse College",
    "Enrolled",
    "Waitlist Capacity",
    "Primary Ind",
    "Override",
    "1st Date Reg Opens",
    "Last Date Reg Opens",
    "OL Range From Date",
    "OL Range To Date",
    "OL Numb Units",
    "OL Duration Code",
    "Grade Mode",
    "Gradable Ind",
    "Tuit Waiver",
    "Course Override {Meet Override}",
    "Lvl Res Ind",
    "Cmp Res Ind",
    "Rate Code",
    "Cohort Code",
    "St Attr Code",
    "Coll SC",
    "DegC SC",
    "Prog SC",
    "FOS Type",
    "Term Code",
    "Part-Of-Term",
    "Camp SC",
]

HEADER_DEFAULT_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
HEADER_BLUE_FILL    = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
HEADER_GREY_FILL    = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

HEADER_FILL_BY_NAME = {h: HEADER_BLUE_FILL for h in BLUE_HEADERS}
HEADER_FILL_BY_NAME.update({h: HEADER_GREY_FILL for h in GREY_HEADERS})

GROUP_DEFS = [
    ('Q', 'V'),
    ('X', 'Z'),
    ('AB', 'AH'),
    ('AJ', 'AW'),
    ('AY', 'BA'),
    ('BC', 'CV')
]

FIXED_WIDTHS = {
    "Notes {not in Banner}": 40,
    "Instructor Name {Instr Name}": 20,
    "Instructor Email {Instr Email}": 20,
    "Instructor ID {Instr ID}": 10,
    "account to be charged {not in Banner}": 20,
}

def build_dropdown_ranges_from_dd_sheet(wb, dd_sheet_name="DropDownMenu"):
    if dd_sheet_name not in wb.sheetnames:
        return {}, {}

    ws_dd = wb[dd_sheet_name]
    dropdown_ranges = {}
    dropdown_values = {}

    for col_idx in range(1, ws_dd.max_column + 1):
        header = ws_dd.cell(row=1, column=col_idx).value
        if header is None or str(header).strip() == "":
            continue

        values = []
        for r in range(2, ws_dd.max_row + 1):
            v = ws_dd.cell(row=r, column=col_idx).value
            if v is None or str(v).strip() == "":
                continue
            values.append(str(v))

        if not values:
            continue

        col_letter = get_column_letter(col_idx)
        last_row = 1 + len(values)
        dropdown_ranges[header] = f"{quote_sheetname(dd_sheet_name)}!${col_letter}$2:${col_letter}${last_row}"
        dropdown_values[header] = values

    return dropdown_ranges, dropdown_values

def reapply_dropdown_validations(sheet, dropdown_ranges, dropdown_values, max_rows=5000):
    headers = [cell.value for cell in sheet[1]]
    for header, formula_range in dropdown_ranges.items():
        if header not in headers:
            continue

        cidx = headers.index(header) + 1
        cL = get_column_letter(cidx)

        dv = DataValidation(type="list", formula1=f"={formula_range}", allow_blank=True)
        dv.showDropDown = False
        sheet.add_data_validation(dv)
        dv.add(f"${cL}$2:${cL}${max_rows}")

        values_list = dropdown_values.get(header, [])
        if "False" in values_list:
            for r in range(2, min(sheet.max_row, max_rows) + 1):
                cell = sheet[f"{cL}{r}"]
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = "False"

def group_columns(sheet, group_defs=GROUP_DEFS, outline_level=1):
    sheet.sheet_properties.outlinePr.summaryRight = False
    sheet.sheet_properties.outlinePr.summaryBelow = True
    sheet.sheet_properties.outlinePr.applyStyles = True

    max_col = sheet.max_column or 1
    for start_col, end_col in group_defs:
        start_idx = column_index_from_string(start_col)
        end_idx   = column_index_from_string(end_col)

        if start_idx > max_col:
            continue

        end_idx = min(end_idx, max_col)
        for col_idx in range(start_idx, end_idx + 1):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].outline_level = outline_level

def autofit_columns_ignore_header(sheet, min_width=4, max_width=60, padding=2, max_rows=5000):
    max_row = min(sheet.max_row or 1, max_rows)
    max_col = sheet.max_column or 1

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        for r in range(2, max_row + 1):
            v = sheet.cell(row=r, column=col_idx).value
            if v is None:
                continue
            max_length = max(max_length, len(str(v)))

        sheet.column_dimensions[col_letter].width = max(min_width, min(max_length + padding, max_width))

def apply_fixed_widths(sheet, fixed_widths=FIXED_WIDTHS):
    headers = [cell.value for cell in sheet[1]]
    for header, width in fixed_widths.items():
        if header in headers:
            col_idx = headers.index(header) + 1
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = width

def format_sheet_like_cell2(sheet):
    headers = [cell.value for cell in sheet[1]]

    if "Notes {not in Banner}" in headers:
        notes_idx = headers.index("Notes {not in Banner}") + 1
        sheet.freeze_panes = f"{get_column_letter(notes_idx + 1)}2"
    else:
        sheet.freeze_panes = "A2"

    last_col = get_column_letter(sheet.max_column)
    sheet.auto_filter.ref = f"A1:{last_col}{sheet.max_row}"

    for cell in sheet[1]:
        if cell.value is None or str(cell.value).strip() == "":
            continue
        header_text = str(cell.value).strip()

        cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color="FF0000" if header_text == "Notes {not in Banner}" else "000000",
        )
        cell.fill = HEADER_DEFAULT_FILL
        cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True, text_rotation=90)

    sheet.row_dimensions[1].height = 150

    for col_idx, header_val in enumerate(headers, start=1):
        if header_val is None:
            continue
        header_text = str(header_val).strip()
        fill = HEADER_FILL_BY_NAME.get(header_text)
        if fill is None:
            continue

        sheet.cell(row=1, column=col_idx).fill = fill
        if COLOR_DATA_CELLS:
            for r in range(2, sheet.max_row + 1):
                sheet.cell(row=r, column=col_idx).fill = fill

    if "Status" in headers:
        status_col_idx = headers.index("Status") + 1
        inactive_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for r in range(2, sheet.max_row + 1):
            v = sheet.cell(row=r, column=status_col_idx).value
            if str(v).strip().upper() == "I":
                for c in range(1, sheet.max_column + 1):
                    sheet.cell(row=r, column=c).fill = inactive_fill

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_grey_border

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            f = cell.font or Font()
            cell.font = Font(
                name="Arial",
                size=10,
                bold=f.bold,
                italic=f.italic,
                underline=f.underline,
                color=f.color,
            )

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=min(7, sheet.max_column)):
        for cell in row:
            f = cell.font or Font()
            cell.font = Font(
                name="Arial",
                size=10,
                bold=f.bold,
                italic=f.italic,
                underline=f.underline,
                color=f.color,
            )

dropdown_ranges, dropdown_values = build_dropdown_ranges_from_dd_sheet(wb_final, dd_sheet_name="DropDownMenu")

for sname in [sn for sn in wb_final.sheetnames if sn != "DropDownMenu"]:
    ws = wb_final[sname]
    format_sheet_like_cell2(ws)
    group_columns(ws)
    autofit_columns_ignore_header(ws)
    apply_fixed_widths(ws)
    reapply_dropdown_validations(ws, dropdown_ranges, dropdown_values, max_rows=5000)

wb_final.save(Step3_OUTPUT_DiscrepencyReport_Color_full)
wb_final.close()
print("🎨 Step3_OUTPUT_DiscrepencyReport_Color_full formatted with the SAME styling rules as Cell 2.")


# In[ ]:





# In[ ]:




