#!/usr/bin/env python
# coding: utf-8

# # CSV to XLSX Converter with Custom Tab Organization
# 
# This notebook converts a CSV file to an XLSX file with:
# - Course-based tab organization
# - PhD tab for courses 600+
# - DMin tab for courses 700-799
# - MA religion tab for INT campus courses
# - MDiv Hispanic tab for Florida location courses
# - MAPM Engl. tab for Forest Lake/Burman courses (sections 30-199)
# - Custom column headers
# - Column mapping from CSV to XLSX
# - Section extraction from Course column

# In[1]:


import os
import re
import glob
import numpy as np
from copy import copy
from collections import defaultdict
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation


# In[2]:


# Making sure we are working within the same directory as the jupyter notebook (this will allow for poratbility of the project)
import os
## Verify we're in the right directory
print("Current directory:", os.getcwd())


# In[3]:


import pandas as pd

# Show all columns
pd.set_option('display.max_columns', None)

# Show all rows (optional — be careful if your BannerPrevious is huge)
pd.set_option('display.max_rows', None)

# Don’t truncate long strings
pd.set_option('display.max_colwidth', None)

# Make the display wider in the notebook
pd.set_option('display.width', 0)


# ## Configuration

# In[4]:


import glob

# File Paths
## First Input File: Downloaded AU Course Schedule (supports CSV, XLS, XLSX)
BannerPreviousYearFile =    (
                            glob.glob('Step1_input_downloaded_AU_BannerPrevious/*.csv') + 
                            glob.glob('Step1_input_downloaded_AU_BannerPrevious/*.xls') + 
                            glob.glob('Step1_input_downloaded_AU_BannerPrevious/*.xlsx')
                            )
if not BannerPreviousYearFile:
    raise FileNotFoundError("No CSV, XLS, or XLSX files found in Step1_input_downloaded_AU_BannerPrevious directory")
BannerPreviousYearFile_INPUT = BannerPreviousYearFile[0]
print(f"Found Banner file: {BannerPreviousYearFile_INPUT}")

DropDown_RETRIEVAL = '0000_BuildingFiles/source_DropDownMenus.xlsx'

#New_SEM_course_updates
### First Output File: Basic Calculated Spreadsheet File
New_SEM_course_updates = 'x_SEM_CourseUpdate_LiveFolder_office365/New_SEM_course_updates_YYYYsemester.xlsx'
#New_SEM_course_updates = 'Step1_output_SEM_CourseUpdates_Spreadsheet/New_SEM_course_updates_YYYYsemester.xlsx'


# # Identical Cells in `Step3 code` (exception: `BannerPrevious` [Step1] = `BannerRecentFile` [Step3])

# ## Loading Banner Data

# In[5]:


import pandas as pd
import re

def find_header_row(file_path, max_rows_to_check=20):
    """
    Automatically identify the header row in an Excel file.
    Reads everything as text to preserve exact cell values,
    but detects headers by checking for mostly non-numeric strings.
    """
    # Read file as text
    BannerPrevious_raw = pd.read_excel(
        file_path,
        header=None,
        nrows=max_rows_to_check,
        dtype=str,
        na_filter=False
    )

    best_row = 0
    max_score = 0

    for idx in range(len(BannerPrevious_raw)):
        row = BannerPrevious_raw.iloc[idx].fillna('')

        # Define helper functions
        def looks_numeric(val):
            return bool(re.match(r"^\s*-?\d+(\.\d+)?\s*$", val.strip()))

        # Count non-empty cells
        non_empty_count = sum(val.strip() != '' for val in row)

        # Count values that look like text (not numeric)
        text_like_count = sum((val.strip() != '' and not looks_numeric(val)) for val in row)

        # Scoring logic:
        #   reward text-like cells, reward non-empty rows,
        #   penalize rows with mostly numeric values
        score = text_like_count * 2 + non_empty_count

        # Bonus for having a reasonable fill ratio
        if non_empty_count > len(row) * 0.3:
            score += 10

        if score > max_score:
            max_score = score
            best_row = idx

    return best_row


def load_dataframe(file_path, max_rows_to_check=20):
    """
    Load Excel file with auto header detection and exact text preservation.
    """
    header_row = find_header_row(file_path, max_rows_to_check)
    print(f"Header row identified at row index: {header_row}")

    BannerPrevious = pd.read_excel(
        file_path,
        header=header_row,
        dtype=str,
        na_filter=False
    )
    return BannerPrevious


# === Example usage ===
if __name__ == "__main__":
    file_path = BannerPreviousYearFile_INPUT
    BannerPrevious = load_dataframe(file_path)

    print(f"\nDataFrame shape: {BannerPrevious.shape}")
    print(f"\nColumn names:")
    print(BannerPrevious.columns.tolist())
    print(f"\nFirst 2 rows:")
    print(BannerPrevious.head(2))
    print(f"\nData types:")
    print(BannerPrevious.dtypes)


# ### Adapting Banner Input

# In[6]:


BannerPrevious.head(2)


# In[7]:


import pandas as pd

def add_crn_sorted_column(BannerPrevious):
    print("=== Starting add_crn_sorted_column ===")

    df = BannerPrevious.copy()

    if 'CRN' not in df.columns:
        raise KeyError("CRN column not found in DataFrame")
    print("Found CRN column")

    crn_series = df['CRN'].astype(str).str.strip()
    print("Converted CRN to string")

    crn_numeric = pd.to_numeric(crn_series, errors='coerce')
    is_digitlike = crn_series.str.match(r'^\d+$')
    df['__crn_group'] = crn_series.where(~is_digitlike, crn_numeric.astype('Int64').astype(str))
    print("Created __crn_group helper column")

    df['__occurrence'] = df.groupby('__crn_group', sort=False).cumcount() + 1
    print("Created __occurrence helper column")

    crn_sorted = crn_series + '-' + df['__occurrence'].astype(str).str.zfill(2)
    empty_mask = crn_series == ''
    crn_sorted.loc[empty_mask] = ''
    print("Built CRN sorted values")

    df.drop(columns=['__crn_group', '__occurrence'], inplace=True)
    print("Dropped helper columns")

    crn_pos = df.columns.get_loc('CRN')
    if 'CRN sorted {not in Banner}' in df.columns:
        df.drop(columns='CRN sorted {not in Banner}', inplace=True)
    df.insert(crn_pos + 1, 'CRN sorted {not in Banner}', crn_sorted)
    print("Inserted 'CRN sorted {not in Banner}' column successfully")

    print("=== Function complete ===")
    return df


# In[8]:


BannerPrevious = add_crn_sorted_column(BannerPrevious)
print(BannerPrevious.columns.tolist())
print(BannerPrevious.head())


# In[9]:


BannerPrevious.head(2)


# In[10]:


def normalize_empty_values(series):
    """
    Normalize a pandas Series by:
    - Converting 'nan', 'none', 'na', 'n/a' (case-insensitive) to np.nan
    - Converting empty strings or whitespace-only to np.nan
    - Stripping leading/trailing spaces
    - Preserving original casing otherwise
    """
    # First ensure string type and strip whitespace
    s = series.astype(str).str.strip()

    # Replace known string nulls and empty strings
    s = s.replace(
        to_replace=['', 'nan', 'NaN', 'NAN', 'none', 'None', 'NA', 'N/A', 'na', 'n/a'],
        value=np.nan
    )

    return s


columns_to_process = [
    'CRN sorted {not in Banner}', 'CRN', 'Activity Date', 'Term Code', 'Subject', 'Crse Num',
    'Seq Crse Num', 'X Lst', 'Catalog Title', 'Section Title', 'Status',
    'Cat Crs', 'Sect Crs', 'Bill Hrs', 'Sect Sch Type', 'Preq Areas',
    'Inst Method', 'Link Ind', 'Integration Code', 'Link Conn', 'Crs Attr',
    'Comments', 'Ssasect Campus', 'Camp Restr CC', 'Major Restr',
    'Scacrse College', 'Scacrse Dept', 'Special Apvl', 'Max Enrl', 'Enrolled',
    'Waitlist Capacity', 'Instr ID', 'Instr Name', 'Instr Email',
    '% Responsibility', 'Primary Ind', 'Override', 'Part-Of-Term',
    'Soaterm Start Date', 'Soaterm End Date', 'Weeks', '1st Date Reg Opens',
    'Last Date Reg Opens', 'OL Range From Date', 'OL Range To Date',
    'OL Numb Units', 'OL Duration Code', 'Grade Mode', 'Gradable Ind',
    'Tuit Waiver', 'Meeting Type', 'SUN', 'MON', 'TUE', 'WED', 'THU', 'FRI',
    'SAT', 'Meet Start Date', 'Meet End Date', 'Meet Beg Time', 'Meet End Time',
    'Meet Bldg', 'Meet Room', 'Meet Override', 'Lvl Res Ind', 'Cmp Res Ind', 'Fee Term', 'Fee Level',
    'Fees Amt', 'Fee Type', 'Fee Ind', 'Detl Code', 'Rate Code', 'Cohort Code',
    'St Attr Code', 'Level SC', 'Camp SC', 'Coll SC', 'DegC SC', 'Prog SC',
    'FOS Type', 'FOS Code', 'Dept Code', 'Admit Term', 'Curr Rate', 'Curr Styp',
    'Curricula'
]

for col in columns_to_process:
    if col in BannerPrevious.columns:
        BannerPrevious[col] = normalize_empty_values(BannerPrevious[col])
        print(f"{col}: {BannerPrevious[col].dtype}")
    else:
        print(f"Warning: Column '{col}' not found in BannerPrevious")


# In[11]:


BannerPrevious.head()


# In[12]:


# ==============================================================================
# BANNER COLUMN RENAMING
# ==============================================================================

rename_dict = {
    'Crse Num': 'Course Number {Crse Num}',
    'Seq Crse Num': 'Course Section {Seq Crse Num}',
    'Cat Crs': 'Credit Category {Cat Crs}',
    'Sect Crs': 'Credits {Sect Crs}',
    'Preq Areas': 'Prerequisites {Preq Areas}',
    'Inst Method': 'Instruction Method {Inst Method}',
    'Camp Restr CC': 'Campus Restriction {Camp Restr CC}',
    'Scacrse Dept': 'SEM Department {Scacrse Dept}',
    'Special Apvl': 'Special Approval {Special Apvl}',
    'Max Enrl': 'Enrollment Cap {Max Enrl}',
    'Instr ID': 'Instructor ID {Instr ID}',
    'Instr Name': 'Instructor Name {Instr Name}',
    'Instr Email': 'Instructor Email {Instr Email}',
    'Meet Start Date': 'Course Start Date {Meet Start Date}',
    'Meet End Date': 'Course End Date {Meet End Date}',
    'Meet Beg Time': 'Course Beginning Time {Meet Beg Time}',
    'Meet End Time': 'Course Ending Time {Meet End Time}',
    'Meet Bldg': 'Building {Meet Bldg}',
    'Meet Room': 'Room {Meet Room}',
    'Meet Override': 'Course Override {Meet Override}',
    'Fees Amt': 'Fee Amount {Fees Amt}',
    'Detl Code': 'Fee Code {Detl Code}',
    'Level SC': 'Level {Level SC}',
    'Ssasect Campus': 'Campus {Ssasect Campus}',
    'Soaterm Start Date':'Semester Start Date {Soaterm Start Date}',
    'Soaterm End Date':'Semester End Date {Soaterm End Date}',
    'Cat Crs':'Credit Catalog {Cat Crs}'
}

BannerPrevious.rename(columns=rename_dict, inplace=True)

BannerPrevious.head(2)


# In[13]:


site_name_dict = {
'ACW':'Wellness Center [ACW]',	
'BH':'Bell Hall [BH]',	
'BUL':'Buller Hall [BUL]',	
'CSH':'Chan Shun Hall [CSH]',	
'GHA':'Griggs Hall A [GHA]',
'GHB':'Griggs Hall B [GHB]',	
'HORN':'Horn Museum [HORN]',	
'JGYM':'Johnson Gym [JGYM]',	
'JWL':'James White Library [JWL]',	
'NH':'Nethery Hall [NH]',	
'OCARG':'Univ Adventista del Plata [OCARG]',	
'OCBRA1':'Centro Univ Adv de Sao Paulo [OCBRA1]',	
'OCCABU':'Burman University [OCCABU]',	
'OCCALL':'Loma Linda University [OCCALL]',	
'OCCASC':'Southeastern California Conf [OCCASC]',	
'OCCHK':'Hong Kong Adventist College [OCCHK]',	
'OCFLAH':'Advent Health University [OCFLAH]',	
'OCFLFC':'Florida Conference of SDA [OCFLFC]',	
'OCFLFL':'Forest Lake SDA Church [OCFLFL]',	
'OCGBR':'Newbold College [OCGBR]',	
'OCMDND':'North American Division of SDA [OCMDND]',	
'OCMENE':'Northern New England Conf [OCMENE]',	
'OCNEUC':'Union College [OCNEUC]',	
'OCOKOC':'Oklahoma City Central [OCOKOC]',	
'OCPOL':'Polish Senior College Theo&Hum [OCPOL]',	
'OCROU':'Universitatea Adventus din Cer [OCROU]',	
'OCRUS':'Zaokski Theo Seminary [OCRUS]',	
'OCTHA':'Asia-Pacific Int Univ AIU [OCTHA]',	
'OCTWN':'Taiwan Adventist College [OCTWN]',	
'OCUKR':'Ukrainan Adv Center Higher ED [OCUKR]',	
'OCWANP':'North Pacific Union Conference [OCWANP]',	
'PMC':'Pioneer Memorial Church [PMC]',	
'SEM':'Seminary Building [SEM]'
}

def replace_site_prefix(value):
    """Replace site prefix before ':' using site_name_dict."""
    if not isinstance(value, str) or ':' not in value:
        return value  # leave non-string or missing values unchanged

    prefix, rest = value.split(':', 1)
    prefix = prefix.strip()
    rest = rest.strip()

    # Replace prefix if found in dict
    if prefix in site_name_dict:
        return f"{site_name_dict[prefix]}: {rest}"
    else:
        return value

# Apply to Meet Bldg column
# Replace short codes with full names
BannerPrevious["Building {Meet Bldg}"] = BannerPrevious["Building {Meet Bldg}"].replace(site_name_dict)

print("✅ 'Building {Meet Bldg}' values replaced successfully.")

print("✅ 'Building {Meet Bldg}' prefixes replaced using site_name_dict")
BannerPrevious["Building {Meet Bldg}"].head(2)


# In[14]:


BannerPrevious["Building {Meet Bldg}"].unique()


# In[15]:


BannerPrevious.head(2)


# In[16]:


BannerPrevious = pd.concat(
    [pd.DataFrame(columns=["Don {not in Banner}", "Mismatching {not in Banner}", "Dept/Prog Admin {not in Banner}", "Mona {not in Banner}", "Karen {not in Banner}", "DONE {not in Banner}", "Notes {not in Banner}", "Program {not in Banner}", "Crosslist Details {not in Banner}", "Schedule Type {not in Banner}", "Pre-work Start Date {not in Banner}", "Pre-work End Date {not in Banner}", "Post-work Start Date {not in Banner}", "Post-work End Date {not in Banner}", "load/contract {not in Banner}", "costs per credit {not in Banner}", "total costs {not in Banner}", "account to be charged {not in Banner}"]), BannerPrevious],
    axis=1
)
BannerPrevious.head(2)


# In[17]:


# Define the exact column order
desired_order = [
    "Don {not in Banner}", 
    "Mismatching {not in Banner}", 
    "Dept/Prog Admin {not in Banner}",
    "Mona {not in Banner}", 
    "Karen {not in Banner}",
    "DONE {not in Banner}", 
    "Notes {not in Banner}",
    "CRN sorted {not in Banner}",
    "CRN",
    "Subject",
    "Course Number {Crse Num}",
    "Course Section {Seq Crse Num}",
    "Program {not in Banner}",
    "Catalog Title",
    "Section Title",
    "Crosslist Details {not in Banner}", #added and not in banner
    "X Lst",
    "Campus {Ssasect Campus}",
    "Campus Restriction {Camp Restr CC}",   #Did 'Campus' get back in?? WHen, where, and how can we get it out
    "Schedule Type {not in Banner}",
    "Instruction Method {Inst Method}", #not working
    "Level {Level SC}",
    "Credit Catalog {Cat Crs}",
    "Credits {Sect Crs}",
    "Enrollment Cap {Max Enrl}",
    "Meeting Type",
    "Semester Start Date {Soaterm Start Date}",
    "Pre-work Start Date {not in Banner}",
    "Pre-work End Date {not in Banner}",
    "Course Start Date {Meet Start Date}",
    "Course End Date {Meet End Date}",
    "Post-work Start Date {not in Banner}",
    "Post-work End Date {not in Banner}",
    "Semester End Date {Soaterm End Date}",
    "Course Beginning Time {Meet Beg Time}",
    "Course Ending Time {Meet End Time}",
    "SUN",
    "MON",
    "TUE",
    "WED",
    "THU",
    "FRI",
    "SAT",
    "Room {Meet Room}",
    "Building {Meet Bldg}",
    "Instructor Name {Instr Name}",
    "Instructor Email {Instr Email}",
    "Instructor ID {Instr ID}",
    "% Responsibility",
    "load/contract {not in Banner}",
    "costs per credit {not in Banner}",
    "total costs {not in Banner}", 
    "account to be charged {not in Banner}",
    "Fee Amount {Fees Amt}",
    "Fee Code {Detl Code}",
    "Fee Type",
    "Fee Level",
    "Fee Ind",
    "Fee Term",
    "SEM Department {Scacrse Dept}",
    "Activity Date",
    "Status",
    "Bill Hrs",
    "Sect Sch Type",
    "Prerequisites {Preq Areas}",
    "Link Ind",
    "Integration Code",
    "Link Conn",
    "Crs Attr",
    "Comments",
    "Major Restr",
    "Scacrse College",
    "Enrolled",
    "Waitlist Capacity",
    "Primary Ind",
    "Override",
    "1st Date Reg Opens",
    "Last Date Reg Opens",
    "OL Range From Date",
    "OL Range To Date",
    "OL Numb Units",
    "OL Duration Code",
    "Grade Mode",
    "Gradable Ind",
    "Tuit Waiver",
    "Course Override {Meet Override}",
    "Lvl Res Ind",
    "Cmp Res Ind", 
    "Rate Code",
    "Cohort Code",
    "St Attr Code",
    "Coll SC",
    "DegC SC",
    "Prog SC",
    "FOS Type",
    "Special Approval {Special Apvl}",
    "Weeks",
    "Term Code",
    "Part-Of-Term",
    "Camp SC" #Is this the one tha appears as 'Campus' get back in?? WHen, where, and how can we get it out

]

# Reorder columns
BannerPrevious = BannerPrevious[[col for col in desired_order if col in BannerPrevious.columns]]

# (Optional) Report any columns that were missing
missing_cols = [col for col in desired_order if col not in BannerPrevious.columns]
if missing_cols:
    print("⚠️ The following columns were not found in BannerPrevious:")
    for col in missing_cols:
        print("  -", col)
else:
    print("✅ All columns reordered successfully.")

# Preview the reordered DataFrame
BannerPrevious.head(2)


# In[18]:


BannerPrevious.shape


# # Reducing Dataset to DonSheet relevant data.

# In[19]:


BannerPrevious = BannerPrevious[
     (BannerPrevious["Scacrse College"] == "70")
     &
    ~(BannerPrevious["Status"] == "I")
].copy()



print(f"✅ Filtered rows: {len(BannerPrevious)} out of {len(BannerPrevious)} total")
BannerPrevious["Scacrse College"].value_counts()


# In[20]:


BannerPrevious.shape


# In[21]:


cols_to_check = ["Lvl Res Ind", "Cmp Res Ind"]

existing = [c for c in cols_to_check if c in BannerPrevious.columns]

if existing:
    BannerPrevious = BannerPrevious[
        BannerPrevious[existing].isna().all(axis=1)
    ].copy()


# In[22]:


BannerPrevious.shape


# ### Minor Testing

# In[23]:


BannerPreviousDummy = BannerPrevious
DUMMY = BannerPreviousDummy[(
    (
        #(BannerPreviousDummy["Scacrse College"] == 70) 
        #& 
        (BannerPreviousDummy['CRN'] == "4730")
        & 
        (BannerPreviousDummy['Course Section {Seq Crse Num}'] == 'XXX')
    )
)].copy()
DUMMY.head()


# In[24]:


BannerPreviousDummy = BannerPrevious
DUMMY = BannerPreviousDummy[(
    (
        #(BannerPreviousDummy["Scacrse College"] == 70) 
        #& 
        (BannerPreviousDummy['Subject'] == "OTST")
        #& 
        #(BannerPreviousDummy['Course Section {Seq Crse Num}'] == 'XXX')
    )
)].copy()
DUMMY.head()


# In[25]:


BannerPrevious.head()


# In[26]:


BannerPrevious.to_excel("0000_source_files/banner_DonLike_dump.xlsx")


# In[27]:


#BannerPrevious = pd.read_excel("/Users/oliverglanz/Library/CloudStorage/OneDrive-AndrewsUniversity/0000_EfficiencyWithIT/SmartScheduling/0000_source_files/banner_DonLike_dump.xlsx")
#BannerPrevious.head()


# In[28]:


#BannerPrevious.drop(columns=['Unnamed: 0'], inplace=True)
#BannerPrevious.head()


# # Creating Unified DonSheet (DonSheet with just one tab/sheet)

# In[29]:


import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------
# Output path
# ------------------------------------------------------------
output_file = ("0000_source_files/DonSheet_unified.xlsx"
)

# ------------------------------------------------------------
# 1. Export DataFrame (no formatting yet)
# ------------------------------------------------------------
BannerPrevious.to_excel(output_file, index=False)

# ------------------------------------------------------------
# 2. Load workbook for formatting
# ------------------------------------------------------------
wb = load_workbook(output_file)
ws = wb.active

# ------------------------------------------------------------
# 3. Header formatting: vertical, bottom-center
# ------------------------------------------------------------
header_alignment = Alignment(
    textRotation=90,
    vertical="bottom",
    horizontal="center",
    wrap_text=False
)

for cell in ws[1]:
    cell.alignment = header_alignment

# ------------------------------------------------------------
# 4. Freeze header row
# ------------------------------------------------------------
ws.freeze_panes = "A2"

# ------------------------------------------------------------
# 5. Apply autofilter
# ------------------------------------------------------------
ws.auto_filter.ref = ws.dimensions

# ------------------------------------------------------------
# 6. Autofit column width (IGNORE header row)
# ------------------------------------------------------------
for col_idx in range(1, ws.max_column + 1):
    col_letter = get_column_letter(col_idx)
    max_length = 0

    for row in ws.iter_rows(
        min_row=2,              # <-- ignore header row
        min_col=col_idx,
        max_col=col_idx
    ):
        cell_value = row[0].value
        if cell_value is not None:
            max_length = max(max_length, len(str(cell_value)))

    # conservative padding
    ws.column_dimensions[col_letter].width = max(10, min(max_length + 2, 60))

# ------------------------------------------------------------
# 7. Save
# ------------------------------------------------------------
wb.save(output_file)

print(f"Formatted file written to:\n{output_file}")


# # Creating DonSheet with a Sheet/Tab per Department

# ## Creating different DF for each Department and Program (No longer idential to Step3!)

# In[30]:


# ------------------------------------------------------------
# Normalize key filter columns (robust across years)
# Put this RIGHT AFTER you load BannerPrevious
# ------------------------------------------------------------
BannerPrevious['Subject'] = BannerPrevious['Subject'].astype(str).str.strip()

BannerPrevious['Campus {Ssasect Campus}'] = (
    BannerPrevious['Campus {Ssasect Campus}']
    .astype(str)
    .str.strip()
)

# Convert course number ONCE (no need to reconvert inside every filter)
BannerPrevious['Course Number {Crse Num}'] = pd.to_numeric(
    BannerPrevious['Course Number {Crse Num}'], errors='coerce'
)

# (Optional: only if you actually use it later)
# BannerPrevious['Course Section {Seq Crse Num}'] = pd.to_numeric(
#     BannerPrevious['Course Section {Seq Crse Num}'], errors='coerce'
# )

# Helpful masks to avoid repeating logic everywhere
mask_main_campus = BannerPrevious['Campus {Ssasect Campus}'].isin(['0', '7'])
mask_not_700s = (
    (BannerPrevious['Course Number {Crse Num}'] < 700) |
    (BannerPrevious['Course Number {Crse Num}'] > 799)
)

# ------------------------------------------------------------
# Program DataFrames
# ------------------------------------------------------------
NTST = BannerPrevious[
    (BannerPrevious['Subject'] == 'NTST') &
    mask_not_700s &
    mask_main_campus
].copy()

OTST = BannerPrevious[
    (BannerPrevious['Subject'].isin(['OTST', 'ANEA'])) &
    mask_not_700s &
    mask_main_campus
].copy()

CHIS = BannerPrevious[
    (BannerPrevious['Subject'] == 'CHIS') &
    mask_not_700s &
    mask_main_campus
].copy()

DSLE = BannerPrevious[
    (BannerPrevious['Subject'] == 'DSLE') &
    mask_not_700s &
    mask_main_campus
].copy()

GSEM = BannerPrevious[
    (BannerPrevious['Subject'] == 'GSEM') &
    mask_not_700s &
    mask_main_campus
].copy()

MSSN = BannerPrevious[
    (BannerPrevious['Subject'] == 'MSSN') &
    mask_not_700s &
    mask_main_campus
].copy()

PATH = BannerPrevious[
    (BannerPrevious['Subject'] == 'PATH') &
    mask_not_700s &
    mask_main_campus
].copy()

THST = BannerPrevious[
    (BannerPrevious['Subject'] == 'THST') &
    mask_not_700s &
    mask_main_campus
].copy()

# NOTE: Your comment says '9OM' here; keep as-is.
# Also: this one intentionally does NOT apply mask_not_700s or mask_main_campus.
MDivHISP_MAPmENGL_MAPmHISP = BannerPrevious[
    (BannerPrevious['Campus {Ssasect Campus}'] == '9OM')
].copy()

# NOTE: This one intentionally does NOT apply mask_main_campus.
MA_Religion = BannerPrevious[
    (BannerPrevious['Campus {Ssasect Campus}'] == 'ZOM')
].copy()

# DMIN = 700–799 (inclusive)
DMIN = BannerPrevious[
    BannerPrevious['Course Number {Crse Num}'].between(700, 799, inclusive='both')
].copy()


# In[31]:


OTST.head()


# In[32]:


# Ensure numeric comparison works: convert 'Course Number {Seq Crse Num}' to numeric
BannerPrevious['Course Number {Crse Num}'] = pd.to_numeric(
    BannerPrevious['Course Number {Crse Num}'], errors='coerce'
)

# Filter rows where:
# - Subject = "OTST" AND Course Number {Seq Crse Num} < 600
# OR
# - Subject = "ANEA" (any course number)

# Preview
print(f"✅ NTST DataFrame created with {len(NTST)} rows")
NTST.head(2)

# Preview
print(f"✅ OTST DataFrame created with {len(OTST)} rows")
print(f"   - OTST courses < 600: {len(OTST[OTST['Subject'] == 'OTST'])}")
print(f"   - ANEA courses (all): {len(OTST[OTST['Subject'] == 'ANEA'])}")
OTST.head(2)

# Preview
print(f"✅ CHIS DataFrame created with {len(CHIS)} rows")
CHIS.head(2)

# Preview
print(f"✅ DSLE DataFrame created with {len(DSLE)} rows")
DSLE.head(2)

# Preview
print(f"✅ GSEM DataFrame created with {len(GSEM)} rows")
GSEM.head(2)

# Preview
print(f"✅ MSSN DataFrame created with {len(MSSN)} rows")
MSSN.head(2)

# Preview
print(f"✅ PATH DataFrame created with {len(PATH)} rows")
PATH.head(2)

# Preview
print(f"✅ THST DataFrame created with {len(THST)} rows")
THST.head(2)

# Preview
print(f"✅ MDivHISP_MAPmENGL_MAPmHISP DataFrame created with {len(MDivHISP_MAPmENGL_MAPmHISP)} rows")
MDivHISP_MAPmENGL_MAPmHISP.head(2)

# Preview
print(f"✅ MA_Religion DataFrame created with {len(MA_Religion)} rows")
MA_Religion.head(2)

# Preview
print(f"✅ DMIN DataFrame created with {len(DMIN)} rows")
DMIN.head(2)


# ## Creating DropDown Menu (Identical to Step3!)

# In[33]:


import pandas as pd
import pprint  # pretty-print

# --- Step 1: Load the Excel file ---
# Replace 'your_file.xlsx' with your filename
DropDownMenu_dict = pd.read_excel(DropDown_RETRIEVAL, dtype=str)  # forces all columns to be read as strings
DropDownMenu_dict.head()


# In[34]:


# --- Build Instructor mapping (Name -> Email, ID) BEFORE unique lists are created ---
instr_name_col  = "Instructor Name {Instr Name}"
instr_email_col = "Instructor Email {Instr Email}"
instr_id_col    = "Instructor ID {Instr ID}"

InstructorMap_df = None
if all(c in DropDownMenu_dict.columns for c in [instr_name_col, instr_email_col, instr_id_col]):
    InstructorMap_df = (
        DropDownMenu_dict[[instr_name_col, instr_email_col, instr_id_col]]
        .copy()
    )

    # normalize
    for c in [instr_name_col, instr_email_col, instr_id_col]:
        InstructorMap_df[c] = InstructorMap_df[c].fillna("").astype(str).str.strip()

    # keep only rows that actually have a name
    InstructorMap_df = InstructorMap_df[InstructorMap_df[instr_name_col] != ""]

    # dedupe by name (keep first)
    InstructorMap_df = InstructorMap_df.drop_duplicates(subset=[instr_name_col], keep="first")



print("=== DEBUG: InstructorMap_df ===")
print("InstructorMap_df is None:", InstructorMap_df is None)
if InstructorMap_df is not None:
    print("InstructorMap_df rows:", len(InstructorMap_df))
    print(InstructorMap_df.head(10))
    # quick sanity check for empties
    empties = (InstructorMap_df[instr_email_col] == "").sum()
    print("Empty emails:", empties)
    empties_id = (InstructorMap_df[instr_id_col] == "").sum()
    print("Empty IDs:", empties_id)


# In[35]:


# --- Step 2: Create dictionaries ---
# Each column becomes a key; values are the unique entries (as strings)
DropDownMenu_dict = {
    col: DropDownMenu_dict[col].dropna().astype(str).unique().tolist()
    for col in DropDownMenu_dict.columns
}

# --- Step 3: (Optional) Display the result nicely ---
for key, values in DropDownMenu_dict.items():
    print(f"{key}: {values[:5]}{'...' if len(values) > 5 else ''}")


# ## Exporting the new DonSheet with separate Tabs/Sheets

# ### Building the Excel File Contents

# In[36]:


import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# NOTE: DataValidation no longer needed in Cell 1

# ===========================================
# CELL 1: Build workbook + helper sheets + formulas
#   - Writes DropDownMenu (hidden list source) ONLY
#   - Writes InstructorMap (hidden lookup table) ONLY (NOW from source_AutoFill_names.xlsx)
#   - Writes dependent autofill formulas (Email/ID)
#   - DOES NOT add DataValidation dropdowns (handled in Cell 2)
# ===========================================

# ------------------------------------------------------------
# InstructorMap SOURCE (NEW)
# ------------------------------------------------------------
AUTOFILL_NAMES_FILE = "0000_BuildingFiles/source_AutoFill_names.xlsx"
AUTOFILL_SHEET_NAME = "Sheet1"

# These are the column headers used in your exported sheets
instr_name_col  = "Instructor Name {Instr Name}"
instr_email_col = "Instructor Email {Instr Email}"
instr_id_col    = "Instructor ID {Instr ID}"

# ------------------------------------------------------------
# Load DropDown retrieval (unchanged)
# ------------------------------------------------------------
DropDown_RETRIEVAL_df = pd.read_excel(DropDown_RETRIEVAL, dtype=str)
DropDown_RETRIEVAL_df = DropDown_RETRIEVAL_df.apply(lambda s: s.fillna("").astype(str).str.strip())

# ------------------------------------------------------------
# Build InstructorMap_df from source_AutoFill_names.xlsx (NEW)
# ------------------------------------------------------------
InstructorMap_df = None
try:
    autofill_df = pd.read_excel(AUTOFILL_NAMES_FILE, sheet_name=AUTOFILL_SHEET_NAME, dtype=str)
    autofill_df = autofill_df.apply(lambda s: s.fillna("").astype(str).str.strip())

    # Expect these columns in the source file
    src_name_col = "Faculty Name"
    src_id_col   = "ID#"
    src_email_col = "email"

    missing = [c for c in [src_name_col, src_id_col, src_email_col] if c not in autofill_df.columns]
    if missing:
        raise ValueError(
            f"Missing required columns in {AUTOFILL_NAMES_FILE} / {AUTOFILL_SHEET_NAME}: {missing}"
        )

    InstructorMap_df = autofill_df[[src_name_col, src_email_col, src_id_col]].copy()

    # Rename to match your workbook headers
    InstructorMap_df.columns = [instr_name_col, instr_email_col, instr_id_col]

    # Keep only rows with a name
    InstructorMap_df = InstructorMap_df[InstructorMap_df[instr_name_col] != ""]

    # Deduplicate by name (first match wins)
    InstructorMap_df = InstructorMap_df.drop_duplicates(subset=[instr_name_col], keep="first")

except FileNotFoundError:
    raise FileNotFoundError(f"Could not find InstructorMap source file: {AUTOFILL_NAMES_FILE}")
except Exception as e:
    raise RuntimeError(f"Failed building InstructorMap_df from {AUTOFILL_NAMES_FILE}: {e}")

# ------------------------------------------------------------
# Build dropdown lists (unique values per column) -> for DropDownMenu sheet (unchanged)
# ------------------------------------------------------------
DropDownMenu_dict = {}
for col in DropDown_RETRIEVAL_df.columns:
    vals = (
        DropDown_RETRIEVAL_df[col]
        .loc[lambda s: s != ""]
        .drop_duplicates()
        .tolist()
    )
    DropDownMenu_dict[col] = vals

# ------------------------------------------------------------
# 1) Export main DataFrames (unchanged)
# ------------------------------------------------------------
with pd.ExcelWriter(New_SEM_course_updates, engine="openpyxl") as writer:
    CHIS.to_excel(writer, index=False, sheet_name="CHIS")
    DSLE.to_excel(writer, index=False, sheet_name="DSLE")
    GSEM.to_excel(writer, index=False, sheet_name="GSEM")
    MSSN.to_excel(writer, index=False, sheet_name="MSSN")
    NTST.to_excel(writer, index=False, sheet_name="NTST")
    OTST.to_excel(writer, index=False, sheet_name="OTST")
    PATH.to_excel(writer, index=False, sheet_name="PATH")
    THST.to_excel(writer, index=False, sheet_name="THST")
    MDivHISP_MAPmENGL_MAPmHISP.to_excel(writer, index=False, sheet_name="MDivHISP_MAPmENGL_MAPmHISP")
    MA_Religion.to_excel(writer, index=False, sheet_name="MA_Religion")
    DMIN.to_excel(writer, index=False, sheet_name="DMIN")

print(f"✅ Exported all DataFrames to {New_SEM_course_updates}")

wb = load_workbook(New_SEM_course_updates)

# ------------------------------------------------------------
# 2a) Create/populate DropDownMenu helper sheet (unchanged)
# ------------------------------------------------------------
if "DropDownMenu" in wb.sheetnames:
    ws_dd = wb["DropDownMenu"]
    ws_dd.delete_rows(1, ws_dd.max_row)
else:
    ws_dd = wb.create_sheet("DropDownMenu")

col_idx = 1
for header, values in DropDownMenu_dict.items():
    if not values:
        continue

    ws_dd.cell(row=1, column=col_idx, value=header)
    for row_idx, val in enumerate(values, start=2):
        ws_dd.cell(row=row_idx, column=col_idx, value=val)

    col_idx += 1

ws_dd.sheet_state = "hidden"

# ------------------------------------------------------------
# 2b) Create/populate InstructorMap helper sheet (NOW from AutoFill file)
# ------------------------------------------------------------
if InstructorMap_df is not None and not InstructorMap_df.empty:
    if "InstructorMap" in wb.sheetnames:
        ws_map = wb["InstructorMap"]
        ws_map.delete_rows(1, ws_map.max_row)
    else:
        ws_map = wb.create_sheet("InstructorMap")

    ws_map["A1"] = instr_name_col
    ws_map["B1"] = instr_email_col
    ws_map["C1"] = instr_id_col

    for r_idx, row in enumerate(InstructorMap_df.itertuples(index=False), start=2):
        ws_map.cell(r_idx, 1, row[0])
        ws_map.cell(r_idx, 2, row[1])
        ws_map.cell(r_idx, 3, row[2])

    ws_map.sheet_state = "hidden"
else:
    print("⚠️ InstructorMap_df was empty; InstructorMap sheet not populated.")

# ------------------------------------------------------------
# 2c) Add dependent autofill formulas (Email/ID) to exported sheets (unchanged)
# ------------------------------------------------------------
export_sheets = [
    "CHIS", "DSLE", "GSEM", "MSSN", "NTST", "OTST",
    "PATH", "THST", "MDivHISP_MAPmENGL_MAPmHISP", "MA_Religion", "DMIN",
]

for sname in export_sheets:
    sheet = wb[sname]
    headers = [cell.value for cell in sheet[1]]

    if instr_name_col in headers and "InstructorMap" in wb.sheetnames:
        name_col = headers.index(instr_name_col) + 1
        name_L = get_column_letter(name_col)

        if instr_email_col in headers:
            email_col = headers.index(instr_email_col) + 1
            email_L = get_column_letter(email_col)
            for r in range(2, min(sheet.max_row, 5000) + 1):
                sheet[f"{email_L}{r}"].value = (
                    f'=IF(${name_L}{r}="","",'
                    f'IFERROR(INDEX(InstructorMap!$B:$B, MATCH(${name_L}{r}, InstructorMap!$A:$A, 0)),""))'
                )

        if instr_id_col in headers:
            id_col = headers.index(instr_id_col) + 1
            id_L = get_column_letter(id_col)
            for r in range(2, min(sheet.max_row, 5000) + 1):
                sheet[f"{id_L}{r}"].value = (
                    f'=IF(${name_L}{r}="","",'
                    f'IFERROR(INDEX(InstructorMap!$C:$C, MATCH(${name_L}{r}, InstructorMap!$A:$A, 0)),""))'
                )

wb.save(New_SEM_course_updates)
wb.close()

print("✅ Cell 1 done: workbook built + DropDownMenu/InstructorMap saved (from source_AutoFill_names.xlsx) + formulas written (validations handled in Cell 2).")


# ### Caching the Excel Formula Outputs

# In[37]:


#import xlwings as xw
#
#with xw.App(visible=False) as app:
#    wb_xw = app.books.open(New_SEM_course_updates)
#    wb_xw.app.calculate()
#    wb_xw.save()
#    wb_xw.close()
#
#print("✅ Excel recalculated + saved (cached formula results should now exist).")
#
#for name in export_sheets:
#    ws = wb[name]
#    print(name, "data validations:", len(ws.data_validations.dataValidation))



# ### Applying distinct Formatting to the Excel File

# In[38]:


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string, quote_sheetname
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, date

# ===========================================
# CELL 2: Load workbook + formatting + grouping + autofit
#       + Re-apply dropdown validations (to restore dropdown arrows/UI)
#       + Force all NON-formula cells to be stored as text (strings)
# ===========================================

# Load the workbook we will WRITE to (single workbook; no cached-values workbook)
wb = load_workbook(New_SEM_course_updates, data_only=False)

# ------------------------------------------------------------
# Rebuild dropdown ranges from hidden DropDownMenu sheet
# ------------------------------------------------------------
def build_dropdown_ranges_from_dd_sheet(wb, dd_sheet_name="DropDownMenu"):
    """
    Rebuild dropdown_ranges and dropdown_values from the hidden DropDownMenu sheet.
    This makes Cell 2 independent from variables created in Cell 1.
    """
    if dd_sheet_name not in wb.sheetnames:
        return {}, {}

    ws_dd = wb[dd_sheet_name]
    dropdown_ranges = {}
    dropdown_values = {}

    for col_idx in range(1, ws_dd.max_column + 1):
        header = ws_dd.cell(row=1, column=col_idx).value
        if header is None or str(header).strip() == "":
            continue

        values = []
        for r in range(2, ws_dd.max_row + 1):
            v = ws_dd.cell(row=r, column=col_idx).value
            if v is None or str(v).strip() == "":
                continue
            values.append(str(v))

        if not values:
            continue

        col_letter = get_column_letter(col_idx)
        last_row = 1 + len(values)

        dropdown_ranges[header] = f"{quote_sheetname(dd_sheet_name)}!${col_letter}$2:${col_letter}${last_row}"
        dropdown_values[header] = values

    return dropdown_ranges, dropdown_values


def reapply_dropdown_validations(sheet, dropdown_ranges, dropdown_values, max_rows=5000):
    """
    Re-add DV rules for dropdown columns to refresh Excel UI dropdown arrows.
    """
    headers = [cell.value for cell in sheet[1]]

    for header, formula_range in dropdown_ranges.items():
        if header not in headers:
            continue

        cidx = headers.index(header) + 1
        cL = get_column_letter(cidx)

        dv = DataValidation(type="list", formula1=f"={formula_range}", allow_blank=True)

        # IMPORTANT: False = show dropdown arrow in Excel UI
        dv.showDropDown = False

        sheet.add_data_validation(dv)
        dv.add(f"${cL}$2:${cL}${max_rows}")

        # Optional: default blanks to "False" if False is an allowed option
        values_list = dropdown_values.get(header, [])
        if "False" in values_list:
            for r in range(2, min(sheet.max_row, max_rows) + 1):
                cell = sheet[f"{cL}{r}"]
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = "False"


dropdown_ranges, dropdown_values = build_dropdown_ranges_from_dd_sheet(wb, dd_sheet_name="DropDownMenu")

# ------------------------------------------------------------
# Force NON-formula cells to TEXT (strings) (rows 2..max_row)
# ------------------------------------------------------------
from datetime import datetime, date

def force_nonformula_cells_to_text(sheet, start_row=2, max_rows=5000, remove_leading_apostrophe=True):
    """
    Converts ALL non-formula cells to strings and sets number_format to Text ("@").
    ALSO sets number_format to Text ("@") for EMPTY cells, so Excel treats blanks as text-ready.
    Leaves formula cells untouched.
    """
    max_row = min(sheet.max_row or 1, max_rows)
    max_col = sheet.max_column or 1

    for row in sheet.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            v = cell.value

            # Skip formulas (even if they evaluate to blank)
            if cell.data_type == "f" or (isinstance(v, str) and v.startswith("=")):
                continue

            # Ensure Excel stores/behaves as Text even for blanks
            cell.number_format = "@"

            # If blank, we're done (keep it blank, just formatted as text)
            if v is None:
                continue

            # Normalize to string
            if isinstance(v, (datetime, date)):
                s = v.isoformat()
            else:
                s = str(v)

            # Remove a visible leading apostrophe created by a previous run
            if remove_leading_apostrophe and isinstance(s, str) and s.startswith("'"):
                s = s[1:]

            cell.value = s
            # IMPORTANT: do NOT set cell.quotePrefix = True


# ------------------------------------------------------------
# Grouping
# ------------------------------------------------------------
GROUP_DEFS = [
    ('Q', 'V'),
    ('X', 'Z'),
    ('AB', 'AH'),
    ('AJ', 'AW'),
    ('AY', 'BA'),
    ('BC', 'CV')
]

def group_columns(sheet, group_defs=GROUP_DEFS, outline_level=1):
    sheet.sheet_properties.outlinePr.summaryRight = False
    sheet.sheet_properties.outlinePr.summaryBelow = True
    sheet.sheet_properties.outlinePr.applyStyles = True

    max_col = sheet.max_column or 1
    for start_col, end_col in group_defs:
        start_idx = column_index_from_string(start_col)
        end_idx   = column_index_from_string(end_col)

        if start_idx > max_col:
            continue

        end_idx = min(end_idx, max_col)
        for col_idx in range(start_idx, end_idx + 1):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].outline_level = outline_level

# ------------------------------------------------------------
# Autofit widths (NO cached values)
# ------------------------------------------------------------
def autofit_columns_ignore_header(sheet, min_width=4, max_width=60, padding=2, max_rows=5000):
    """
    Autofit based on THIS workbook's cell values.
    NOTE: If a cell contains a formula, openpyxl sees the formula text (e.g. "=XLOOKUP(...)"),
    which can inflate widths. Use FIXED_WIDTHS to override those columns.
    """
    max_row = min(sheet.max_row or 1, max_rows)
    max_col = sheet.max_column or 1

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        for r in range(2, max_row + 1):
            v = sheet.cell(row=r, column=col_idx).value
            if v is None:
                continue
            max_length = max(max_length, len(str(v)))

        sheet.column_dimensions[col_letter].width = max(min_width, min(max_length + padding, max_width))

# ------------------------------------------------------------
# Fixed widths (override AFTER autofit)
# ------------------------------------------------------------
FIXED_WIDTHS = {
    "Notes {not in Banner}": 40,
    "Instructor Name {Instr Name}": 20,
    "Instructor Email {Instr Email}": 20,
    "Instructor ID {Instr ID}": 10,
    "account to be charged {not in Banner}": 20,
    "Semester Start Date {Soaterm Start Date}": 10, 
    "Pre-work Start Date {not in Banner}": 10, 
    "Pre-work End Date {not in Banner}": 10, 
    "Course Start Date {Meet Start Date}": 10, 
    "Course End Date {Meet End Date}": 10, 
    "Post-work Start Date {not in Banner}": 10, 
    "Post-work End Date {not in Banner}": 10, 
    "Semester End Date {Soaterm End Date}": 10, 
}

def apply_fixed_widths(sheet, fixed_widths=FIXED_WIDTHS):
    headers = [cell.value for cell in sheet[1]]
    for header, width in fixed_widths.items():
        if header in headers:
            col_idx = headers.index(header) + 1
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = width

# ------------------------------------------------------------
# Formatting
# ------------------------------------------------------------
thin_grey_border = Border(
    left=Side(style='thin', color="D3D3D3"),
    right=Side(style='thin', color="D3D3D3"),
    top=Side(style='thin', color="D3D3D3"),
    bottom=Side(style='thin', color="D3D3D3"),
)

COLOR_DATA_CELLS = False

BLUE_HEADERS = [
    "Don {not in Banner}",
    "Mismatching {not in Banner}",
    "Dept/Prog Admin {not in Banner}",
    "Mona {not in Banner}",
    "Karen {not in Banner}",
    "DONE {not in Banner}",
    "Notes {not in Banner}",
    "Subject",
    "Course Number {Crse Num}",
    "Course Section {Seq Crse Num}",
    "Program {not in Banner}",
    "Catalog Title",
    "Section Title",
    "Crosslist Details {not in Banner}",
    "Campus {Ssasect Campus}",
    "Campus Restriction {Camp Restr CC}",
    "Schedule Type {not in Banner}",
    "Instruction Method {Inst Method}",
    "Level {Level SC}",
    "Part-Of-Term",
    "Credits {Sect Crs}",
    "Enrollment Cap {Max Enrl}",
    "Meeting Type",
    "Pre-work Start Date {not in Banner}",
    "Pre-work End Date {not in Banner}",
    "Course Start Date {Meet Start Date}",
    "Course End Date {Meet End Date}",
    "Post-work Start Date {not in Banner}",
    "Post-work End Date {not in Banner}",
    "Course Beginning Time {Meet Beg Time}",
    "Course Ending Time {Meet End Time}",
    "SUN",
    "MON",
    "TUE",
    "WED",
    "THU",
    "FRI",
    "SAT",
    "Room {Meet Room}",
    "Building {Meet Bldg}",
    "Instructor Name {Instr Name}",
    "Instructor Email {Instr Email}",
    "Instructor ID {Instr ID}",
    "% Responsibility",
    "load/contract {not in Banner}",
    "costs per credit {not in Banner}",
    "total costs {not in Banner}",
    "account to be charged {not in Banner}",
]

GREY_HEADERS = [
    "CRN sorted {not in Banner}",
    "CRN",
    "X Lst",
    "Special Approval {Special Apvl}",
    "Credit Catalog {Cat Crs}",
    "Semester Start Date {Soaterm Start Date}",
    "Semester End Date {Soaterm End Date}",
    "Weeks",
    "Fee Amount {Fees Amt}",
    "Fee Code {Detl Code}",
    "Fee Type",
    "Fee Level",
    "Fee Ind",
    "Fee Term",
    "SEM Department {Scacrse Dept}",
    "Activity Date",
    "Status",
    "Bill Hrs",
    "Sect Sch Type",
    "Prerequisites {Preq Areas}",
    "Link Ind",
    "Integration Code",
    "Link Conn",
    "Crs Attr",
    "Comments",
    "Major Restr",
    "Scacrse College",
    "Enrolled",
    "Waitlist Capacity",
    "Primary Ind",
    "Override",
    "1st Date Reg Opens",
    "Last Date Reg Opens",
    "OL Range From Date",
    "OL Range To Date",
    "OL Numb Units",
    "OL Duration Code",
    "Grade Mode",
    "Gradable Ind",
    "Tuit Waiver",
    "Course Override {Meet Override}",
    "Lvl Res Ind",
    "Cmp Res Ind",
    "Rate Code",
    "Cohort Code",
    "St Attr Code",
    "Coll SC",
    "DegC SC",
    "Prog SC",
    "FOS Type",
    "Term Code",
    "Part-Of-Term",
    "Camp SC",
]

HEADER_DEFAULT_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
HEADER_BLUE_FILL    = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
HEADER_GREY_FILL    = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

HEADER_FILL_BY_NAME = {}
for h in BLUE_HEADERS:
    HEADER_FILL_BY_NAME[h] = HEADER_BLUE_FILL
for h in GREY_HEADERS:
    HEADER_FILL_BY_NAME[h] = HEADER_GREY_FILL


def format_sheet(sheet):
    headers = [cell.value for cell in sheet[1]]

    # Freeze panes
    if "Notes {not in Banner}" in headers:
        notes_idx = headers.index("Notes {not in Banner}") + 1
        sheet.freeze_panes = f"{get_column_letter(notes_idx + 1)}2"
    else:
        sheet.freeze_panes = "A2"

    # AutoFilter
    last_col = get_column_letter(sheet.max_column)
    sheet.auto_filter.ref = f"A1:{last_col}{sheet.max_row}"

    # Header style (base)
    for cell in sheet[1]:
        if cell.value is None or str(cell.value).strip() == "":
            continue

        header_text = str(cell.value).strip()

        cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color="FF0000" if header_text == "Notes {not in Banner}" else "000000",
        )

        cell.fill = HEADER_DEFAULT_FILL
        cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True, text_rotation=90)

    sheet.row_dimensions[1].height = 150

    # Apply explicit header color overrides
    for col_idx, header_val in enumerate(headers, start=1):
        if header_val is None:
            continue
        header_text = str(header_val).strip()
        fill = HEADER_FILL_BY_NAME.get(header_text)
        if fill is None:
            continue

        sheet.cell(row=1, column=col_idx).fill = fill

        if COLOR_DATA_CELLS:
            for r in range(2, sheet.max_row + 1):
                sheet.cell(row=r, column=col_idx).fill = fill

    # Grey-out inactive rows (Status == "I")
    if "Status" in headers:
        status_col_idx = headers.index("Status") + 1
        inactive_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for r in range(2, sheet.max_row + 1):
            v = sheet.cell(row=r, column=status_col_idx).value
            if str(v).strip().upper() == "I":
                for c in range(1, sheet.max_column + 1):
                    sheet.cell(row=r, column=c).fill = inactive_fill

    # Borders
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_grey_border

    # Fonts
    # 1) Global default: Arial 10
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            f = cell.font or Font()
            cell.font = Font(
                name="Arial",
                size=10,
                bold=f.bold,
                italic=f.italic,
                underline=f.underline,
                color=f.color,
            )

    # 2) Override columns A–G to Arial 10 (DATA ROWS ONLY)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=7):
        for cell in row:
            f = cell.font or Font()
            cell.font = Font(
                name="Arial",
                size=10,
                bold=f.bold,
                italic=f.italic,
                underline=f.underline,
                color=f.color,
            )

# ------------------------------------------------------------
# Apply formatting + grouping + autofit + fixed widths + dropdowns
# ------------------------------------------------------------
export_sheets = [
    "CHIS", "DSLE", "GSEM", "MSSN", "NTST", "OTST",
    "PATH", "THST", "MDivHISP_MAPmENGL_MAPmHISP", "MA_Religion", "DMIN",
]

for name in export_sheets:
    ws = wb[name]

    # ✅ 1) Force all NON-formula cells to be text (strings)
    force_nonformula_cells_to_text(ws, start_row=2, max_rows=5000)

    # 2) Normal formatting pipeline
    format_sheet(ws)
    group_columns(ws)
    autofit_columns_ignore_header(ws)  # no cached values

    # Override some columns to fixed widths (must be AFTER autofit)
    apply_fixed_widths(ws)

    # Re-apply dropdown validations to refresh dropdown arrows in Excel
    reapply_dropdown_validations(ws, dropdown_ranges, dropdown_values, max_rows=5000)

wb.save(New_SEM_course_updates)
wb.close()

print("✅ Cell 2 done: formatting + grouping + autofit + fixed widths + dropdown validations refreshed + non-formulas stored as text.")


# # Creating Executables

# ## Simple Course of action
# 1. First I ran this code:
# `jupyter nbconvert --to script /Users/oliverglanz/Library/CloudStorage/OneDrive-AndrewsUniversity/0000_EfficiencyWithIT/SmartScheduling/StepsFrom1to3_v20251030_working_executable.ipynb`
# 
# 2. this code:
# `python StepsFrom1to3_v20251030_working_executable.py`
# 
# 3. If the program should be selfcontained (all libraries packages within one executable) one has to continue:
# ` pip install pyinstaller      `
# 
# 4. finally: Now one has to zip/archive the folders and the py program into one file and send it to the secretaries
# 
# 
# ## Another option is to create a double-clickable version. 
# 1. For Windows do this:
# >>> create a bat file with this content and name it something like "RunSmartSchedule.bat":
# >>> 
# >>> ```python
# >>> @echo off
# >>> setlocal enabledelayedexpansion
# >>> title Smart Schedule - Step 1
# >>> echo =========================================
# >>> echo Starting Smart Schedule Step 1...
# >>> echo =========================================
# >>> echo.
# >>> 
# >>> REM === Navigate to the directory where this batch file is located ===
# >>> cd /d "%~dp0"
# >>> 
# >>> REM === Use flexible Anaconda Python path (works for any user) ===
# >>> set "PYTHON_PATH=%LOCALAPPDATA%\anaconda3\python.exe"
# >>> 
# >>> if not exist "%PYTHON_PATH%" (
# >>>     set "PYTHON_PATH=%LOCALAPPDATA%\Programs\anaconda3\python.exe"
# >>> )
# >>> 
# >>> if not exist "%PYTHON_PATH%" (
# >>>     set "PYTHON_PATH=%USERPROFILE%\anaconda3\python.exe"
# >>> )
# >>> 
# >>> if not exist "%PYTHON_PATH%" (
# >>>     set "PYTHON_PATH=%USERPROFILE%\Anaconda3\python.exe"
# >>> )
# >>> 
# >>> if not exist "%PYTHON_PATH%" (
# >>>     set "PYTHON_PATH=%ProgramData%\anaconda3\python.exe"
# >>> )
# >>> 
# >>> if not exist "%PYTHON_PATH%" (
# >>>     echo ERROR: Could not find Anaconda Python installation.
# >>>     echo Please install Anaconda for this program to run.
# >>>     echo Expected in one of:
# >>>     echo   %%LOCALAPPDATA%%\anaconda3\
# >>>     echo   %%LOCALAPPDATA%%\Programs\anaconda3\
# >>>     echo   %%USERPROFILE%%\anaconda3\
# >>>     echo   %%USERPROFILE%%\Anaconda3\
# >>>     echo   %%ProgramData%%\anaconda3\
# >>>     echo.
# >>>     pause
# >>>     exit /b 1
# >>> )
# >>> 
# >>> echo Using Python at: %PYTHON_PATH%
# >>> echo.
# >>> 
# >>> REM === Activate base Conda environment if possible ===
# >>> if exist "%LOCALAPPDATA%\anaconda3\Scripts\activate.bat" (
# >>>     echo Activating base Conda environment...
# >>>     call "%LOCALAPPDATA%\anaconda3\Scripts\activate.bat"
# >>> ) else if exist "%USERPROFILE%\anaconda3\Scripts\activate.bat" (
# >>>     echo Activating base Conda environment...
# >>>     call "%USERPROFILE%\anaconda3\Scripts\activate.bat"
# >>> ) else if exist "%ProgramData%\anaconda3\Scripts\activate.bat" (
# >>>     echo Activating base Conda environment...
# >>>     call "%ProgramData%\anaconda3\Scripts\activate.bat"
# >>> ) else (
# >>>     echo Proceeding without explicit environment activation.
# >>> )
# >>> 
# >>> echo.
# >>> echo Checking required Python packages...
# >>> echo.
# >>> 
# >>> REM === Check and install required packages ===
# >>> "%PYTHON_PATH%" -c "import openpyxl" 2>nul
# >>> if errorlevel 1 (
# >>>     echo Installing openpyxl...
# >>>     "%PYTHON_PATH%" -m pip install openpyxl
# >>> )
# >>> 
# >>> "%PYTHON_PATH%" -c "import xlrd" 2>nul
# >>> if errorlevel 1 (
# >>>     echo Installing xlrd...
# >>>     "%PYTHON_PATH%" -m pip install xlrd
# >>> )
# >>> 
# >>> echo.
# >>> echo All required packages are installed.
# >>> echo.
# >>> 
# >>> REM Run the Python script
# >>> "%PYTHON_PATH%" SmartSchedule_Step01_v20251031_working.py
# >>> 
# >>> echo.
# >>> echo =========================================
# >>> echo Script completed!
# >>> echo =========================================
# >>> pause
# >>> ```
# 
# 2. For MacOS create the file below and call it something like this "Run_SmartSchedule_Step01.command":
# 
# >>> create a bat file with this content and name it something like "RunSmartSchedule.command":
# >>> 
# >>> ```python
# >>> #!/bin/zsh
# >>> 
# >>> echo "========================================="
# >>> echo "Starting Smart Schedule Step 1..."
# >>> echo "========================================="
# >>> echo ""
# >>> 
# >>> # Navigate to the directory where this script is located
# >>> cd "$(dirname "$0")"
# >>> 
# >>> # Try to find Python in common Anaconda locations
# >>> PYTHON_PATH=""
# >>> 
# >>> if [ -f "$HOME/anaconda3/bin/python" ]; then
# >>>     PYTHON_PATH="$HOME/anaconda3/bin/python"
# >>> elif [ -f "/opt/anaconda3/bin/python" ]; then
# >>>     PYTHON_PATH="/opt/anaconda3/bin/python"
# >>> elif [ -f "$HOME/opt/anaconda3/bin/python" ]; then
# >>>     PYTHON_PATH="$HOME/opt/anaconda3/bin/python"
# >>> elif [ -f "/usr/local/anaconda3/bin/python" ]; then
# >>>     PYTHON_PATH="/usr/local/anaconda3/bin/python"
# >>> elif [ -f "/opt/anaconda3/bin/python" ]; then
# >>>     PYTHON_PATH="/opt/anaconda3/bin/python"
# >>> fi
# >>> 
# >>> if [ -z "$PYTHON_PATH" ]; then
# >>>     echo "ERROR: Could not find Anaconda Python installation"
# >>>     echo "Please check your Anaconda installation"
# >>>     read -p "Press Enter to exit..."
# >>>     exit 1
# >>> fi
# >>> 
# >>> echo "Using Python at: $PYTHON_PATH"
# >>> echo ""
# >>> 
# >>> # Run the Python script
# >>> "$PYTHON_PATH" SmartSchedule_Step01_v20251031_working.py
# >>>  
# >>> # Capture exit status
# >>> EXIT_CODE=$?
# >>> 
# >>> echo ""
# >>> echo "========================================="
# >>> if [ $EXIT_CODE -eq 0 ]; then
# >>>     echo "Script completed successfully!"
# >>> else
# >>>     echo "Script failed with error code: $EXIT_CODE"
# >>> fi
# >>> echo "========================================="
# >>> echo ""
# >>> read -p "Press Enter to exit..."
# >>> 
# >>> ```
# >> After that it is important to run the command ```chmod +x Run_SmartSchedule_Step01.command``` before sharing the file with others
# 
# 3. Now one has to zip/archive the folders, the py program, and the bat/command file into one file and send it to the secretaries

# 

# 
