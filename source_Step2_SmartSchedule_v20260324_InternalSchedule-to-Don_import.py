#!/usr/bin/env python
# coding: utf-8

# In[1]:


#Initializing essential packages

import re
import os
import glob
import math
from datetime import datetime
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from collections import defaultdict


# In[2]:


#Definine the current directory in which the code is running

print("Current directory:", os.getcwd())


# In[3]:


# File Paths
## Input File: Downloaded Seminary Smart Schedule (supports CSV, XLS, XLSX)
smartschudulefile = glob.glob('Step2_input_downloaded_Seminary_Smart_Schedule/*.csv') + glob.glob('Step2_input_downloaded_Seminary_Smart_Schedule/*.xls') + glob.glob('Step2_input_downloaded_Seminary_Smart_Schedule/*.xlsx')
if not smartschudulefile:
    raise FileNotFoundError("No CSV, XLS, or XLSX files found in Step2_input_downloaded_Seminary_Smart_Schedule directory")
if '~$' in smartschudulefile[0]:
    raise RuntimeError("Save and close CSV, XLS, or XLSX file in Step2_input_downloaded_Seminary_Smart_Schedule directory before continue")
smartschudulefile_INPUT = smartschudulefile[0]
print(f"Input file: {smartschudulefile_INPUT}")

## Input File: Class information update
Bannerclassesfile = glob.glob('x_SEM_CourseUpdate_LiveFolder_office365/*.csv') + glob.glob('x_SEM_CourseUpdate_LiveFolder_office365/*.xls') + glob.glob('x_SEM_CourseUpdate_LiveFolder_office365/*.xlsx')
if not Bannerclassesfile:
    raise FileNotFoundError("No CSV, XLS, or XLSX files found in Step2_input_course_updates directory")
if '~$' in Bannerclassesfile[0]:
    raise RuntimeError("Save and close CSV, XLS, or XLSX file in Step2_input_course_updates directory directory before continue")
Bannerclassesfile_INPUT = Bannerclassesfile[0]
print(f"Input file: {Bannerclassesfile_INPUT}")

## Output File
# it is assumed that the output file will be same as the input file,
# therefore, the code replace the old file with the new one
# there is another option which is considering a version of the code, which is commented below
# version= f'_v' + str(datetime.now().strftime("%Y%m%d"))
smartschudulefile_OUTPUT = Bannerclassesfile_INPUT
print(f"Output file: {smartschudulefile_OUTPUT}")


# In[4]:


# using first three sheets of the input file in the following order
# 0 - semester schedule
# 1 - synchronous classes
# 2 - asynchronous classes

all_sheets = pd.read_excel(smartschudulefile_INPUT, sheet_name=[0, 1, 2], header=None)

# saving the pandas dataframe of sheets as a dictionary
# empty cells are filled with 'NA' values

all_sheets = {
    name: df.fillna('NA') 
    for name, df in all_sheets.items()
}

# testing if the name of sheet 1 and sheet 2 correspond to synchronous and asynchronous classes, respectively
xls = pd.ExcelFile(smartschudulefile_INPUT)
sheet_names = xls.sheet_names
if 'SYNC' not in sheet_names[1] or 'ASYNC' not in sheet_names[2]:
    msg = (
        f"Rearange or rename properly the order of sheets in input file:{smartschudulefile_INPUT}\n"
        f"Expected: semester schedule, synchronous classes (SYNC), asynchronous classes (ASYNC)\n"
        f"Found: {sheet_names[0]}, {sheet_names[1]}, {sheet_names[2]}"
    )
    raise RuntimeError(msg)


# In[5]:


# cleaner function for the cells
# it transforms different types of data in string,
# and if a word stars or end with an empty space, it removes that
# this step is necessary to avoid errors due to different datatype from the input sheet

def cleaner(value):
    if type(value) != str:
        value = str(value)

    if len(value) > 1:    
        if value[0] == ' ':
            value = value[1:]

        if value[-1] == ' ':
            value = value.strip()

    return value


# In[6]:


# the first part of the code works with the first sheet (semester schedule) with the steps:
# 1 - discovering the gap columns (columns in which all the cells are empty)
# 2 - run a loop from each large portion of the sheet (range of columns)
# 3 - in the first range of blocks, it is defined the semester value and when it start and ends
# the information about the semester is important, because it defines what is the semester
# the code should use in the other sheets (synchronous classes, and asynchronous classes)
# 4 - the sheet is mapped into a matrix
# 5 - for each column, it is defined the type of algorithm to gather information
# 6 - depending on the type of algorithm, the code does a loop for each column and gathers the information for the classes


# In[7]:


exclude_time = ['Tues. - 11:30 CHAPEL','Lunch 12:30 - 1:20 PM','M, W, R - Lunch 11:30 - 1:20, T - 12:30-1:30', 'NA'] # excluded cells
department = ['CHIS', 'DSLE', 'GSEM', 'MSSN', 'NTST', 'OTST', 'PATH', 'THST', 'ANEA'] # definition of the department names
room = {1: 'N 108', 2: 'N 110', 3: 'N 135', 4: 'N 150', 5: 'S 115', 6: 'S 120', 7: 'N215', 8: 'N 235', 9: 'N 310', 10: 'N 335', 11: 'S 340'} # definition of rooms

df=all_sheets[0]

data = {}
gapcolumns = df.columns[(df == 'NA').all()].tolist() # discovering the gap columns (colums in which all the cells are empty)
initialcolumn = 0

for gapcolumn in gapcolumns:

    matrix = {}
    df1 = df.iloc[:, initialcolumn:gapcolumn]

    # defining the column range
    column_range = range(initialcolumn, gapcolumn)

    if initialcolumn == 0:
        ## defining when the semester starts and ends
        # semester starts
        try:
            semesterstart = datetime.strptime(
                str(df.iloc[3, gapcolumn-1]), '%Y-%m-%d %H:%M:%S'
            ).strftime('%d-%b-%y').upper()
        except Exception:
            semesterstart = '####'

        # semester ends
        try:
            semesterend = datetime.strptime(
                str(df.iloc[4, gapcolumn-1]), '%Y-%m-%d %H:%M:%S'
            ).strftime('%d-%b-%y').upper()
        except Exception:
            semesterend = '####'

        code='####'

        if semesterstart != '####':
            datestart = semesterstart.split('-')
            if datestart[1] in ['JAN', 'FEB']:
                code = '21' #it means Spring semester
                summersemester = False
            elif datestart[1] in ['APR', 'MAY']:
                code = '31' #it means Summer semester
                summersemester = True
            elif datestart[1] in ['AUG', 'SEP']:
                code = '41' #it means Fall semester
                summersemester = False

            if code != '####':
                termcode = '20' + datestart[2]+ code
            else:
                termcode = '####'
        else:
            termcode = '####'

    # indicating that the chosen semester
    chosensemester = termcode

    # droping the header of the dataframe
    df1 = df1[6:].reset_index(drop=True)

    # mapping the cells in the column range to a matrix

    for i in column_range:
        line=0
        for column in df1[i]:
            if column and column not in exclude_time:
                matrix[(line,i)] = cleaner(column)
            line=line+1

    # discovering gap rows to differentiate two algorithm methods
    gaprows = (df1 == 'NA').all(axis=1)
    cut = gaprows & gaprows.shift(-1) # it is considered a cut row when there are two empty rows in the column range

    # saving information of the algorithm method for each column
    mode = {}
    if cut.any():
        cut_row = cut.index[cut].min()
        cut_index_label = cut.index[cut].min()
        cut_iloc = df1.index.get_loc(cut_index_label)

        for i, idx in enumerate(df1.index):
            if i < cut_iloc:
                mode[idx] = 'column'
            else:
                mode[idx] = 'row'
    else:
        cut_row = ''
        mode = {idx: 'column' for idx in df1.index}

    current_time_key = None
    class_count = 0
    timekey = {}
    termkey = {}
    column_old = None
    term_old = None

    for i in column_range:
        line = 0

        if mode[line] == 'column':

            for column in df1[i]:
                repeatedclass = False
                cl = '####'
                section = '####'
                credit = '####'
                crn = '####'
                cap = '####'
                day = '####'
                hour = '####'
                note = '####'

                column = cleaner(column)

                if i == 0:

                    # saving information for time frame of the class as indicated in the first column of the column range
                    if column != column_old and column not in exclude_time:
                        timekey[line] = column
                        column_old = column
                    else:
                        timekey[line] = column_old

                    if 'Term' in column:
                        summersemester = True
                        match = re.search(r'(\b[A-Za-z]{3,}\s+\d{1,2}\s*)(?: To | to )(\b[A-Za-z]{3,}\s+\d{1,2}\b)', column)

                        if match:
                            initial_date = datetime.strptime(match.group(1), '%B %d')
                            classstart = initial_date.strftime('%d-%b').upper() + '-' + semesterstart[-2:]

                            final_date = datetime.strptime(match.group(2), '%B %d')
                            classend = final_date.strftime('%d-%b').upper() + '-' + semesterstart[-2:]
                        else:
                            classstart = '####'
                            classend = '####'

                        termkey[line] = [classstart,classend]
                        term_old = [classstart,classend]

                    else:
                        if summersemester == False:
                            termkey[line] = [semesterstart,semesterend]
                        else:
                            termkey[line] = term_old

                if column and column not in exclude_time: # using only cells with values and that are not in the excluded cells dictionary

                    # saving cells that start with parentesis as comments (notes)
                    if column[0] == '(':
                        class_data = {'note': column
                                     }
                        class_list = data.setdefault(timekey[line], [])
                        class_list.append(class_data)

                    if len(column) > 4: # considering cells with more text to avoid cells with only the name of the department
                        if column[:4] in department and column[4].isnumeric(): # considering cells that the name belongs to
                                                                               # the department dictionary and the next character is numeric (e.g., NTST500-001)

                            # code for cross-listed classes
                            if '/' in column:
                                cl_dict = {}
                                classes = column.split('/')
                                for item in enumerate(classes):
                                    cl_dict[item[1]] = item[0]

                                crosslist = "/".join(classes)

                                for cls in classes:
                                    match = re.search(r'([A-Z]{3,}\d{3})-(\d{3})', cls)
                                    if match:
                                        cl = match.group(1).strip()
                                        section = match.group(2).strip()
                                    else:
                                        cl = '####'
                                        section = '####'

                                    # checking if the class information was already presented in the sheet
                                    for item in data:
                                        for newitem in data[item]:
                                            if 'class' in newitem.keys():
                                                 if newitem['class'] == cl and newitem['section'] == section:
                                                    repeatedclass = True
                                                    break
                                                 else:
                                                    repeatedclass = False

                                    # saving only information for the classes one time
                                    if repeatedclass == False:

                                        # class information are obtained by the matrix
                                        if (line+2,i) in matrix:
                                            instructor = matrix[(line+2,i)]

                                        if (line+3+cl_dict[cls],i) in matrix:
                                            match = re.search(r'^(?:.*?[-\s])?([\d\-\,]+)\s*Cr\.?\s*\(CRN\s*([\w#]+)\)\s*(\d+|#+)\s*St\.?\s*$', matrix[(line+3+cl_dict[cls],i)])
                                            if match:
                                                credit = match.group(1).strip()
                                                crn = match.group(2).strip()
                                                cap = match.group(3).strip()
                                            else:
                                                credit = '####'
                                                crn = '####'
                                                cap = '####'

                                        if (line+5+len(cl_dict)-1, i) in matrix:                
                                            note = matrix[(line+5+len(cl_dict)-1, i)]

                                        if (line+4+len(cl_dict)-1,i) in matrix:
                                            # considering classes with more than one day of class
                                            if '/' in matrix[(line+4+len(cl_dict)-1,i)]:
                                                dayhour1, dayhour2 = matrix[(line+4+len(cl_dict)-1,i)].split('/')

                                                day1, hour1 = dayhour1.split(' ', 1)
                                                day2, hour2 = dayhour2.split(' ', 1)

                                                class_data = {'class': cl,
                                                      'section': section,
                                                      'title': matrix[(line + 1, i)],
                                                      'instructor': instructor,
                                                      'cap': cap,
                                                      'credit': credit,
                                                      'crn': crn,
                                                      'day': day1,
                                                      'hour': hour1,
                                                      'room': room[i],
                                                      'note': note,
                                                      'semesterstart' : semesterstart,
                                                      'semesterend' : semesterend,
                                                      'termcode' : termcode,
                                                      'crosslist' : crosslist,
                                                      'classstart': termkey[line][0],
                                                      'classend': termkey[line][1]
                                                     }
                                                class_list = data.setdefault(timekey[line], [])
                                                class_list.append(class_data)

                                                class_data = {'class': cl,
                                                      'section': section,
                                                      'title': matrix[(line + 1, i)],
                                                      'instructor': instructor,
                                                      'cap': cap,
                                                      'credit': credit,
                                                      'crn': crn,
                                                      'day': day2,
                                                      'hour': hour2,
                                                      'room': room[i],
                                                      'note': note,
                                                      'semesterstart' : semesterstart,
                                                      'semesterend' : semesterend,
                                                      'termcode' : termcode,
                                                      'crosslist' : crosslist,
                                                      'classstart': termkey[line][0],
                                                      'classend': termkey[line][1]
                                                     }
                                                class_list = data.setdefault(timekey[line], [])
                                                class_list.append(class_data)


                                            else:
                                                if ': ' in matrix[(line+4+len(cl_dict)-1,i)]:
                                                    day, hour = matrix[(line+4+len(cl_dict)-1,i)].split(': ', 1)
                                                else:
                                                    day, hour = matrix[(line+4+len(cl_dict)-1,i)].split(' ', 1)

                                                class_data = {'class': cl,
                                                            'section': section,
                                                            'title': matrix[(line + 1, i)],
                                                            'instructor': instructor,
                                                            'cap': cap,
                                                            'credit': credit,
                                                            'crn': crn,
                                                            'day': day,
                                                            'hour': hour,
                                                            'room': room[i],
                                                            'note': note,
                                                            'semesterstart' : semesterstart,
                                                            'semesterend' : semesterend,
                                                            'termcode' : termcode,
                                                            'crosslist' : crosslist,
                                                            'classstart': termkey[line][0],
                                                            'classend': termkey[line][1]
                                                            }
                                                class_list = data.setdefault(timekey[line], [])
                                                class_list.append(class_data)

                            else:

                                # checking if the class information was already presented in the sheet
                                match = re.search(r'([A-Z]{3,}\d{3})(?:-(\d{3}))?', column)
                                if match:
                                    cl = (match.group(1) or '####').strip()
                                    section = (match.group(2) or '####').strip() 
                                else:
                                    cl = '####'
                                    section = '####'

                                repeatedclass = False

                                for item in data:
                                    for newitem in data[item]:
                                        if 'class' in newitem:
                                            if newitem['class'] == cl and newitem['section'] == section:
                                                repeatedclass = True
                                                break

                                    if repeatedclass: 
                                        break

                                # saving only information for the classes one time
                                if repeatedclass == False:

                                    # class information are obtained by the matrix
                                    if (line+2,i) in matrix:
                                        instructor = matrix[(line+2,i)]

                                    if (line+3,i) in matrix:
                                        match = re.search(r'^(?:.*?[-\s])?([\d\-\,]+)\s*Cr\.?\s*\(CRN\s*([\w#]+)\)\s*(\d+|#+)\s*St\.?\s*$', matrix[(line+3,i)])
                                        if match:
                                            credit = match.group(1).strip()
                                            crn = match.group(2).strip()
                                            cap = match.group(3).strip()
                                        else:
                                            credit = '####'
                                            crn = '####'
                                            cap = '####'

                                    if (line+5,i) in matrix:
                                        note = matrix[(line+5,i)]

                                    if (line+4,i) in matrix:
                                        # considering classes with more than one day of class
                                        if '/' in matrix[(line+4,i)]:
                                            dayhour1, dayhour2 = matrix[(line+4,i)].split('/')

                                            day1, hour1 = dayhour1.split(' ', 1)
                                            day2, hour2 = dayhour2.split(' ', 1)

                                            class_data = {'class': column[:7],
                                                  'section': column[-3:],
                                                  'title': matrix[(line + 1, i)],
                                                  'instructor': instructor,
                                                  'cap': cap,
                                                  'credit': credit,
                                                  'crn': crn,
                                                  'day': day1,
                                                  'hour': hour1,
                                                  'note': note,
                                                  'room': room[i],
                                                  'semesterstart' : semesterstart,
                                                  'semesterend' : semesterend,
                                                  'termcode' : termcode,
                                                  'classstart': termkey[line][0],
                                                  'classend': termkey[line][1]
                                                 }
                                            class_list = data.setdefault(timekey[line], [])
                                            class_list.append(class_data)

                                            class_data = {'class': column[:7],
                                                  'section': column[-3:],
                                                  'title': matrix[(line + 1, i)],
                                                  'instructor': instructor,
                                                  'cap': cap,
                                                  'credit': credit,
                                                  'crn': crn,
                                                  'day': day2,
                                                  'hour': hour2,
                                                  'note': note,
                                                  'room': room[i],
                                                  'semesterstart' : semesterstart,
                                                  'semesterend' : semesterend,
                                                  'termcode' : termcode,
                                                  'classstart': termkey[line][0],
                                                  'classend': termkey[line][1]
                                                 }
                                            class_list = data.setdefault(timekey[line], [])
                                            class_list.append(class_data)

                                        else:
                                            if ': ' in matrix[(line+4,i)]:

                                                day, hour = matrix[(line+4,i)].split(': ', 1)
                                            else:
                                                day, hour = matrix[(line+4,i)].split(' ', 1)

                                            class_data = {'class': column[:7],
                                                          'section': column[-3:],
                                                          'title': matrix[(line + 1, i)],
                                                          'instructor': instructor,
                                                          'cap': cap,
                                                          'credit': credit,
                                                          'crn': crn,
                                                          'day': day,
                                                          'hour': hour,
                                                          'note': note,
                                                          'room': room[i],
                                                          'semesterstart' : semesterstart,
                                                          'semesterend' : semesterend,
                                                          'termcode' : termcode,
                                                          'classstart': termkey[line][0],
                                                          'classend': termkey[line][1]
                                                         }
                                            class_list = data.setdefault(timekey[line], [])
                                            class_list.append(class_data)

                line = line + 1

    initialcolumn = gapcolumn + 1

####################################################################################################################################################################
# last range of column for the semester schedule have the information for arranged classes
# this information is displayed using rows, therefore, it is used a different algorithm
# it reads information by column, but access and save that information with is stored in rows

if initialcolumn < len(df.columns):
    df1 = df.iloc[:, initialcolumn:len(df.columns)]
    df1 = df1[6:].reset_index(drop=True)
    df1.columns = range(len(df1.columns))
    data['Arranged Courses'] = []

    for i in range(0, len(df1)):
        cls = df1.iloc[i]
        day = '####'
        hour = '####'

        if cls[0][:4] in department and cls[0][5].isnumeric():
            if '-' in cls[0]:
                cl, section = cleaner(cls[0]).strip().split('-')
            else:
                cl = cleaner(cls[0]).strip()

            crn = '####' if cleaner(cls[1]) == 'NA' else cleaner(cls[1])

            title = '####' if cleaner(cls[2]) == 'NA' else cleaner(cls[2])

            note = '####' if cleaner(cls[3]) == 'NA' else cleaner(cls[3])

            credit = '####' if cleaner(cls[4]) == 'NA' else cleaner(cls[4])

            instructor = '####' if cleaner(cls[5]) == 'NA' else cleaner(cls[5])

            cap = '####' if cleaner(cls[6]) == 'NA' else cleaner(cls[6])

            if cleaner(cls[7]) == 'NA':
                day = '####'
            else:
                if '-' in cleaner(cls[7]):
                    day = cleaner(cls[7])
                    hour = '####'
                else:
                    day = '####'
                    hour = '####'

            if cleaner(cls[8]) == 'NA':
                room = '####'
            else:
                room = cls[8]

            class_data = {'class': cl,
                'section': section,
                'title': title,
                'instructor': instructor,
                'cap': cap,
                'credit': credit,
                'crn': crn,
                'day': day,
                'hour': hour,
                'note': note,
                'room' : room,
                'semesterstart' : semesterstart,
                'semesterend' : semesterend,
                'termcode' : termcode
                }
            class_list = data.setdefault('Arranged Courses', [])
            class_list.append(class_data)


# In[8]:


# the second part of the code works with the second sheet (synchronous classes)
# it follows the same logic of the first part for the first sheet


# In[9]:


exclude_time = ['Tues. - 11:30 CHAPEL','Lunch 12:30 - 1:20 PM','M, W, R - Lunch 11:30 - 1:20, T - 12:30-1:30', 'NA'] # excluded cells
department = ['CHIS', 'DSLE', 'GSEM', 'MSSN', 'NTST', 'OTST', 'PATH', 'THST', 'ANEA'] # definition of the department names

df=all_sheets[1]

gapcolumns = df.columns[(df == 'NA').all()].tolist() # discovering the gap columns (colums in which all the cells are empty)
initialcolumn = 0

for gapcolumn in gapcolumns:

    matrix = {}
    df1 = df.iloc[:, initialcolumn:gapcolumn]

    # defining the column range
    column_range = range(initialcolumn, gapcolumn)

    ## defining when the semester starts and ends
    # semester starts
    try:
        semesterstart = datetime.strptime(
            str(df.iloc[3, gapcolumn-1]), '%Y-%m-%d %H:%M:%S'
        ).strftime('%d-%b-%y').upper()
    except Exception:
        semesterstart = '####'

    # semester ends
    try:
        semesterend = datetime.strptime(
            str(df.iloc[4, gapcolumn-1]), '%Y-%m-%d %H:%M:%S'
        ).strftime('%d-%b-%y').upper()
    except Exception:
        semesterend = '####'

    code='####'

    if semesterstart != '####':
        datestart = semesterstart.split('-')
        if datestart[1] in ['JAN', 'FEB']:
            code = '21' #it means Spring semester
        elif datestart[1] in ['APR', 'MAY']:
            code = '31' #it means Summer semester
        elif datestart[1] in ['AUG', 'SEP']:
            code = '41' #it means Fall semester

        if code != '####':
            termcode = '20' + datestart[2]+ code
        else:
            termcode = '####'
    else:
        termcode = '####'

    # working only which the chosen semester from the first sheet
    if termcode == chosensemester:

        # droping the header of the dataframe
        df1 = df1[6:].reset_index(drop=True)

        # mapping the cells in the column range to a matrix
        for i in column_range:
            line=0
            for column in df1[i]:
                if column and column not in exclude_time:
                    matrix[(line,i)] = cleaner(column)
                line=line+1

        # discovering gap rows to differentiate two algorithm methods
        gaprows = (df1 == 'NA').all(axis=1)
        cut = gaprows & gaprows.shift(-1) # it is considered a cut row when there are two empty rows in the column range

        # saving information of the algorithm method for each column
        mode = {}
        if cut.any():
            cut_row = cut.index[cut].min()
            cut_index_label = cut.index[cut].min()
            cut_iloc = df1.index.get_loc(cut_index_label)

            for i, idx in enumerate(df1.index):
                if i < cut_iloc:
                    mode[idx] = 'column'
                else:
                    mode[idx] = 'row'
        else:
            cut_row = ''
            mode = {idx: 'column' for idx in df1.index}

        current_time_key = None
        class_count = 0
        timekey = {}
        termkey = {}
        column_old = None
        term_old = None
        rowlines = {}

        for i in column_range:
            line = 0

            for column in df1[i]:
                cl = '####'
                section = '####'
                credit = '####'
                crn = '####'
                cap = '####'
                day = '####'
                hour = '####'
                note = '####'

                column = cleaner(column)

                if mode[line] == 'column':

                    if i == 0:
                        # saving information for time frame of the class as indicated in the first column of the column range
                        if column != column_old and column not in exclude_time:
                            timekey[line] = column
                            column_old = column
                        else:
                            timekey[line] = column_old

                        if 'Term' in column:
                            match = re.search(r'(\b[A-Za-z]{3,}\s+\d{1,2}\s*)(?: To | to )(\b[A-Za-z]{3,}\s+\d{1,2}\b)', column)

                            if match:
                                initial_date = datetime.strptime(match.group(1), '%B %d')
                                classstart = initial_date.strftime('%d-%b').upper() + '-' + semesterstart[-2:]

                                final_date = datetime.strptime(match.group(2), '%B %d')
                                classend = final_date.strftime('%d-%b').upper() + '-' + semesterstart[-2:]
                            else:
                                classstart = '####'
                                classend = '####'

                            termkey[line] = [classstart,classend]
                            term_old = [classstart,classend]

                        else:
                            if summersemester == False:
                                termkey[line] = [semesterstart,semesterend]
                            else:
                                termkey[line] = term_old

                    if column and column not in exclude_time: # using only cells with values and that are not in the excluded cells dictionary

                        # saving cells that start with parentesis as comments (notes)
                        if column[0] == '(':
                            class_data = {'note': column
                                            }
                            class_list = data.setdefault(timekey[line], [])
                            class_list.append(class_data)

                        if len(column) > 4: # considering cells with more text to avoid cells with only the name of the department
                            if column[:4] in department and column[4].isnumeric(): # considering cells that the name belongs to
                                                                                   # the department dictionary and the next character is numeric (e.g., NTST500-001)

                                # code for cross-listed classes
                                if '/' in column:
                                    cl_dict = {}
                                    classes = column.split('/')
                                    for item in enumerate(classes):
                                        cl_dict[item[1]] = item[0]

                                    crosslist = "/".join(classes)

                                    for cls in classes:
                                        match = re.search(r'([A-Z]{3,}\d{3})-(\d{3})', cls)
                                        if match:
                                            cl = match.group(1).strip()
                                            section = match.group(2).strip()
                                        else:
                                            cl = '####'
                                            section = '####'

                                        # checking if the class information was already presented in the sheet
                                        for item in data:
                                            for newitem in data[item]:
                                                if 'class' in newitem.keys():
                                                    if newitem['class'] == cl and newitem['section'] == section:
                                                        repeatedclass = True
                                                        break

                                        # saving only information for the classes one time
                                        if repeatedclass == False:

                                            # class information are obtained by the matrix
                                            if (line+2,i) in matrix:
                                                instructor = matrix[(line+2,i)]

                                            if (line+3+cl_dict[cls],i) in matrix:
                                                match = re.search(r'^(?:.*?[-\s])?([\d\-\,]+)\s*Cr\.?\s*\(CRN\s*([\w#]+)\)\s*(\d+|#+)\s*St\.?\s*$', matrix[(line+3+cl_dict[cls],i)])
                                                if match:
                                                    credit = match.group(1).strip()
                                                    crn = match.group(2).strip()
                                                    cap = match.group(3).strip()
                                                else:
                                                    credit = '####'
                                                    crn = '####'
                                                    cap = '####'

                                            if (line+5+len(cl_dict)-1, i) in matrix:                
                                                note = matrix[(line+5+len(cl_dict)-1, i)]

                                            if (line+4+len(cl_dict)-1,i) in matrix:
                                                # considering classes with more than one day of class
                                                if '/' in matrix[(line+4+len(cl_dict)-1,i)]:
                                                    dayhour1, dayhour2 = matrix[(line+4+len(cl_dict)-1,i)].split('/')

                                                    day1, hour1 = dayhour1.split(' ', 1)
                                                    day2, hour2 = dayhour2.split(' ', 1)

                                                    class_data = {'class': cl,
                                                            'section': section,
                                                            'title': matrix[(line + 1, i)],
                                                            'instructor': instructor,
                                                            'cap': cap,
                                                            'credit': credit,
                                                            'crn': crn,
                                                            'day': day1,
                                                            'hour': hour1,
                                                            'note': note,
                                                            'semesterstart' : semesterstart,
                                                            'semesterend' : semesterend,
                                                            'termcode' : termcode,
                                                            'crosslist' : crosslist,
                                                            'classstart': termkey[line][0],
                                                            'classend': termkey[line][1]
                                                            }
                                                    class_list = data.setdefault(timekey[line], [])
                                                    class_list.append(class_data)

                                                    class_data = {'class': cl,
                                                            'section': section,
                                                            'title': matrix[(line + 1, i)],
                                                            'instructor': instructor,
                                                            'cap': cap,
                                                            'credit': credit,
                                                            'crn': crn,
                                                            'day': day2,
                                                            'hour': hour2,
                                                            'note': note,
                                                            'semesterstart' : semesterstart,
                                                            'semesterend' : semesterend,
                                                            'termcode' : termcode,
                                                            'crosslist' : crosslist,
                                                            'classstart': termkey[line][0],
                                                            'classend': termkey[line][1]
                                                            }
                                                    class_list = data.setdefault(timekey[line], [])
                                                    class_list.append(class_data)


                                                else:
                                                    if ': ' in matrix[(line+4+len(cl_dict)-1,i)]:
                                                        day, hour = matrix[(line+4+len(cl_dict)-1,i)].split(': ', 1)
                                                    else:
                                                        day, hour = matrix[(line+4+len(cl_dict)-1,i)].split(' ', 1)

                                                    class_data = {'class': cl,
                                                                'section': section,
                                                                'title': matrix[(line + 1, i)],
                                                                'instructor': instructor,
                                                                'cap': cap,
                                                                'credit': credit,
                                                                'crn': crn,
                                                                'day': day,
                                                                'hour': hour,
                                                                'note': note,
                                                                'semesterstart' : semesterstart,
                                                                'semesterend' : semesterend,
                                                                'termcode' : termcode,
                                                                'crosslist' : crosslist,
                                                                'classstart': termkey[line][0],
                                                                'classend': termkey[line][1]
                                                                }
                                                    class_list = data.setdefault(timekey[line], [])
                                                    class_list.append(class_data)

                                else:

                                    # checking if the class information was already presented in the sheet
                                    match = re.search(r'([A-Z]{3,}\d{3})(?:-(\d{3}))?', column)
                                    if match:
                                        cl = (match.group(1) or '####').strip()
                                        section = (match.group(2) or '####').strip() 
                                    else:
                                        cl = '####'
                                        section = '####'

                                    repeatedclass = False

                                    for item in data:
                                        for newitem in data[item]:
                                            if 'class' in newitem:
                                                if newitem['class'] == cl and newitem['section'] == section:
                                                    repeatedclass = True
                                                    break

                                        if repeatedclass: 
                                            break

                                    # saving only information for the classes one time
                                    if repeatedclass == False:

                                        match = re.search(r'([A-Z]{3,}\d{3})(?:-(\d{3}))?', column)
                                        if match:
                                            cl = (match.group(1) or '####').strip()
                                            section = (match.group(2) or '####').strip() 
                                        else:
                                            cl = '####'
                                            section = '####'

                                        # class information are obtained by the matrix
                                        if (line+2,i) in matrix:
                                            instructor = matrix[(line+2,i)]

                                        if (line+3,i) in matrix:
                                            match = re.search(r'^(?:.*?[-\s])?([\d\-\,]+)\s*Cr\.?\s*\(CRN\s*([\w#]+)\)\s*(\d+|#+)\s*St\.?\s*$', matrix[(line+3,i)])
                                            if match:
                                                credit = match.group(1).strip()
                                                crn = match.group(2).strip()
                                                cap = match.group(3).strip()
                                            else:
                                                credit = '####'
                                                crn = '####'
                                                cap = '####'

                                        if (line+5,i) in matrix:
                                            note = matrix[(line+5,i)]

                                        if (line+4,i) in matrix:
                                            # considering classes with more than one day of class
                                            if '/' in matrix[(line+4,i)]:
                                                dayhour1, dayhour2 = matrix[(line+4,i)].split('/')

                                                day1, hour1 = dayhour1.split(' ', 1)
                                                day2, hour2 = dayhour2.split(' ', 1)

                                                class_data = {'class': cl,
                                                        'section': section,
                                                        'title': matrix[(line + 1, i)],
                                                        'instructor': instructor,
                                                        'cap': cap,
                                                        'credit': credit,
                                                        'crn': crn,
                                                        'day': day1,
                                                        'hour': hour1,
                                                        'note': note,
                                                        'semesterstart' : semesterstart,
                                                        'semesterend' : semesterend,
                                                        'termcode' : termcode,
                                                        'classstart': termkey[line][0],
                                                        'classend': termkey[line][1]
                                                        }
                                                class_list = data.setdefault(timekey[line], [])
                                                class_list.append(class_data)

                                                class_data = {'class': cl,
                                                        'section': section,
                                                        'title': matrix[(line + 1, i)],
                                                        'instructor': instructor,
                                                        'cap': cap,
                                                        'credit': credit,
                                                        'crn': crn,
                                                        'day': day2,
                                                        'hour': hour2,
                                                        'note': note,
                                                        'semesterstart' : semesterstart,
                                                        'semesterend' : semesterend,
                                                        'termcode' : termcode,
                                                        'classstart': termkey[line][0],
                                                        'classend': termkey[line][1]
                                                        }
                                                class_list = data.setdefault(timekey[line], [])
                                                class_list.append(class_data)

                                            else:
                                                if ': ' in matrix[(line+4,i)]:

                                                    day, hour = matrix[(line+4,i)].split(': ', 1)
                                                else:                                                    
                                                    day, hour = matrix[(line+4,i)].split(' ', 1)

                                                class_data = {'class': cl,
                                                                'section': section,
                                                                'title': matrix[(line + 1, i)],
                                                                'instructor': instructor,
                                                                'cap': cap,
                                                                'credit': credit,
                                                                'crn': crn,
                                                                'day': day,
                                                                'hour': hour,
                                                                'note': note,
                                                                'semesterstart' : semesterstart,
                                                                'semesterend' : semesterend,
                                                                'termcode' : termcode,
                                                                'classstart': termkey[line][0],
                                                                'classend': termkey[line][1]
                                                                }
                                                class_list = data.setdefault(timekey[line], [])
                                                class_list.append(class_data)

                elif mode[line] == 'row':
                    # the information for arranged classes is displayed using rows, therefore, it is used a different algorithm
                    # the code reads information by column, but access and save that information which is stored in rows

                    cls = df1.iloc[line]

                    day = '####'
                    hour = '####'

                    if cls.iloc[0][:4] in department and cls.iloc[0][5].isnumeric(): # considering cells that the name belongs to 
                                                                                     # the department dictionary and the next character is numeric (e.g., NTST500-001)

                        if '-' in cls.iloc[0]:
                            cl, section = cleaner(cls.iloc[0]).strip().split('-')
                        else:
                            cl = cleaner(cls.iloc[0]).strip()
                            section = '####'

                        # checking if the class information was already presented in the sheet
                        for item in data:
                            for newitem in data[item]:
                                if 'class' in newitem.keys():
                                    if newitem['class'] == cl and newitem['section'] == section:
                                        repeatedclass = True
                                        break

                        # saving only information for the classes one time
                        if repeatedclass == False:

                            crn = '####' if cleaner(cls.iloc[2]) == 'NA' else cleaner(cls.iloc[2])

                            title = '####' if cleaner(cls.iloc[3]) == 'NA' else cleaner(cls.iloc[3])

                            note = '####' if cleaner(cls.iloc[5]) == 'NA' else cleaner(cls.iloc[5])

                            credit = '####' if cleaner(cls.iloc[8]) == 'NA' else cleaner(cls.iloc[8])

                            instructor = '####' if cleaner(cls.iloc[9]) == 'NA' else cleaner(cls.iloc[9])

                            cap = '####' if cleaner(cls.iloc[11]) == 'NA' else cleaner(cls.iloc[11])

                            if cleaner(cls.iloc[12]) == 'NA':
                                day = '####'
                            else:
                                if '-' in cleaner(cls.iloc[12]):
                                    day = cleaner(cls.iloc[12])
                                    hour = '####'
                                else:
                                    day = '####'
                                    hour = '####'

                            class_data = {'class': cl,
                                'section': section,
                                'title': title,
                                'instructor': instructor,
                                'cap': cap,
                                'credit': credit,
                                'crn': crn,
                                'day': day,
                                'hour': hour,
                                'note': note,
                                'semesterstart' : semesterstart,
                                'semesterend' : semesterend,
                                'termcode' : termcode
                                }
                            class_list = data.setdefault('Arranged Courses', [])
                            class_list.append(class_data)

                line = line + 1

        initialcolumn = gapcolumn + 1

################################# LAST COLUMN
# the code repeats for the last column

df1 = df.iloc[:, initialcolumn:len(df.columns)]
df1 = df1[6:].reset_index(drop=True)  # droping the header of the dataframe

# defining the column range
column_range = range(initialcolumn, len(df.columns))

## defining when the semester starts and ends
# semester starts
try:
    semesterstart = datetime.strptime(
        str(df.iloc[3, len(df.columns)-1]), '%Y-%m-%d %H:%M:%S'
    ).strftime('%d-%b-%y').upper()
except Exception:
    semesterstart = '####'

# semester ends
try:
    semesterend = datetime.strptime(
        str(df.iloc[4, len(df.columns)-1]), '%Y-%m-%d %H:%M:%S'
    ).strftime('%d-%b-%y').upper()
except Exception:
    semesterend = '####'

code='####'

if semesterstart != '####':
    datestart = semesterstart.split('-')
    if datestart[1] in ['JAN', 'FEB']:
        code = '21' #it means Spring semester
    elif datestart[1] in ['APR', 'MAY']:
        code = '31' #it means Summer semester
    elif datestart[1] in ['AUG', 'SEP']:
        code = '41' #it means Fall semester

    if code != '####':
        termcode = '20' + datestart[2]+ code
    else:
        termcode = '####'
else:
    termcode = '####'

if termcode == chosensemester:

    # mapping the cells in the column range to a matrix
    matrix = {}
    for i in column_range:
        line=0
        for column in df1[i]:
            if column and column not in exclude_time:
                matrix[(line,i)] = cleaner(column)
            line=line+1

    # discovering gap rows to differentiate two algorithm methods
    gaprows = (df1 == 'NA').all(axis=1)
    cut = gaprows & gaprows.shift(-1) # it is considered a cut row when there are two empty rows in the column range

    # saving information of the algorithm method for each column
    mode = {}
    if cut.any():
        cut_row = cut.index[cut].min()
        cut_index_label = cut.index[cut].min()
        cut_iloc = df1.index.get_loc(cut_index_label)

        for i, idx in enumerate(df1.index):
            if i < cut_iloc:
                mode[idx] = 'column'
            else:
                mode[idx] = 'row'
    else:
        cut_row = ''
        mode = {idx: 'column' for idx in df1.index}

    current_time_key = None
    class_count = 0
    timekey = {}
    termkey = {}
    column_old = None
    term_old = None
    rowlines = {}

    for i in column_range:
        line = 0

        for column in df1[i]:
            cl = '####'
            section = '####'
            credit = '####'
            crn = '####'
            cap = '####'
            day = '####'
            hour = '####'
            note = '####'

            column = cleaner(column)

            if mode[line] == 'column':

                print(column)

                if i == 0:

                    # saving information for time frame of the class as indicated in the first column of the column range
                    if column != column_old and column not in exclude_time:
                        timekey[line] = column
                        column_old = column
                    else:
                        timekey[line] = column_old

                    if 'Term' in column:
                        summersemester = True
                        match = re.search(r'(\b[A-Za-z]{3,}\s+\d{1,2}\s*)(?: To | to )(\b[A-Za-z]{3,}\s+\d{1,2}\b)', column)

                        if match:
                            initial_date = datetime.strptime(match.group(1), '%B %d')
                            classstart = initial_date.strftime('%d-%b').upper() + '-' + semesterstart[-2:]

                            final_date = datetime.strptime(match.group(2), '%B %d')
                            classend = final_date.strftime('%d-%b').upper() + '-' + semesterstart[-2:]
                        else:
                            classstart = '####'
                            classend = '####'

                        termkey[line] = [classstart,classend]
                        term_old = [classstart,classend]

                    else:
                        if summersemester == False:
                            termkey[line] = [semesterstart,semesterend]
                        else:
                            termkey[line] = term_old


                if column and column not in exclude_time:

                    # saving cells that start with parentesis as comments (notes)
                    if column[0] == '(':
                        class_data = {'note': column
                                        }
                        class_list = data.setdefault(timekey[line], [])
                        class_list.append(class_data)

                    if len(column) > 4: # considering cells with more text to avoid cells with only the name of the department
                        if column[:4] in department and column[4].isnumeric(): # considering cells that the name belongs to
                                                                               # the department dictionary and the next character is numeric (e.g., NTST500-001)

                            # code for cross-listed classes
                            if '/' in column:
                                cl_dict = {}
                                classes = column.split('/')
                                for item in enumerate(classes):
                                    cl_dict[item[1]] = item[0]

                                crosslist = "/".join(classes)

                                for cls in classes:
                                    match = re.search(r'([A-Z]{3,}\d{3})-(\d{3})', cls)
                                    if match:
                                        cl = match.group(1).strip()
                                        section = match.group(2).strip()
                                    else:
                                        cl = '####'
                                        section = '####'

                                    # checking if the class information was already presented in the sheet
                                    for item in data:
                                        for newitem in data[item]:
                                            if 'class' in newitem.keys():
                                                if newitem['class'] == cl and newitem['section'] == section:
                                                    repeatedclass = True
                                                    break
                                                else:
                                                    repeatedclass = False

                                    # saving only information for the classes one time
                                    if repeatedclass == False:

                                        # class information are obtained by the matrix
                                        if (line+2,i) in matrix:
                                            instructor = matrix[(line+2,i)]

                                        if (line+3+cl_dict[cls],i) in matrix:
                                            match = re.search(r'^(?:.*?[-\s])?([\d\-\,]+)\s*Cr\.?\s*\(CRN\s*([\w#]+)\)\s*(\d+|#+)\s*St\.?\s*$', matrix[(line+3+cl_dict[cls],i)])
                                            if match:
                                                credit = match.group(1).strip()
                                                crn = match.group(2).strip()
                                                cap = match.group(3).strip()
                                            else:
                                                credit = '####'
                                                crn = '####'
                                                cap = '####'

                                        if (line+5+len(cl_dict)-1, i) in matrix:                
                                            note = matrix[(line+5+len(cl_dict)-1, i)]

                                        if (line+4+len(cl_dict)-1,i) in matrix:
                                            # considering classes with more than one day of class
                                            if '/' in matrix[(line+4+len(cl_dict)-1,i)]:
                                                dayhour1, dayhour2 = matrix[(line+4+len(cl_dict)-1,i)].split('/')

                                                day1, hour1 = dayhour1.split(' ', 1)
                                                day2, hour2 = dayhour2.split(' ', 1)

                                                class_data = {'class': cl,
                                                        'section': section,
                                                        'title': matrix[(line + 1, i)],
                                                        'instructor': instructor,
                                                        'cap': cap,
                                                        'credit': credit,
                                                        'crn': crn,
                                                        'day': day1,
                                                        'hour': hour1,
                                                        'note': note,
                                                        'semesterstart' : semesterstart,
                                                        'semesterend' : semesterend,
                                                        'termcode' : termcode,
                                                        'crosslist' : crosslist,
                                                        'classstart': termkey[line][0],
                                                        'classend': termkey[line][1]
                                                        }
                                                class_list = data.setdefault(timekey[line], [])
                                                class_list.append(class_data)

                                                class_data = {'class': cl,
                                                        'section': section,
                                                        'title': matrix[(line + 1, i)],
                                                        'instructor': instructor,
                                                        'cap': cap,
                                                        'credit': credit,
                                                        'crn': crn,
                                                        'day': day2,
                                                        'hour': hour2,
                                                        'note': note,
                                                        'semesterstart' : semesterstart,
                                                        'semesterend' : semesterend,
                                                        'termcode' : termcode,
                                                        'crosslist' : crosslist,
                                                        'classstart': termkey[line][0],
                                                        'classend': termkey[line][1]
                                                        }
                                                class_list = data.setdefault(timekey[line], [])
                                                class_list.append(class_data)


                                            else:
                                                if ': ' in matrix[(line+4+len(cl_dict)-1,i)]:
                                                    day, hour = matrix[(line+4+len(cl_dict)-1,i)].split(': ', 1)
                                                else:
                                                    day, hour = matrix[(line+4+len(cl_dict)-1,i)].split(' ', 1)

                                                class_data = {'class': cl,
                                                            'section': section,
                                                            'title': matrix[(line + 1, i)],
                                                            'instructor': instructor,
                                                            'cap': cap,
                                                            'credit': credit,
                                                            'crn': crn,
                                                            'day': day,
                                                            'hour': hour,
                                                            'note': note,
                                                            'semesterstart' : semesterstart,
                                                            'semesterend' : semesterend,
                                                            'termcode' : termcode,
                                                            'crosslist' : crosslist,
                                                            'classstart': termkey[line][0],
                                                            'classend': termkey[line][1]
                                                            }
                                                class_list = data.setdefault(timekey[line], [])
                                                class_list.append(class_data)

                            else:

                                # checking if the class information was already presented in the sheet
                                match = re.search(r'([A-Z]{3,}\d{3})-(\d{3})', column)
                                if match:
                                    cl, section = match.group(1).strip(), match.group(2).strip()
                                else:
                                    cl, section = '####', '####'

                                repeatedclass = False

                                for item in data:
                                    for newitem in data[item]:
                                        if 'class' in newitem:
                                            if newitem['class'] == cl and newitem['section'] == section:
                                                repeatedclass = True
                                                break

                                    if repeatedclass: 
                                        break

                                # saving only information for the classes one time
                                if repeatedclass == False:

                                    # class information are obtained by the matrix
                                    if (line+2,i) in matrix:
                                        instructor = matrix[(line+2,i)]

                                    if (line+3,i) in matrix:
                                        match = re.search(r'^(?:.*?[-\s])?([\d\-\,]+)\s*Cr\.?\s*\(CRN\s*([\w#]+)\)\s*(\d+|#+)\s*St\.?\s*$', matrix[(line+3,i)])
                                        if match:
                                            credit = match.group(1).strip()
                                            crn = match.group(2).strip()
                                            cap = match.group(3).strip()
                                        else:
                                            credit = '####'
                                            crn = '####'
                                            cap = '####'

                                    if (line+5,i) in matrix:
                                        note = matrix[(line+5,i)]

                                    if (line+4,i) in matrix:
                                        # considering classes with more than one day of class
                                        if '/' in matrix[(line+4,i)]:
                                            dayhour1, dayhour2 = matrix[(line+4,i)].split('/')

                                            day1, hour1 = dayhour1.split(' ', 1)
                                            day2, hour2 = dayhour2.split(' ', 1)

                                            class_data = {'class': cl,
                                                    'section': section,
                                                    'title': matrix[(line + 1, i)],
                                                    'instructor': instructor,
                                                    'cap': cap,
                                                    'credit': credit,
                                                    'crn': crn,
                                                    'day': day1,
                                                    'hour': hour1,
                                                    'note': note,
                                                    'semesterstart' : semesterstart,
                                                    'semesterend' : semesterend,
                                                    'termcode' : termcode,
                                                    'classstart': termkey[line][0],
                                                    'classend': termkey[line][1]
                                                    }
                                            class_list = data.setdefault(timekey[line], [])
                                            class_list.append(class_data)

                                            class_data = {'class': cl,
                                                    'section': section,
                                                    'title': matrix[(line + 1, i)],
                                                    'instructor': instructor,
                                                    'cap': cap,
                                                    'credit': credit,
                                                    'crn': crn,
                                                    'day': day2,
                                                    'hour': hour2,
                                                    'note': note,
                                                    'semesterstart' : semesterstart,
                                                    'semesterend' : semesterend,
                                                    'termcode' : termcode,
                                                    'classstart': termkey[line][0],
                                                    'classend': termkey[line][1]
                                                    }
                                            class_list = data.setdefault(timekey[line], [])
                                            class_list.append(class_data)

                                        else:
                                            if ': ' in matrix[(line+4,i)]:

                                                day, hour = matrix[(line+4,i)].split(': ', 1)
                                            else:
                                                day, hour = matrix[(line+4,i)].split(' ', 1)

                                            class_data = {'class': cl,
                                                            'section': section,
                                                            'title': matrix[(line + 1, i)],
                                                            'instructor': instructor,
                                                            'cap': cap,
                                                            'credit': credit,
                                                            'crn': crn,
                                                            'day': day,
                                                            'hour': hour,
                                                            'note': note,
                                                            'semesterstart' : semesterstart,
                                                            'semesterend' : semesterend,
                                                            'termcode' : termcode,
                                                            'classstart': termkey[line][0],
                                                            'classend': termkey[line][1]
                                                            }
                                            class_list = data.setdefault(timekey[line], [])
                                            class_list.append(class_data)

            elif mode[line] == 'row':
                # the information for arranged classes is displayed using rows, therefore, it is used a different algorithm
                # the code reads information by column, but access and save that information which is stored in rows

                cls = df1.iloc[line]

                day = '####'
                hour = '####'

                if cls.iloc[0][:4] in department and cls.iloc[0][5].isnumeric():
                    if '-' in cls.iloc[0]:
                        cl, section = cleaner(cls.iloc[0]).strip().split('-')
                    else:
                        cl = cleaner(cls.iloc[0]).strip()
                        section = '####'

                    crn = '####' if cleaner(cls.iloc[2]) == 'NA' else cleaner(cls.iloc[2])

                    title = '####' if cleaner(cls.iloc[3]) == 'NA' else cleaner(cls.iloc[3])

                    note = '####' if cleaner(cls.iloc[5]) == 'NA' else cleaner(cls.iloc[5])

                    credit = '####' if cleaner(cls.iloc[8]) == 'NA' else cleaner(cls.iloc[8])

                    instructor = '####' if cleaner(cls.iloc[9]) == 'NA' else cleaner(cls.iloc[9])

                    cap = '####' if cleaner(cls.iloc[11]) == 'NA' else cleaner(cls.iloc[11])

                    if cleaner(cls.iloc[12]) == 'NA':
                        day = '####'
                    else:
                        if '-' in cleaner(cls.iloc[12]):
                            day = cleaner(cls.iloc[12])
                            hour = '####'
                        else:
                            day = '####'
                            hour = '####'

                    class_data = {'class': cl,
                        'section': section,
                        'title': title,
                        'instructor': instructor,
                        'cap': cap,
                        'credit': credit,
                        'crn': crn,
                        'day': day,
                        'hour': hour,
                        'note': note,
                        'semesterstart' : semesterstart,
                        'semesterend' : semesterend,
                        'termcode' : termcode
                        }
                    class_list = data.setdefault('Arranged Courses', [])
                    class_list.append(class_data)

            line = line + 1


# In[10]:


# the third part of the code works with the third sheet (asynchronous classes)
# it follows the same logic of the algorith which reads information from rows


# In[11]:


exclude_time = ['Tues. - 11:30 CHAPEL','Lunch 12:30 - 1:20 PM','M, W, R - Lunch 11:30 - 1:20, T - 12:30-1:30', 'NA'] # excluded cells
department = ['CHIS', 'DSLE', 'GSEM', 'MSSN', 'NTST', 'OTST', 'PATH', 'THST', 'ANEA'] # definition of the department names
room = {1: 'N 108', 2: 'N 110', 3: 'N 135', 4: 'N 150', 5: 'S 115', 6: 'S 120', 7: 'N215', 8: 'N 235', 9: 'N 310', 10: 'N 335', 11: 'S 340'} # definition of rooms

df=all_sheets[2]

gapcolumns = df.columns[(df == 'NA').all()].tolist() # discovering the gap columns (colums in which all the cells are empty)
initialcolumn = 0

for gapcolumn in gapcolumns:

    df1 = df.iloc[:, initialcolumn:gapcolumn]

    # defining the column range
    column_range = range(initialcolumn, gapcolumn)

    ## defining when the semester starts and ends
    # semester starts
    try:
        semesterstart = datetime.strptime(
            str(df.iloc[3, gapcolumn-1]), '%Y-%m-%d %H:%M:%S'
        ).strftime('%d-%b-%y').upper()
    except Exception:
        semesterstart = '####'

    # semester ends
    try:
        semesterend = datetime.strptime(
            str(df.iloc[4, gapcolumn-1]), '%Y-%m-%d %H:%M:%S'
        ).strftime('%d-%b-%y').upper()
    except Exception:
        semesterend = '####'

    code='####'

    if semesterstart != '####':
        datestart = semesterstart.split('-')
        if datestart[1] in ['JAN', 'FEB']:
            code = '21' #it means Spring semester
        elif datestart[1] in ['APR', 'MAY']:
            code = '31' #it means Summer semester
        elif datestart[1] in ['AUG', 'SEP']:
            code = '41' #it means Fall semester

        if code != '####':
            termcode = '20' + datestart[2]+ code
        else:
            termcode = '####'
    else:
        termcode = '####'

    # working only which the chosen semester from the first sheet
    if termcode == chosensemester:

        # droping the header of the dataframe
        df1 = df1[6:].reset_index(drop=True)

        for line in range(0,len(df1.index)):

            cls = df1.iloc[line]

            day = '####'
            hour = '####'

            if cls.iloc[0][:4] in department and cls.iloc[0][5].isnumeric():

                if '-' in cls.iloc[0]:
                    cl, section = cleaner(cls.iloc[0]).strip().split('-')
                else:
                    cl = cleaner(cls.iloc[0]).strip()
                    section = '####'

                crn = '####' if cleaner(cls.iloc[1]) == 'NA' else cleaner(cls.iloc[1])

                title = '####' if cleaner(cls.iloc[2]) == 'NA' else cleaner(cls.iloc[2])

                note = '####' if cleaner(cls.iloc[3]) == 'NA' else cleaner(cls.iloc[3])

                credit = '####' if cleaner(cls.iloc[4]) == 'NA' else cleaner(cls.iloc[4])

                instructor = '####' if cleaner(cls.iloc[5]) == 'NA' else cleaner(cls.iloc[5])

                cap = '####' if cleaner(cls.iloc[6]) == 'NA' else cleaner(cls.iloc[6])

                class_data = {'class': cl,
                    'section': section,
                    'title': title,
                    'instructor': instructor,
                    'cap': cap,
                    'credit': credit,
                    'crn': crn,
                    'note': note,
                    'semesterstart' : semesterstart,
                    'semesterend' : semesterend,
                    'termcode' : termcode
                    }
                class_list = data.setdefault('Arranged Courses', [])
                class_list.append(class_data)

    initialcolumn = gapcolumn + 1


# In[12]:


# the fourth part of the code works cleans the information gathered from the spreadsheet and was saved in the class_data dictionary
# this step is crucial to save gathered information in the desired output name
# if some information is not defined or is not properly defined, it is assigned the place holder value '####' in the output


# In[13]:


room_based_locations = {
    'Wellness Center [ACW]': 'ACW',
    'Bell Hall [BH]': 'BH',
    'Buller Hall [BUL]': 'BUL',
    'Chan Shun Hall [CSH]': 'CSH',
    'Griggs Hall A [GHA]': 'GHA',
    'Griggs Hall B [GHB]': 'GHB',
    'Horn Museum [HORN]': 'HORN',
    'Johnson Gym [JGYM]': 'JGYM',
    'James White Library [JWL]': 'JWL',
    'Nethery Hall [NH]': 'NH',
    'Univ Adventista del Plata [OCARG]': 'OCARG',
    'Centro Univ Adv de Sao Paulo [OCBRA1]': 'OCBRA1',
    'Burman University [OCCABU]': 'OCCABU',
    'Loma Linda University [OCCALL]': 'OCCALL',
    'Southeastern California Conf [OCCASC]': 'OCCASC',
    'Hong Kong Adventist College [OCCHK]': 'OCCHK',
    'Advent Health University [OCFLAH]': 'OCFLAH',
    'Florida Conference of SDA [OCFLFC]': 'OCFLFC',
    'Forest Lake SDA Church [OCFLFL]': 'OCFLFL',
    'Newbold College [OCGBR]': 'OCGBR',
    'North American Division of SDA [OCMDND]': 'OCMDND',
    'Northern New England Conf [OCMENE]': 'OCMENE',
    'Union College [OCNEUC]': 'OCNEUC',
    'Oklahoma City Central [OCOKOC]': 'OCOKOC',
    'Polish Senior College Theo&Hum [OCPOL]': 'OCPOL',
    'Universitatea Adventus din Cer [OCROU]': 'OCROU',
    'Zaokski Theo Seminary [OCRUS]': 'OCRUS',
    'Asia-Pacific Int Univ AIU [OCTHA]': 'OCTHA',
    'Taiwan Adventist College [OCTWN]': 'OCTWN',
    'Ukrainan Adv Center Higher ED [OCUKR]': 'OCUKR',
    'North Pacific Union Conference [OCWANP]': 'OCWANP',
    'Pioneer Memorial Church [PMC]': 'PMC'
}


# In[14]:


# this dictionary works as a standard for the nomenclature used for the day of the week in the sheets
DAY_MAP = {
    'Mon.': 'M', 'Mon': 'M',
    'Tues.': 'T', 'Tue': 'T', 'Tues': 'T',
    'Wed.': 'W', 'Weds': 'W', 'Weds.': 'W', 'Wed': 'W',
    'Thurs': 'R', 'Thurs.': 'R', 'Thur.': 'R', 'Thu.': 'R'
}

# this is the dictionary with entries of Seminary off-campus locations based on class sections
offcampus_locations_class_sections = {'134': 'Advent Health University [OCFLAH]',
                                      '133': 'Asia-Pacific Int Univ AIU [OCTHA]',
                                      '118': 'Burman University [OCCABU]',
                                      '127': 'Centro Univ Adv de Sao Paulo [OCBRA1]',
                                      '112': 'Florida Conference of SDA [OCFLFC]',
                                      '065': 'Forest Lake SDA Church [OCFLFL}',
                                      '095': 'Hong Kong Adventist College [OCCHK]',
                                      '031': 'Seminary Building [SEM]',
                                      '049': 'Loma Linda University [OCCALL]',
                                      '079': 'Newbold College [OCGBR]',
                                      '132': 'North American Division of SDA [OCMDND]',
                                      '111': 'North Pacific Union Conference [OCWANP]',
                                      '038': 'Northern New England Conf [OCMENE]',
                                      '130': 'Oklahoma City Central [OCOKOC]',
                                      '092': 'Polish Senior College Theo&Hum [OCPOL]',
                                      '116': 'Southeastern California Conf [OCCASC]',
                                      '128': 'Taiwan Adventist College [OCTWN]',
                                      '091': 'Ukrainan Adv Center Higher ED [OCUKR]',
                                      '113': 'Union College [OCNEUC]',
                                      '055': 'Universitatea Adventus din Cer [OCROU]',
                                      '069': 'Zaokski Theo Seminary [OCRUS]'
                                      }

room_based_locations = {
    'Wellness Center [ACW]': 'ACW',
    'Bell Hall [BH]': 'BH',
    'Buller Hall [BUL]': 'BUL',
    'Chan Shun Hall [CSH]': 'CSH',
    'Griggs Hall A [GHA]': 'GHA',
    'Griggs Hall B [GHB]': 'GHB',
    'Horn Museum [HORN]': 'HORN',
    'Johnson Gym [JGYM]': 'JGYM',
    'James White Library [JWL]': 'JWL',
    'Nethery Hall [NH]': 'NH',
    'Univ Adventista del Plata [OCARG]': 'OCARG',
    'Centro Univ Adv de Sao Paulo [OCBRA1]': 'OCBRA1',
    'Burman University [OCCABU]': 'OCCABU',
    'Loma Linda University [OCCALL]': 'OCCALL',
    'Southeastern California Conf [OCCASC]': 'OCCASC',
    'Hong Kong Adventist College [OCCHK]': 'OCCHK',
    'Advent Health University [OCFLAH]': 'OCFLAH',
    'Florida Conference of SDA [OCFLFC]': 'OCFLFC',
    'Forest Lake SDA Church [OCFLFL]': 'OCFLFL',
    'Newbold College [OCGBR]': 'OCGBR',
    'North American Division of SDA [OCMDND]': 'OCMDND',
    'Northern New England Conf [OCMENE]': 'OCMENE',
    'Union College [OCNEUC]': 'OCNEUC',
    'Oklahoma City Central [OCOKOC]': 'OCOKOC',
    'Polish Senior College Theo&Hum [OCPOL]': 'OCPOL',
    'Universitatea Adventus din Cer [OCROU]': 'OCROU',
    'Zaokski Theo Seminary [OCRUS]': 'OCRUS',
    'Asia-Pacific Int Univ AIU [OCTHA]': 'OCTHA',
    'Taiwan Adventist College [OCTWN]': 'OCTWN',
    'Ukrainan Adv Center Higher ED [OCUKR]': 'OCUKR',
    'North Pacific Union Conference [OCWANP]': 'OCWANP',
    'Pioneer Memorial Church [PMC]': 'PMC',
    'Seminary Building [SEM]' : ''
}

classes = []
for day_period, sessions in data.items():
    for s in sessions:
        day = ''
        M = ''
        T = ''
        W = ''
        R = ''
        if 'class' in s:
            day = s.get('day', '').strip()

            if 'M' in day:
                M = 'M'

            if 'T' in day:
                T = 'T'

            if 'W' in day:
                W = 'W'

            if 'R' in day:
                R = 'R'

            if '#' in day:
                M = '#'
                T = '#'
                W = '#'
                R = '#'

            if 'hour' in s and len(s.get('hour', '').strip().split('-')) == 2:
                start, end = s.get('hour', '').strip().split('-')
                beginhour = start.replace(':', '').zfill(4)
                endhour = end.replace(':', '').zfill(4)
            else:
                beginhour = '####'
                endhour = '####'
                c = s.get('class', '').strip()
                sec = s.get('section', '').strip()
                #print(f'See hour of class {c}-{sec}')

            if s.get('class', '').strip()[:4] == 'ANEA':
                department = 'OTST'
            else:
                department = s.get('class', '').strip()[:4]

            try:
                date_start = datetime.strptime(s.get('classstart', ''), '%d-%b-%y')
            except ValueError:
                date_start = '####'

            try:
                date_end = datetime.strptime(s.get('classend', ''), '%d-%b-%y')
            except ValueError:
                date_end = '####'

            if date_start != '####' and date_end != '####':
                duration = date_end - date_start
                weeks = math.ceil(duration.days / 7) # calculation and round values for the number of weeks of classes
            else:
                duration = '####'
                weeks = '####'

            if 'room' in s:
                room = s.get('room', '').replace(' ','')
                if room[0] == 'N' or room[0] == 'S':
                    building = 'Seminary Building [SEM]'
                else:
                    section = s.get('section', '').strip()
                    if section in offcampus_locations_class_sections.keys():
                        building = offcampus_locations_class_sections[section]
                        room = room_based_locations[building]
                    else:
                        room = ''
                        building = ''
            else:
                room = ''
                building = ''

            classes.append({
                'Subject': s.get('class', '').strip()[:4],
                'Course Number {Crse Num}': s.get('class', '').strip()[4:],
                'Course Section {Seq Crse Num}': s.get('section', '').strip(),
                'CRN': s.get('crn', '').strip(),
                'Catalog Title': s.get('title', ''),
                'SEM Department {Scacrse Dept}': department,
                'Instructor Name {Instr Name}': s.get('instructor', ''),
                'Credits {Sect Crs}': s.get('credit', '').replace(' ',''),
                'Enrollment Cap {Max Enrl}': s.get('cap', ''),
                'MON': M,
                'TUE': T,
                'WED': W,
                'THU': R,
                'Course Beginning Time {Meet Beg Time}': beginhour,
                'Course Ending Time {Meet End Time}': endhour,
                'Room {Meet Room}': room,
                'Building {Meet Bldg}': building,
                'Semester Start Date {Soaterm Start Date}': s.get('semesterstart', ''),
                'Course Start Date {Meet Start Date}': s.get('semesterstart', ''),
                'Course End Date {Meet End Date}': s.get('semesterend',''),
                'Semester End Date {Soaterm End Date}': s.get('semesterend',''),
                'Crosslist Details {not in Banner}': s.get('crosslist',''),
                'Course Start Date {Meet Start Date}': s.get('classstart', ''),
                'Course End Date {Meet End Date}': s.get('classend',''),
                'Term Code': s.get('termcode',''),
                'Weeks': weeks
            })


# In[15]:


# the fifth and last part of the code saves the information organized by classes in the Excel spreadsheet
# the previous information presented in the input Excel spreadsheet is replaced by the new or updated information gathered in the coode
# it is used the CRN, subject, course number and course section as matching keys


# In[16]:


# this dictonary of group deparments is necessary since ANEA classes belong to the OTST department
groups_department = {
    'GSEM': ['GSEM'], 
    'OTST': ['OTST', 'ANEA'],
    'NTST': ['NTST'], 
    'CHIS': ['CHIS'],
    'THST': ['THST'], 
    'PATH': ['PATH'],
    'DSLE': ['DSLE'], 
    'MSSN': ['MSSN']
}

# course subject and course number for independent studes, topics and directed readings
# these are repeated classes offered regularly that should not appear with red background
exception_classes = {('ANEA', '690'), ('ANEA', '695'), ('ANEA', '880'), ('ANEA', '885'), ('ANEA', '890'), ('ANEA', '895'), ('CHIS', '690'), ('CHIS', '695'), ('CHIS', '696'),
                     ('CHIS', '880'), ('CHIS', '885'), ('CHIS', '890'), ('CHIS', '891'), ('CHIS', '892'), ('CHIS', '895'), ('DSLE', '676'), ('DSLE', '690'), ('DSLE', '885'),
                     ('DSLE', '890'), ('GSEM', '680'), ('GSEM', '685'), ('GSEM', '688'), ('GSEM', '690'), ('GSEM', '695'), ('GSEM', '697'), ('GSEM', '698'), ('GSEM', '700'),
                     ('GSEM', '706'), ('GSEM', '785'), ('GSEM', '788'), ('GSEM', '789'), ('GSEM', '790'), ('GSEM', '793'), ('GSEM', '796'), ('GSEM', '797'), ('GSEM', '810'),
                     ('GSEM', '844'), ('GSEM', '854'), ('GSEM', '880'), ('GSEM', '885'), ('GSEM', '888'), ('GSEM', '898'), ('GSEM', '915'), ('GSEM', '990'), ('GSEM', '995'),
                     ('MSSN', '690'), ('MSSN', '695'), ('MSSN', '795'), ('MSSN', '885'), ('MSSN', '890'), ('MSSN', '897'), ('NTST', '646'), ('NTST', '667'), ('NTST', '690'),
                     ('NTST', '695'), ('NTST', '885'), ('NTST', '890'), ('NTST', '897'), ('OTST', '690'), ('OTST', '695'), ('OTST', '885'), ('OTST', '890'), ('OTST', '895'),
                     ('PATH', '517'), ('PATH', '690'), ('PATH', '715'), ('PATH', '885'), ('PATH', '890'), ('PATH', '897'), ('THST', '690'), ('THST', '695'), ('THST', '885'),
                     ('THST', '890'), ('THST', '895'), ('THST', '896'), ('THST', '897'), ('DSLE', '699'), ('DSLE', '870'), ('DSLE', '878'), ('DSLE', '880'), ('DSLE', '994'),
                     ('DSLE', '995'), ('MSSN', '888'), ('MSSN', '899'), ('PATH', '557')}

# loading our class dataframe
df = pd.DataFrame(classes)

# CRN, subject, course number and course section are matching keys used to replace the informaition
match_keys = ['CRN', 'Subject', 'Course Number {Crse Num}', 'Course Section {Seq Crse Num}']

# loading workbook
wb = load_workbook(Bannerclassesfile_INPUT)
red_font = Font(color='FF0000', bold=True) # it is used red font for the place holder values '####'
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') # it indicates new classes
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid') # it indicates classes presented only in Banner
gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid') # it indicates repeated classes (independent studes, topics and directed readings)

for tab_name, dept_list in groups_department.items():
    if tab_name not in wb.sheetnames:
        continue  # skipping missing sheets

    ws = wb[tab_name]
    data = list(ws.values)
    if not data:
        continue

    # extracting header row
    headers = [str(h).strip() if h else "" for h in data[0]]
    existing_df = pd.DataFrame(data[1:], columns=headers)

    # filtering and making a full copy
    group_df = df[df['SEM Department {Scacrse Dept}'].isin(dept_list)].copy()

    # ensuring string columns safely
    for col in match_keys:
        if col in group_df.columns:
            group_df.loc[:, col] = group_df[col].astype(str)
        if col in existing_df.columns:
            existing_df.loc[:, col] = existing_df[col].astype(str)

    # building an index for existing data for quick lookup
    existing_index = defaultdict(list)

    for idx, row in existing_df.iterrows():
        key = tuple(row[col] for col in match_keys if col in existing_df.columns)
        existing_index[key].append(idx + 2)

    # tracking which rows in the excel sheet were actually updated
    touched_rows = set()

    # iterating over new data
    for _, new_row in group_df.iterrows():
        key = tuple(str(new_row[col]) for col in match_keys)
        row_idxs = existing_index.get(key, []) #considering classes with the same matching key that appears more than once in the spreasheet

        if row_idxs: # existing row are updated if changed
            for row_idx in row_idxs:
                touched_rows.add(row_idx)

                for col_idx, col_name in enumerate(headers, start=1):
                    if col_name not in new_row:
                        continue
                    new_val = new_row[col_name]
                    cell = ws.cell(row=row_idx, column=col_idx)
                    old_val = cell.value
                    if old_val != new_val:
                        cell.value = new_val
                        if str(cell.value) == '#':
                            cell.font = red_font
                        elif str(cell.value) == '##':
                            cell.font = red_font
                        elif str(cell.value) == '###':
                            cell.font = red_font
                        elif str(cell.value) == '####':
                            cell.font = red_font
                        elif str(cell.value) == '#####':
                            cell.font = red_font

        else:
            # new courses are appended as a new row at the end of the data in the repective department sheet
            new_values = [new_row.get(col, "") for col in headers]
            ws.append(new_values)
            new_row_idx = ws.max_row
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=new_row_idx, column=c)
                cell.fill = yellow_fill
                if str(cell.value) == '#':
                    cell.font = red_font
                elif str(cell.value) == '##':
                    cell.font = red_font
                elif str(cell.value) == '###':
                    cell.font = red_font
                elif str(cell.value) == '####':
                    cell.font = red_font
                elif str(cell.value) == '#####':
                    cell.font = red_font

    # working with data that was not changed to apply red or gray background
    subj_col_idx = headers.index('Subject')
    crse_col_idx = headers.index('Course Number {Crse Num}')
    status_col_idx = headers.index('Status')
    section_col_idx = headers.index('Course Section {Seq Crse Num}')
    building_col_idx = headers.index('Building {Meet Bldg}')
    room_col_idx = headers.index('Room {Meet Room}')

    # iterating through all original rows (from 2 up to the count of original data), since data[0] is header, data[1] is row 2
    max_original_row = len(data)

    for r in range(2, max_original_row + 1):

        if r not in touched_rows:
            row_data = data[r-1] 

            current_subj = str(row_data[subj_col_idx])
            current_crse = str(row_data[crse_col_idx])
            current_status = str(row_data[status_col_idx])
            current_section = str(row_data[section_col_idx])
            current_building = str(row_data[building_col_idx])

            if current_status != 'I': # rows with status "I" indicate inactive classes, so it is not necessary to work with them
                if (current_subj, current_crse) in exception_classes or current_section == '006': # gray color to indicate repeated classes (independent studes, topics and directed readings)
                    fill_color = gray_fill
                else:
                    fill_color = red_fill # red color to indicate classes presented only in Banner

                if current_building != 'Seminary Building [SEM]' and current_building in room_based_locations.keys():
                    updatedroom = room_based_locations[current_building]
                else:
                    updatedroom = ''

                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = fill_color
                    ws.cell(row=r, column=room_col_idx+1).value = updatedroom

# save workbook in the output file
wb.save(smartschudulefile_OUTPUT)
print(f"✅ Updated {Bannerclassesfile_INPUT} based on CRN, Subject, Course Number, and Section. Changed cells highlighted in yellow.")


# In[17]:


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string, quote_sheetname
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, date

# ===========================================
# CELL 2: Load workbook + formatting + grouping + autofit
#       + Re-apply dropdown validations (to restore dropdown arrows/UI)
#       + Re-apply Instructor Email/ID formulas for any NEW rows
#       + Force all NON-formula cells to be stored as text (strings)
#       + Option A: also apply Text number_format to EMPTY cells
# ===========================================

# Load the workbook we will WRITE to (single workbook; no cached-values workbook)
wb = load_workbook(smartschudulefile_OUTPUT, data_only=False)

# ------------------------------------------------------------
# Rebuild dropdown ranges from hidden DropDownMenu sheet
# ------------------------------------------------------------
def build_dropdown_ranges_from_dd_sheet(wb, dd_sheet_name="DropDownMenu"):
    if dd_sheet_name not in wb.sheetnames:
        return {}, {}

    ws_dd = wb[dd_sheet_name]
    dropdown_ranges = {}
    dropdown_values = {}

    for col_idx in range(1, ws_dd.max_column + 1):
        header = ws_dd.cell(row=1, column=col_idx).value
        if header is None or str(header).strip() == "":
            continue

        values = []
        for r in range(2, ws_dd.max_row + 1):
            v = ws_dd.cell(row=r, column=col_idx).value
            if v is None or str(v).strip() == "":
                continue
            values.append(str(v))

        if not values:
            continue

        col_letter = get_column_letter(col_idx)
        last_row = 1 + len(values)

        dropdown_ranges[header] = f"{quote_sheetname(dd_sheet_name)}!${col_letter}$2:${col_letter}${last_row}"
        dropdown_values[header] = values

    return dropdown_ranges, dropdown_values


def reapply_dropdown_validations(sheet, dropdown_ranges, dropdown_values, max_rows=5000):
    headers = [cell.value for cell in sheet[1]]

    for header, formula_range in dropdown_ranges.items():
        if header not in headers:
            continue

        cidx = headers.index(header) + 1
        cL = get_column_letter(cidx)

        dv = DataValidation(type="list", formula1=f"={formula_range}", allow_blank=True)
        dv.showDropDown = False  # show dropdown arrow in Excel

        sheet.add_data_validation(dv)
        dv.add(f"${cL}$2:${cL}${max_rows}")

        values_list = dropdown_values.get(header, [])
        if "False" in values_list:
            for r in range(2, min(sheet.max_row, max_rows) + 1):
                cell = sheet[f"{cL}{r}"]
                if cell.value is None or str(cell.value).strip() == "":
                    cell.value = "False"


dropdown_ranges, dropdown_values = build_dropdown_ranges_from_dd_sheet(wb, dd_sheet_name="DropDownMenu")

# ------------------------------------------------------------
# Re-apply dependent autofill formulas (Email/ID) for NEW rows
# ------------------------------------------------------------
instr_name_col  = "Instructor Name {Instr Name}"
instr_email_col = "Instructor Email {Instr Email}"
instr_id_col    = "Instructor ID {Instr ID}"

def reapply_instructor_autofill_formulas(sheet, max_rows=5000, overwrite_existing=False):
    """
    Ensures Instructor Email/ID formulas exist down to sheet.max_row (capped by max_rows).
    Uses InstructorMap sheet lookup (same formula as in Process #1).
    """
    if "InstructorMap" not in wb.sheetnames:
        return

    headers = [cell.value for cell in sheet[1]]
    headers_norm = [str(h).strip() if h is not None else "" for h in headers]

    if instr_name_col not in headers_norm:
        return

    name_col_idx = headers_norm.index(instr_name_col) + 1
    name_L = get_column_letter(name_col_idx)

    max_row = min(sheet.max_row or 1, max_rows)

    # Email formula
    if instr_email_col in headers_norm:
        email_col_idx = headers_norm.index(instr_email_col) + 1
        email_L = get_column_letter(email_col_idx)

        for r in range(2, max_row + 1):
            cell = sheet[f"{email_L}{r}"]
            is_formula = cell.data_type == "f" or (isinstance(cell.value, str) and str(cell.value).startswith("="))
            if is_formula and not overwrite_existing:
                continue

            cell.value = (
                f'=IF(${name_L}{r}="","",'
                f'IFERROR(INDEX(InstructorMap!$B:$B, MATCH(${name_L}{r}, InstructorMap!$A:$A, 0)),""))'
            )

    # ID formula
    if instr_id_col in headers_norm:
        id_col_idx = headers_norm.index(instr_id_col) + 1
        id_L = get_column_letter(id_col_idx)

        for r in range(2, max_row + 1):
            cell = sheet[f"{id_L}{r}"]
            is_formula = cell.data_type == "f" or (isinstance(cell.value, str) and str(cell.value).startswith("="))
            if is_formula and not overwrite_existing:
                continue

            cell.value = (
                f'=IF(${name_L}{r}="","",'
                f'IFERROR(INDEX(InstructorMap!$C:$C, MATCH(${name_L}{r}, InstructorMap!$A:$A, 0)),""))'
            )

# ------------------------------------------------------------
# Force NON-formula cells to TEXT (strings) (rows 2..max_row)
#   Option A: apply Text format to EMPTY cells too
# ------------------------------------------------------------
def force_nonformula_cells_to_text(sheet, start_row=2, max_rows=5000, remove_leading_apostrophe=True):
    """
    Converts ALL non-formula cells to strings and sets number_format to Text ("@").
    ALSO sets number_format to Text ("@") for EMPTY cells, so Excel treats blanks as text-ready.
    Leaves formula cells untouched.
    """
    max_row = min(sheet.max_row or 1, max_rows)
    max_col = sheet.max_column or 1

    for row in sheet.iter_rows(min_row=start_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            v = cell.value

            # Skip formulas (even if they evaluate to blank)
            if cell.data_type == "f" or (isinstance(v, str) and v.startswith("=")):
                continue

            # Ensure Excel stores/behaves as Text even for blanks
            cell.number_format = "@"

            # If blank, keep blank
            if v is None:
                continue

            # Normalize to string
            if isinstance(v, (datetime, date)):
                s = v.isoformat()
            else:
                s = str(v)

            # Remove a visible leading apostrophe created by a previous run
            if remove_leading_apostrophe and isinstance(s, str) and s.startswith("'"):
                s = s[1:]

            cell.value = s
            # IMPORTANT: do NOT set cell.quotePrefix = True

# ------------------------------------------------------------
# Grouping
# ------------------------------------------------------------
GROUP_DEFS = [
    ('Q', 'V'),
    ('X', 'Z'),
    ('AB', 'AH'),
    ('AJ', 'AW'),
    ('AY', 'BA'),
    ('BC', 'CV')
]

def group_columns(sheet, group_defs=GROUP_DEFS, outline_level=1):
    sheet.sheet_properties.outlinePr.summaryRight = False
    sheet.sheet_properties.outlinePr.summaryBelow = True
    sheet.sheet_properties.outlinePr.applyStyles = True

    max_col = sheet.max_column or 1
    for start_col, end_col in group_defs:
        start_idx = column_index_from_string(start_col)
        end_idx   = column_index_from_string(end_col)

        if start_idx > max_col:
            continue

        end_idx = min(end_idx, max_col)
        for col_idx in range(start_idx, end_idx + 1):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].outline_level = outline_level

# ------------------------------------------------------------
# Autofit widths (NO cached values)
# ------------------------------------------------------------
def autofit_columns_ignore_header(sheet, min_width=4, max_width=60, padding=2, max_rows=5000):
    max_row = min(sheet.max_row or 1, max_rows)
    max_col = sheet.max_column or 1

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0

        for r in range(2, max_row + 1):
            v = sheet.cell(row=r, column=col_idx).value
            if v is None:
                continue
            max_length = max(max_length, len(str(v)))

        sheet.column_dimensions[col_letter].width = max(min_width, min(max_length + padding, max_width))

# ------------------------------------------------------------
# Fixed widths (override AFTER autofit)
# ------------------------------------------------------------
FIXED_WIDTHS = {
    "Notes {not in Banner}": 40,
    "Instructor Name {Instr Name}": 20,
    "Instructor Email {Instr Email}": 20,
    "Instructor ID {Instr ID}": 10,
    "account to be charged {not in Banner}": 20,
    "Semester Start Date {Soaterm Start Date}": 10,
    "Pre-work Start Date {not in Banner}": 10,
    "Pre-work End Date {not in Banner}": 10,
    "Course Start Date {Meet Start Date}": 10,
    "Course End Date {Meet End Date}": 10,
    "Post-work Start Date {not in Banner}": 10,
    "Post-work End Date {not in Banner}": 10,
    "Semester End Date {Soaterm End Date}": 10,
}

def apply_fixed_widths(sheet, fixed_widths=FIXED_WIDTHS):
    headers = [cell.value for cell in sheet[1]]
    for header, width in fixed_widths.items():
        if header in headers:
            col_idx = headers.index(header) + 1
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = width

# ------------------------------------------------------------
# Formatting (your original)
# ------------------------------------------------------------
thin_grey_border = Border(
    left=Side(style='thin', color="D3D3D3"),
    right=Side(style='thin', color="D3D3D3"),
    top=Side(style='thin', color="D3D3D3"),
    bottom=Side(style='thin', color="D3D3D3"),
)

COLOR_DATA_CELLS = False

BLUE_HEADERS = [
    "Don {not in Banner}",
    "Mismatching {not in Banner}",
    "Dept/Prog Admin {not in Banner}",
    "Mona {not in Banner}",
    "Karen {not in Banner}",
    "DONE {not in Banner}",
    "Notes {not in Banner}",
    "Subject",
    "Course Number {Crse Num}",
    "Course Section {Seq Crse Num}",
    "Program {not in Banner}",
    "Catalog Title",
    "Section Title",
    "Crosslist Details {not in Banner}",
    "Campus {Ssasect Campus}",
    "Campus Restriction {Camp Restr CC}",
    "Schedule Type {not in Banner}",
    "Instruction Method {Inst Method}",
    "Level {Level SC}",
    "Part-Of-Term",
    "Credits {Sect Crs}",
    "Enrollment Cap {Max Enrl}",
    "Meeting Type",
    "Pre-work Start Date {not in Banner}",
    "Pre-work End Date {not in Banner}",
    "Course Start Date {Meet Start Date}",
    "Course End Date {Meet End Date}",
    "Post-work Start Date {not in Banner}",
    "Post-work End Date {not in Banner}",
    "Course Beginning Time {Meet Beg Time}",
    "Course Ending Time {Meet End Time}",
    "SUN","MON","TUE","WED","THU","FRI","SAT",
    "Room {Meet Room}",
    "Building {Meet Bldg}",
    "Instructor Name {Instr Name}",
    "Instructor Email {Instr Email}",
    "Instructor ID {Instr ID}",
    "% Responsibility",
    "load/contract {not in Banner}",
    "costs per credit {not in Banner}",
    "total costs {not in Banner}",
    "account to be charged {not in Banner}",
]

GREY_HEADERS = [
    "CRN sorted {not in Banner}",
    "CRN",
    "X Lst",
    "Special Approval {Special Apvl}",
    "Credit Catalog {Cat Crs}",
    "Semester Start Date {Soaterm Start Date}",
    "Semester End Date {Soaterm End Date}",
    "Weeks",
    "Fee Amount {Fees Amt}",
    "Fee Code {Detl Code}",
    "Fee Type",
    "Fee Level",
    "Fee Ind",
    "Fee Term",
    "SEM Department {Scacrse Dept}",
    "Activity Date",
    "Status",
    "Bill Hrs",
    "Sect Sch Type",
    "Prerequisites {Preq Areas}",
    "Link Ind",
    "Integration Code",
    "Link Conn",
    "Crs Attr",
    "Comments",
    "Major Restr",
    "Scacrse College",
    "Enrolled",
    "Waitlist Capacity",
    "Primary Ind",
    "Override",
    "1st Date Reg Opens",
    "Last Date Reg Opens",
    "OL Range From Date",
    "OL Range To Date",
    "OL Numb Units",
    "OL Duration Code",
    "Grade Mode",
    "Gradable Ind",
    "Tuit Waiver",
    "Course Override {Meet Override}",
    "Lvl Res Ind",
    "Cmp Res Ind",
    "Rate Code",
    "Cohort Code",
    "St Attr Code",
    "Coll SC",
    "DegC SC",
    "Prog SC",
    "FOS Type",
    "Term Code",
    "Part-Of-Term",
    "Camp SC",
]

HEADER_DEFAULT_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
HEADER_BLUE_FILL    = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
HEADER_GREY_FILL    = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

HEADER_FILL_BY_NAME = {}
for h in BLUE_HEADERS:
    HEADER_FILL_BY_NAME[h] = HEADER_BLUE_FILL
for h in GREY_HEADERS:
    HEADER_FILL_BY_NAME[h] = HEADER_GREY_FILL


def format_sheet(sheet):
    headers = [cell.value for cell in sheet[1]]

    # Freeze panes
    if "Notes {not in Banner}" in headers:
        notes_idx = headers.index("Notes {not in Banner}") + 1
        sheet.freeze_panes = f"{get_column_letter(notes_idx + 1)}2"
    else:
        sheet.freeze_panes = "A2"

    # AutoFilter
    last_col = get_column_letter(sheet.max_column)
    sheet.auto_filter.ref = f"A1:{last_col}{sheet.max_row}"

    # Header style (base)
    for cell in sheet[1]:
        if cell.value is None or str(cell.value).strip() == "":
            continue

        header_text = str(cell.value).strip()
        cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color="FF0000" if header_text == "Notes {not in Banner}" else "000000",
        )
        cell.fill = HEADER_DEFAULT_FILL
        cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True, text_rotation=90)

    sheet.row_dimensions[1].height = 150

    # Apply explicit header color overrides
    for col_idx, header_val in enumerate(headers, start=1):
        if header_val is None:
            continue
        header_text = str(header_val).strip()
        fill = HEADER_FILL_BY_NAME.get(header_text)
        if fill is None:
            continue

        sheet.cell(row=1, column=col_idx).fill = fill

        if COLOR_DATA_CELLS:
            for r in range(2, sheet.max_row + 1):
                sheet.cell(row=r, column=col_idx).fill = fill

    # Grey-out inactive rows (Status == "I")
    if "Status" in headers:
        status_col_idx = headers.index("Status") + 1
        inactive_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for r in range(2, sheet.max_row + 1):
            v = sheet.cell(row=r, column=status_col_idx).value
            if str(v).strip().upper() == "I":
                for c in range(1, sheet.max_column + 1):
                    sheet.cell(row=r, column=c).fill = inactive_fill

    # Borders
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_grey_border

    # Fonts
    # 1) Global default: Arial 10
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            f = cell.font or Font()
            cell.font = Font(
                name="Arial",
                size=10,
                bold=f.bold,
                italic=f.italic,
                underline=f.underline,
                color=f.color,
            )

    # 2) Override columns A–G to Arial 10 (DATA ROWS ONLY)
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=7):
        for cell in row:
            f = cell.font or Font()
            cell.font = Font(
                name="Arial",
                size=10,
                bold=f.bold,
                italic=f.italic,
                underline=f.underline,
                color=f.color,
            )

# ------------------------------------------------------------
# Apply formatting + grouping + autofit + fixed widths + dropdowns
# ------------------------------------------------------------
export_sheets = [
    "CHIS", "DSLE", "GSEM", "MSSN", "NTST", "OTST",
    "PATH", "THST", "MDivHISP_MAPmENGL_MAPmHISP", "MA_Religion", "DMIN",
]

for name in export_sheets:
    ws = wb[name]

    # ✅ 0) Extend formulas for newly-added rows
    reapply_instructor_autofill_formulas(ws, max_rows=5000, overwrite_existing=False)

    # ✅ 1) Force all NON-formula cells to be text (strings) + apply Text format to blanks
    force_nonformula_cells_to_text(ws, start_row=2, max_rows=5000)

    # 2) Normal formatting pipeline
    format_sheet(ws)
    group_columns(ws)
    autofit_columns_ignore_header(ws)
    apply_fixed_widths(ws)
    reapply_dropdown_validations(ws, dropdown_ranges, dropdown_values, max_rows=5000)

wb.save(smartschudulefile_OUTPUT)
wb.close()

print("✅ Cell 2 done: formatting + grouping + autofit + fixed widths + dropdown validations refreshed + formulas extended + non-formulas stored as text.")

